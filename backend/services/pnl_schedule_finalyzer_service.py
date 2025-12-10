"""Service for generating PNL Schedule Finalyzer with detailed line items."""

import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from backend.config.settings import settings
from backend.models.financial_statement import PLScheduleGenerationResponse
from backend.services.path_service import PathService


class PNLScheduleFinalyzerService:
    """Service for generating PNL Schedule Finalyzer with detailed breakdown."""

    ENTITY_PNL_CONFIGS = {
        "default": {
            "income_notes": ["24", "25"],
            "expense_notes": ["26", "27", "28", "29", "30", "31"],
            "tax_notes": ["32"],
        },
    }

    NOTE_TITLES = {
        "24": "Revenue from Operations",
        "25": "Interest Income",  # Will be updated based on note content
        "26": "Purchases of traded goods",
        "27": "Changes in inventories of finished goods, stock-in-trade and work-in-progress",
        "28": "Employee Benefit Expenses",
        "29": "Finance Cost",
        "30": "Depreciation and Amortisation Expense",
        "31": "Other expenses",
        "32": "Tax Expense",
    }

    @staticmethod
    def _get_entity_config(company_name: str) -> Dict[str, List[str]]:
        """Get PNL configuration for a specific entity."""
        return PNLScheduleFinalyzerService.ENTITY_PNL_CONFIGS.get(
            company_name, PNLScheduleFinalyzerService.ENTITY_PNL_CONFIGS["default"]
        )

    @staticmethod
    def _find_latest_note_file(company_name: str, note_number: str) -> Optional[Path]:
        """Find the most recent generated note file."""
        entity_notes_dir = settings.get_entity_generated_notes_dir(company_name)

        if not entity_notes_dir.exists():
            return None

        note_files = list(entity_notes_dir.glob(f"note{note_number}_*.md"))
        if not note_files:
            note_files = list(entity_notes_dir.glob(f"Note_{note_number}_*.md"))

        if not note_files:
            return None

        return max(note_files, key=lambda p: p.stat().st_mtime)

    @staticmethod
    def _parse_note_details(md_content: str, note_number: str) -> Dict:
        """
        Parse markdown note to extract line items with consol codes and amounts.
        
        Returns:
            Dict with 'title', 'line_items' (list of dicts), and 'total'
        """
        result = {
            "title": None,
            "line_items": [],
            "total": None,
        }

        # Extract note title from NOTE XX: header
        title_match = re.search(
            r'\*\*NOTE \d+:\s*([^*\n]+)\*\*',
            md_content,
            re.IGNORECASE
        )
        if title_match:
            result["title"] = title_match.group(1).strip()

        # For tax note (32), parse specifically for Current and Deferred tax
        if note_number == "32":
            # Look for Current tax expense
            current_tax_match = re.search(
                r'\|\s*Current tax(?:\s+expense)?\s*\|\s*₹?([\d,]+(?:\.\d+)?)',
                md_content,
                re.IGNORECASE
            )
            if current_tax_match:
                amount = PNLScheduleFinalyzerService._extract_amount(current_tax_match.group(1))
                result["line_items"].append({
                    "label": "Current tax",
                    "consol_code": "",
                    "amount": amount,
                    "is_total": False
                })

            # Look for Deferred tax expense
            deferred_tax_match = re.search(
                r'\|\s*Deferred tax(?:\s+expense(?:/\(credit\))?|\s+\(expense\))?\s*\|\s*₹?([\d,]+(?:\.\d+)?)',
                md_content,
                re.IGNORECASE
            )
            if deferred_tax_match:
                amount = PNLScheduleFinalyzerService._extract_amount(deferred_tax_match.group(1))
                result["line_items"].append({
                    "label": "Deferred tax",
                    "consol_code": "",
                    "amount": amount,
                    "is_total": False
                })

            # Get total
            total_match = re.search(
                r'\|\s*\*\*TOTAL TAX EXPENSE\*\*\s*\|\s*\*\*₹?([\d,]+(?:\.\d+)?)\*\*',
                md_content,
                re.IGNORECASE
            )
            if total_match:
                result["total"] = PNLScheduleFinalyzerService._extract_amount(total_match.group(1))

            return result

        # For other notes, find the summary table
        note_section_match = re.search(
            r'\*\*NOTE \d+:.*?\*\*.*?\n.*?\| Particulars.*?\n.*?\|.*?---.*?\n(.*?)(?:\n\n|\Z)',
            md_content,
            re.DOTALL | re.IGNORECASE
        )

        if not note_section_match:
            return result

        table_content = note_section_match.group(1)
        lines = table_content.split('\n')

        for line in lines:
            line = line.strip()
            
            if '|' not in line or '---' in line:
                continue

            parts = [p.strip() for p in line.split('|')]
            
            if len(parts) < 3:
                continue

            label = parts[1].strip()
            amount_str = parts[2].strip() if len(parts) > 2 else ""

            if not label or 'Particulars' in label:
                continue

            # Check if TOTAL line
            if 'TOTAL' in label.upper():
                amount = PNLScheduleFinalyzerService._extract_amount(amount_str)
                result["total"] = amount
                continue

            # Check if section header (bold, no amount)
            is_header = (label.startswith('**') and label.endswith('**'))
            
            if is_header:
                label = label.replace('**', '').strip()
                result["line_items"].append({
                    "label": label,
                    "consol_code": "",
                    "amount": None,
                    "is_total": False
                })
            else:
                # Regular line item
                amount = PNLScheduleFinalyzerService._extract_amount(amount_str)
                label = label.replace('**', '').strip()
                
                result["line_items"].append({
                    "label": label,
                    "consol_code": "",  # Empty consol code
                    "amount": amount,
                    "is_total": False
                })

        return result

    @staticmethod
    def _extract_amount(amount_str: str) -> Optional[float]:
        """Extract numeric amount from string."""
        if not amount_str or amount_str == '':
            return None

        amount_str = amount_str.replace('₹', '').replace(',', '').replace('**', '').strip()

        try:
            return float(amount_str)
        except ValueError:
            return None

    @staticmethod
    def _extract_note_schedule_details(
        company_name: str, note_number: str
    ) -> Tuple[str, List[Dict], Optional[float]]:
        """Extract note title, line items, and total from note file."""
        note_file = PNLScheduleFinalyzerService._find_latest_note_file(
            company_name, note_number
        )

        default_title = PNLScheduleFinalyzerService.NOTE_TITLES.get(
            note_number, f"Note {note_number}"
        )

        if not note_file:
            return default_title, [], None

        try:
            with open(note_file, "r", encoding="utf-8") as f:
                content = f.read()

            details = PNLScheduleFinalyzerService._parse_note_details(content, note_number)
            
            title = details["title"] if details["title"] else default_title
            
            return title, details["line_items"], details["total"]

        except Exception as e:
            print(f"Error reading note {note_number}: {e}")
            return default_title, [], None

    @staticmethod
    def _format_amount(value: Optional[float], prefix: str = "", show_prefix: bool = True, convert_to_lakh: bool = False) -> str:
        """Format amount with optional currency prefix and lakh conversion."""
        if value is None:
            return ""
        
        value = abs(value)
        
        # Convert to lakh if requested (divide by 100,000)
        if convert_to_lakh:
            value = value / 100000
        
        formatted = f"{value:,.2f}"
        
        if show_prefix and prefix:
            return f"{prefix} {formatted}"
        else:
            return formatted

    @staticmethod
    def generate_pnl_schedule(
        company_name: str,
        period_label: str = "2025 Mar YTD",
        entity_info: str = "Entity: CPM  Book: 8937,IndAS - IndAS entities,Standalone,MYR,Actual",
        currency: str = "Malaysian Ringgit",
        scenario: str = "Actual",
        show_currency_prefix: bool = True,
        currency_prefix: str = "RM",
        convert_to_lakh: bool = False,
    ) -> PLScheduleGenerationResponse:
        """Generate PNL Schedule Finalyzer with detailed line items."""
        try:
            config = PNLScheduleFinalyzerService._get_entity_config(company_name)

            all_notes = config["income_notes"] + config["expense_notes"] + config["tax_notes"]
            
            note_data = {}
            for note_num in all_notes:
                title, line_items, total = PNLScheduleFinalyzerService._extract_note_schedule_details(
                    company_name, note_num
                )
                note_data[note_num] = {
                    "title": title,
                    "line_items": line_items,
                    "total": total
                }

            output_file = PNLScheduleFinalyzerService._export_to_excel_schedule(
                company_name=company_name,
                period_label=period_label,
                entity_info=entity_info,
                currency=currency,
                scenario=scenario,
                note_data=note_data,
                config=config,
                show_currency_prefix=show_currency_prefix,
                currency_prefix=currency_prefix,
                convert_to_lakh=convert_to_lakh,
            )

            return PLScheduleGenerationResponse(
                success=True,
                message=f"PNL Schedule generated successfully for {company_name}",
                output_file=str(output_file),
                metadata={
                    "generated_at": datetime.now().isoformat(),
                    "notes_included": list(note_data.keys()),
                    "type": "PNL_Schedule",
                },
            )

        except Exception as e:
            return PLScheduleGenerationResponse(
                success=False,
                message=f"Error generating PNL Schedule: {str(e)}"
            )

    @staticmethod
    def _export_to_excel_schedule(
        company_name: str,
        period_label: str,
        entity_info: str,
        currency: str,
        scenario: str,
        note_data: Dict,
        config: Dict[str, List[str]],
        show_currency_prefix: bool = True,
        currency_prefix: str = "RM",
        convert_to_lakh: bool = False,
    ) -> Path:
        """Export PNL Schedule to Excel matching the template."""
        
        path_service = PathService(company_name)
        schedule_dir = (
            path_service.get_financial_statements_dir(company_name) / "PNL_Schedule"
        )
        schedule_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = schedule_dir / f"PNL_Schedule_{timestamp}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "PL Schedule"

        # Define styles
        company_font = Font(name="Calibri", size=11, color="C00000")  # Red
        period_font = Font(name="Calibri", size=11, bold=True)
        entity_font = Font(name="Calibri", size=9)
        header_font = Font(name="Calibri", size=11, bold=True)
        section_font = Font(name="Calibri", size=11, bold=True, color="C00000")  # Red/Orange
        normal_font = Font(name="Calibri", size=11)
        bold_font = Font(name="Calibri", size=11, bold=True)

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        light_gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Set column widths
        ws.column_dimensions["A"].width = 70
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18

        current_row = 1

        # Header section (same as PNL Finalyzer)
        cell = ws.cell(row=current_row, column=1)
        cell.value = company_name
        cell.font = company_font
        current_row += 1

        cell = ws.cell(row=current_row, column=1)
        cell.value = period_label
        cell.font = period_font
        current_row += 1

        current_row += 1  # Empty row

        cell = ws.cell(row=current_row, column=1)
        cell.value = entity_info
        cell.font = entity_font
        current_row += 1

        # PL Schedule header
        cell = ws.cell(row=current_row, column=1)
        cell.value = "PL Schedule"
        cell.font = bold_font
        cell.fill = yellow_fill
        ws.merge_cells(f"A{current_row}:C{current_row}")
        current_row += 1

        current_row += 1  # Empty row

        # Gray header section
        headers_row = current_row
        for col_idx in range(1, 4):
            ws.cell(row=current_row, column=col_idx).fill = light_gray_fill
            ws.cell(row=current_row, column=col_idx).border = thin_border
        ws.cell(row=current_row, column=2).value = "Entity"
        current_row += 1

        # Period Id
        ws.cell(row=current_row, column=2).value = "Period Id"
        ws.cell(row=current_row, column=2).fill = light_gray_fill
        ws.cell(row=current_row, column=2).border = thin_border
        ws.cell(row=current_row, column=3).value = period_label.split()[1].upper() + " (" + period_label.split()[2] + ")"
        ws.cell(row=current_row, column=3).fill = light_gray_fill
        ws.cell(row=current_row, column=3).border = thin_border
        ws.cell(row=current_row, column=3).font = bold_font
        current_row += 1

        # Currency
        ws.cell(row=current_row, column=2).value = "Currency"
        ws.cell(row=current_row, column=2).fill = light_gray_fill
        ws.cell(row=current_row, column=2).border = thin_border
        ws.cell(row=current_row, column=3).value = currency
        ws.cell(row=current_row, column=3).fill = light_gray_fill
        ws.cell(row=current_row, column=3).border = thin_border
        ws.cell(row=current_row, column=3).font = bold_font
        current_row += 1

        # Scenario
        ws.cell(row=current_row, column=2).value = "Scenario"
        ws.cell(row=current_row, column=2).fill = light_gray_fill
        ws.cell(row=current_row, column=2).border = thin_border
        ws.cell(row=current_row, column=3).value = scenario
        ws.cell(row=current_row, column=3).fill = light_gray_fill
        ws.cell(row=current_row, column=3).border = thin_border
        ws.cell(row=current_row, column=3).font = bold_font
        current_row += 1

        # Column headers
        ws.cell(row=current_row, column=1).value = "Item Label / Nature Of Report"
        ws.cell(row=current_row, column=1).fill = light_gray_fill
        ws.cell(row=current_row, column=1).border = thin_border
        ws.cell(row=current_row, column=1).font = header_font

        ws.cell(row=current_row, column=2).value = "Consol Code"
        ws.cell(row=current_row, column=2).fill = light_gray_fill
        ws.cell(row=current_row, column=2).border = thin_border
        ws.cell(row=current_row, column=2).font = header_font

        ws.cell(row=current_row, column=3).value = "Standalone"
        ws.cell(row=current_row, column=3).fill = light_gray_fill
        ws.cell(row=current_row, column=3).border = thin_border
        ws.cell(row=current_row, column=3).font = header_font
        current_row += 1

        # Helper function
        def add_data_row(label: str, consol: str = "", amount: str = "", 
                        is_section: bool = False, is_bold: bool = False):
            nonlocal current_row

            # Column A
            cell = ws.cell(row=current_row, column=1)
            cell.value = label
            cell.border = thin_border
            if is_section:
                cell.font = section_font
            elif is_bold:
                cell.font = bold_font
            else:
                cell.font = normal_font

            # Column B
            cell = ws.cell(row=current_row, column=2)
            cell.value = consol
            cell.border = thin_border
            cell.font = normal_font
            cell.alignment = Alignment(horizontal="left")

            # Column C
            cell = ws.cell(row=current_row, column=3)
            cell.value = amount
            cell.border = thin_border
            if is_bold:
                cell.font = bold_font
            else:
                cell.font = normal_font
            cell.alignment = Alignment(horizontal="right")

            current_row += 1

        # Add "PL Schedule" section marker
        add_data_row("PL Schedule", is_section=True)

        # Add all notes data
        all_notes = config["income_notes"] + config["expense_notes"] + config["tax_notes"]
        
        for note_num in all_notes:
            if note_num not in note_data:
                continue
                
            data = note_data[note_num]
            
            # Section header
            add_data_row(data["title"], is_section=True)
            
            # Add line items
            for item in data["line_items"]:
                consol = item.get("consol_code", "")
                amount_str = PNLScheduleFinalyzerService._format_amount(
                    item["amount"], 
                    prefix=currency_prefix,
                    show_prefix=show_currency_prefix,
                    convert_to_lakh=convert_to_lakh
                ) if item["amount"] else ""
                add_data_row(item["label"], consol, amount_str)
            
            # Total for this note
            if data["total"]:
                total_str = PNLScheduleFinalyzerService._format_amount(
                    data["total"],
                    prefix=currency_prefix,
                    show_prefix=show_currency_prefix,
                    convert_to_lakh=convert_to_lakh
                )
                total_label = f"Total {data['title']}"
                # Special case for note titles that already have descriptive names
                if note_num == "24":
                    total_label = "Total Revenue from operations"
                elif note_num == "25":
                    total_label = "Total Other Income"
                elif note_num == "28":
                    total_label = "Total Employee benefit expense"
                elif note_num == "32":
                    total_label = "Total Tax expense"
                    
                add_data_row(total_label, "", total_str, is_bold=True)

        wb.save(output_file)
        
        return output_file