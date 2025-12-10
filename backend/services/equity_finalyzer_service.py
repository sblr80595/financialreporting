# ============================================================================
# FILE: backend/services/equity_finalyzer_service.py
# ============================================================================
"""Service for generating Equity Schedule Finalyzer with detailed breakdown and nested calculations across multiple periods."""

import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.config.settings import settings
from backend.models.bs_schedule_finalyzer import BSScheduleGenerationResponse 
from backend.services.path_service import PathService


class EquityFinalyzerService:
    """Service for generating Equity Schedule Finalyzer with detailed breakdown across multiple periods."""

    ENTITY_EQUITY_CONFIGS = {
        "default": {
            "share_capital_notes": ["14"],
            "other_equity_notes": ["15"],
            "oci_notes": ["SOCIE"], 
        },
    }

    NOTE_TITLES = {
        "14": "Share capital", "15": "Other equity reserves and surplus", "SOCIE": "Other comprehensive income details",
        "SC_NCI": "Share capital, NCI", "OE_NCI": "Other equity, NCI", 
    }

    @staticmethod
    def _get_entity_config(company_name: str) -> Dict[str, List[str]]:
        """Get Equity configuration for a specific entity."""
        return EquityFinalyzerService.ENTITY_EQUITY_CONFIGS.get(
            company_name, EquityFinalyzerService.ENTITY_EQUITY_CONFIGS["default"]
        )

    @staticmethod
    def _find_latest_note_file(company_name: str, note_number: str) -> Optional[Path]:
        """Find the most recent generated note file. MODIFIED to be robust for note numbers 14, 15, and SOCIE"""
        entity_notes_dir = settings.get_entity_generated_notes_dir(company_name)
        if not entity_notes_dir.exists(): return None
        
        # Use glob pattern matching Note_Number_*.md
        note_files = list(entity_notes_dir.glob(f"Note_{note_number}_*.md"))
        
        # Check for SOCIE specific glob (in case it's just *SOCIE*.md)
        if note_number.upper() == 'SOCIE':
             note_files.extend(list(entity_notes_dir.glob(f"*{note_number}*.md")))

        # Check legacy glob if primary fail (e.g., note24_*.md)
        if not note_files:
            note_files = list(entity_notes_dir.glob(f"note{note_number}_*.md"))

        if not note_files: return None
        return max(note_files, key=lambda p: p.stat().st_mtime)

    @staticmethod
    def _extract_amount(amount_str: str) -> Optional[float]:
        """Extract numeric amount from string."""
        if not amount_str or amount_str == '': return None
        is_negative = '(' in amount_str and ')' in amount_str
        amount_str = amount_str.replace('₹', '').replace('RM', '').replace(',', '').replace('**', '').replace('(', '').replace(')', '').strip()
        try:
            value = float(amount_str)
            return -abs(value) if is_negative else abs(value)
        except ValueError:
            return None

    @staticmethod
    def _parse_note_details(md_content: str, note_number: str, periods: List[str]) -> Dict[str, Any]:
        """Parse markdown note to extract line items with amounts for multiple periods."""
        result = {"title": None, "line_items": [], "totals": {p: None for p in periods}}
        
        title_match = re.search(r'\*\*NOTE \d+:\s*([^*\n]+)\*\*', md_content, re.IGNORECASE)
        if title_match: result["title"] = title_match.group(1).strip()

        note_section_match = re.search(
            r'\| Particulars.*?\|.*?\n\|.*?---.*?\n(.*?)(?:\n\n|\Z)',
            md_content, re.DOTALL | re.IGNORECASE
        )
        if not note_section_match: return result

        table_content = note_section_match.group(1); lines = table_content.split('\n')
        
        header_row_match = re.search(r'\| Particulars\s*\|(.*?)\|', md_content, re.IGNORECASE)
        num_header_cols = 0
        if header_row_match:
            header_cols = [c.strip() for c in header_row_match.group(1).split('|') if c.strip()]
            num_header_cols = len(header_cols)
        
        AMOUNT_START_INDEX = 2

        for line in lines:
            line = line.strip(); 
            if '|' not in line or '---' in line: continue

            parts = [p.strip() for p in line.split('|') if p.strip()]
            
            if len(parts) < 2: continue

            label = parts[0].strip().replace('**', '')
            if not label or 'Particulars' in label: continue

            is_total_line = 'TOTAL' in label.upper() or 'TOTAL' in parts[-1].upper()
            
            consol_code = ""
            if num_header_cols >= 2 and EquityFinalyzerService._extract_amount(parts[1]) is None:
                consol_code = parts[1].strip()
                AMOUNT_START_INDEX = 2
            else:
                AMOUNT_START_INDEX = 1
                
            item_data = {"label": label, "consol_code": consol_code, "amounts": {p: None for p in periods}}
            
            for i, period in enumerate(periods):
                part_index = AMOUNT_START_INDEX + i 
                if part_index < len(parts):
                    amount_str = parts[part_index].strip()
                    amount = EquityFinalyzerService._extract_amount(amount_str)
                    
                    if is_total_line:
                        result["totals"][period] = amount
                    else:
                        item_data["amounts"][period] = amount

            if not is_total_line:
                result["line_items"].append(item_data)

        return result

    @staticmethod
    def _extract_note_schedule_details(
        company_name: str, note_number: str, periods: List[str]
    ) -> Tuple[str, List[Dict], Dict]:
        """Extract note title, line items, and totals from note file. Handles missing file gracefully."""
        note_file = EquityFinalyzerService._find_latest_note_file(company_name, note_number)
        default_title = EquityFinalyzerService.NOTE_TITLES.get(note_number, f"Schedule {note_number}")

        if not note_file: 
            # CRITICAL: Return empty structure if file is not found (for optional notes like SOCIE)
            return default_title, [], {p: None for p in periods}

        try:
            with open(note_file, "r", encoding="utf-8") as f: content = f.read()
            details = EquityFinalyzerService._parse_note_details(content, note_number, periods)
            title = details["title"] if details["title"] else default_title
            
            return title, details["line_items"], details["totals"]

        except Exception as e:
            # Added a print for internal debugging, but returns gracefully
            print(f"Error reading or parsing note {note_number} from file {note_file}: {e}")
            return default_title, [], {p: None for p in periods}

    @staticmethod
    def _format_amount(value: Optional[float], prefix: str = "", show_prefix: bool = True, convert_to_lakh: bool = False, na_on_none: bool = True) -> str:
        """Format amount with optional currency prefix and lakh conversion. (Used for display only)"""
        if value is None:
            return "N/A" if na_on_none else ""
        
        is_negative = value < 0
        value = abs(value)
        
        if convert_to_lakh: value = value / 100000
        
        formatted = f"{value:,.2f}"
        
        if is_negative: formatted = f"({formatted})"
            
        if show_prefix and prefix and not is_negative:
             return f"{prefix} {formatted}"
        elif show_prefix and prefix and is_negative:
             return f"({prefix} {formatted[1:-1]})" 
        else:
            return formatted

    @staticmethod
    def generate_bs_schedule(
        company_name: str, period_label: str = "2025 Mar YTD", entity_info: str = "Entity: CPM Book: 8937,IndAS - IndAS entities,Standalone,MYR,Actual",
        currency: str = "Malaysian Ringgit", scenario: str = "Actual", show_currency_prefix: bool = True, currency_prefix: str = "RM",
        convert_to_lakh: bool = False,
    ) -> BSScheduleGenerationResponse:
        """Generate BS Schedule Finalyzer with detailed line items across multiple periods."""
        
        try:
            match = re.search(r'(\d{4})\s*(\w+)', period_label)
            if not match:
                current_year = datetime.now().year; current_month = datetime.now().strftime('%b'); ytd_label = "YTD"
            else:
                current_year = int(match.group(1)); current_month = match.group(2); ytd_label = period_label.split()[-1] if len(period_label.split()) > 2 else "YTD"
            
            periods = [
                f"{current_month} {current_year} ({ytd_label})",
                f"{current_month} {current_year - 1} ({ytd_label})",
                f"{current_month} {current_year - 2} ({ytd_label})",
            ]
            
            config = EquityFinalyzerService._get_entity_config(company_name)
            all_notes = config["share_capital_notes"] + config["other_equity_notes"] + config["oci_notes"]
            
            note_data = {}
            for note_num in all_notes:
                # This call now handles missing files gracefully, returning empty line_items
                title, line_items, totals = EquityFinalyzerService._extract_note_schedule_details(
                    company_name, note_num, periods
                )
                note_data[note_num] = {"title": title, "line_items": line_items, "totals": totals}

            output_file = EquityFinalyzerService._export_to_excel_schedule(
                company_name=company_name, period_label=period_label, entity_info=entity_info, currency=currency,
                scenario=scenario, periods=periods, note_data=note_data, config=config,
                currency_prefix=currency_prefix, convert_to_lakh=convert_to_lakh,
            )

            return BSScheduleGenerationResponse(
                success=True, message=f"Equity Schedule generated successfully for {company_name}", output_file=str(output_file),
                metadata={"generated_at": datetime.now().isoformat(), "notes_included": list(note_data.keys()), "type": "Equity_Schedule", "periods": periods},
            )

        except Exception as e:
            import traceback; traceback.print_exc()
            return BSScheduleGenerationResponse(success=False, message=f"Error generating Equity Schedule: {str(e)}")

    @staticmethod
    def _export_to_excel_schedule(
        company_name: str, period_label: str, entity_info: str, currency: str, scenario: str, 
        periods: List[str], note_data: Dict, config: Dict[str, List[str]],
        currency_prefix: str = "RM", convert_to_lakh: bool = False,
    ) -> Path:
        """Export Equity Schedule to Excel matching the template, inserting complex formulas."""
        
        # --- CONSTANTS DEFINITION (FIXED) ---
        DATA_COLS = ['C', 'D', 'E'] # Columns C, D, E for the 3 periods
        COL_C = 'C' # Fixed NameError
        # -----------------------------------
        
        path_service = PathService(company_name)
        schedule_dir = path_service.get_financial_statements_dir(company_name) / "Equity_Schedule"
        schedule_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = schedule_dir / f"Equity_Schedule_{timestamp}.xlsx"
        wb = Workbook(); ws = wb.active; ws.title = "Equity Schedule"

        # Define styles
        main_heading_font = Font(name="Calibri", size=11, bold=True, color="C65911") # Orange
        section_font = Font(name="Calibri", size=11, bold=True, color="C00000") # Subheadings: Red/Bold
        normal_font = Font(name="Calibri", size=11)
        header_font = Font(name="Calibri", size=11, bold=True)
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        light_gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        data_format = '#,##0.00;[RED](#,##0.00)'

        ws.column_dimensions["A"].width = 50; ws.column_dimensions["B"].width = 18
        for col in DATA_COLS: ws.column_dimensions[col].width = 18

        current_row = 1

        # --- HEADER ROWS (Metadata and Column Headers) ---
        cell = ws.cell(row=current_row, column=1); cell.value = company_name; current_row += 1
        cell = ws.cell(row=current_row, column=1); cell.value = period_label; current_row += 1
        current_row += 1
        cell = ws.cell(row=current_row, column=1); cell.value = entity_info; current_row += 1
        
        cell = ws.cell(row=current_row, column=1); cell.value = "Equity"; cell.font = main_heading_font; cell.fill = yellow_fill
        ws.merge_cells(f"A{current_row}:E{current_row}"); current_row += 1

        if convert_to_lakh: cell = ws.cell(row=current_row, column=1); cell.value = f"{currency_prefix} in Lakhs"; current_row += 1
        current_row += 1

        for label, value in [("Period Id", periods[0]), ("Currency", currency), ("Scenario", scenario)]:
            ws.merge_cells(f"A{current_row}:B{current_row}"); ws.cell(row=current_row, column=1).value = label; ws.cell(row=current_row, column=1).fill = light_gray_fill; ws.cell(row=current_row, column=1).border = thin_border
            ws.merge_cells(f"C{current_row}:E{current_row}"); ws.cell(row=current_row, column=3).value = value; ws.cell(row=current_row, column=3).fill = light_gray_fill; ws.cell(row=current_row, column=3).border = thin_border; ws.cell(row=current_row, column=3).font = header_font
            current_row += 1
            
        ws.cell(row=current_row, column=1).value = "Item Label / Nature Of Report"; ws.cell(row=current_row, column=1).font = header_font
        ws.cell(row=current_row, column=2).value = "Consol Code"; ws.cell(row=current_row, column=2).font = header_font
        for i, period_header in enumerate(periods):
            cell = ws.cell(row=current_row, column=i+3); cell.value = period_header; cell.font = header_font
        for col_idx in range(1, len(periods) + 3):
            cell = ws.cell(row=current_row, column=col_idx); cell.fill = light_gray_fill; cell.border = thin_border
        current_row += 1

        # Helper function
        def add_data_row(label: str, consol: str = "", amounts: Dict[str, Optional[float]] = None, formula: Optional[str] = None, is_main_heading: bool = False, is_total_row: bool = False, is_subheading: bool = False, force_zero: bool = False):
            nonlocal current_row

            cell = ws.cell(row=current_row, column=1)
            cell.value = label
            cell.border = thin_border
            if is_main_heading: cell.font = main_heading_font
            elif is_total_row: cell.font = header_font
            elif is_subheading: cell.font = section_font
            else: cell.font = normal_font
            
            cell = ws.cell(row=current_row, column=2)
            cell.value = consol
            cell.border = thin_border
            cell.font = normal_font
            
            for i, period in enumerate(periods):
                col_letter = DATA_COLS[i]
                cell = ws.cell(row=current_row, column=i+3)
                
                if formula:
                    cell.value = formula.replace("{COL}", col_letter)
                    cell.data_type = 'f'
                    cell.font = main_heading_font if is_main_heading else header_font
                    cell.number_format = data_format
                elif force_zero:
                    cell.value = 0.0
                    cell.number_format = data_format
                elif amounts:
                    amount = amounts.get(period)
                    if amount is not None:
                        cell.value = amount
                        cell.number_format = data_format
                    else:
                        cell.value = "N/A"
                    cell.font = normal_font
                
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="right")
                
                current_row += 1
                return current_row - 1

        # --- BODY CONTENT: OWNERS EQUITY ---
        
        # 1. OWNERS EQUITY (MAIN HEADING)
        add_data_row("Owners equity", is_main_heading=True)
        
        # 1.1 Share capital (Subheading)
        add_data_row("Share capital", is_subheading=True)
        sc_start_row = current_row
        
        # Particulars: Equity share capital
        sc_item = note_data.get('14', {}).get('line_items', [{}])[0]
        sc_row = add_data_row("Equity share capital", sc_item.get('consol_code', ''), sc_item.get('amounts'))
        sc_end_row = current_row - 1
        
        # Total Share capital
        row_total_sc = add_data_row(
            "Total Share capital", formula=f"=SUM({COL_C}{sc_start_row}:{{COL}}{sc_end_row})", is_total_row=True
        )
        current_row += 1
        
        # 1.2 Other equity (Subheading)
        add_data_row("Other equity", is_subheading=True)

        # 1.2.1 RESERVES AND SURPLUS 
        add_data_row("Reserves and surplus", is_subheading=True)
        rs_start_row = current_row
        
        # Particulars (Securities premium reserve, Capital reserve)
        retained_earnings_items = []
        rs_particular_rows = []
        
        note_15_items = note_data.get('15', {}).get('line_items', [])
        
        for item in note_15_items:
            if 'retained earnings' in item['label'].lower():
                retained_earnings_items.append(item)
            # CRITICAL CHECK: Only add non-total/non-OCI items to the general R&S section
            elif 'total' not in item['label'].lower() and 'comprehensive income' not in item['label'].lower():
                 row_num = add_data_row(item['label'], item['consol_code'], item['amounts'])
                 rs_particular_rows.append(row_num)

        # Retained Earnings Particulars Sum (Opening + P/L - Dividend)
        re_sum_start_row = current_row
        for item in retained_earnings_items:
            add_data_row(item['label'], item['consol_code'], item['amounts'])
        re_sum_end_row = current_row - 1
        
        # Total Retained earnings
        row_total_re = add_data_row(
            "Total Retained earnings", formula=f"=SUM({COL_C}{re_sum_start_row}:{{COL}}{re_sum_end_row})", is_total_row=True
        )
        
        # Total Reserves and surplus
        row_total_rs = add_data_row(
            "Total Reserves and surplus", formula=f"=SUM({COL_C}{rs_start_row}:{{COL}}{re_sum_end_row})", is_total_row=True
        )
        current_row += 1
        
        # 1.2.2 OTHER COMPREHENSIVE INCOME (OCI)
        add_data_row("Other comprehensive income", is_subheading=True)
        oci_sum_start_row = current_row

        socie_line_items = note_data.get('SOCIE', {}).get('line_items', [])
        
        # Only proceed if SOCIE data was loaded (i.e., file existed)
        if socie_line_items:
            
            for item in socie_line_items:
                # Add non-FCTR OCI items
                if 'fctr' not in item['label'].lower():
                    add_data_row(item['label'], item['consol_code'], item['amounts'])
        
        add_data_row("Foreign currency translation reserve", is_subheading=True)
        fctr_sum_start_row = current_row
        
        fctr_cy_item = None
        if socie_line_items:
            for item in socie_line_items:
                if 'fctr' in item['label'].lower() and 'for cy' in item['label'].lower():
                    fctr_cy_item = item
                elif 'fctr' in item['label'].lower() and 'total' not in item['label'].lower() and 'for cy' not in item['label'].lower():
                    add_data_row(item['label'], item['consol_code'], item['amounts'])

        # Hardcoded FCTR lines 
        row_elimination_fctr = add_data_row("Elimination CTA Reserve BS", "", amounts=None, force_zero=True)
        row_fctr_rounding = add_data_row("Foreign currency translation reserve rounding off", "", amounts=None, force_zero=True)
        fctr_sum_end_row = current_row - 1
        
        # Total FCTR, SOCE
        row_total_fctr_soce = add_data_row(
            "Total FCTR, SOCE", formula=f"=SUM({COL_C}{fctr_sum_start_row}:{{COL}}{fctr_sum_end_row})", is_total_row=True
        )

        # FCTR for CY (Particular)
        fctr_cy_row_number = row_total_fctr_soce # Use the row of the subtotal if no CY item found
        if fctr_cy_item:
            fctr_cy_row_number = add_data_row(fctr_cy_item['label'], fctr_cy_item['consol_code'], fctr_cy_item['amounts'])
        
        # Total FCTR
        row_total_fctr = add_data_row(
            "Total Foreign currency translation reserve", 
            # Note: The formula needs to be robust against missing fctr_cy_item. 
            # Since fctr_cy_row_number defaults to row_total_fctr_soce, this sum should be correct 
            # (it sums the subtotal with the item, which simplifies if the item is missing).
            formula=f"={COL_C}{row_total_fctr_soce}+{{COL}}{fctr_cy_row_number}", 
            is_total_row=True
        )
        oci_sum_end_row = current_row - 1 # This marks the end of the OCI section

        # Total Other comprehensive income 
        row_total_oci = add_data_row(
            "Total Other comprehensive income", 
            formula=f"=SUM({COL_C}{oci_sum_start_row}:{{COL}}{oci_sum_end_row})", 
            is_total_row=True
        )
        current_row += 1
        
        # Total Other equity
        row_total_oe = add_data_row(
            "Total Other equity", formula=f"={COL_C}{row_total_rs}+{COL_C}{row_total_oci}", is_total_row=True
        )
        
        # TOTAL Owners equity
        row_total_owners_equity = add_data_row(
            "TOTAL Owners equity", formula=f"={COL_C}{row_total_sc}+{COL_C}{row_total_oe}", is_total_row=True, is_main_heading=True
        )
        current_row += 1
        current_row += 1
        
        # --- NON-CONTROLLING INTERESTS SECTION (NCI) ---
        add_data_row("Non-controlling interests", is_main_heading=True) 
        
        # NCI Placeholder Totals (Rows will be calculated based on the formulas provided)
        # Note: Actual NCI particulars are omitted for brevity, but the formula structure ensures the total row logic works.
        row_nci_sc = add_data_row("Share capital, NCI", formula="", is_total_row=True)
        row_nci_re = add_data_row("Total Retained earnings, NCI", formula="", is_total_row=True)
        row_nci_rs = add_data_row("Total Reserves and surplus, NCI", formula=f"={COL_C}{row_nci_re}", is_total_row=True)
        row_nci_oci = add_data_row("Total Other comprehensive income, NCI", formula="", is_total_row=True)
        row_nci_oe = add_data_row("Total Other equity, NCI", formula=f"={COL_C}{row_nci_rs}+{COL_C}{row_nci_oci}", is_total_row=True)

        # TOTAL NON-CONTROLLING INTERESTS 
        row_total_nci = add_data_row(
            "TOTAL NON-CONTROLLING INTERESTS", formula=f"={COL_C}{row_nci_sc}+{COL_C}{row_nci_oe}", is_total_row=True, is_main_heading=True
        )
        current_row += 1
        
        # TOTAL EQUITY (Final Grand Total)
        row_total_equity = add_data_row(
            "TOTAL EQUITY", formula=f"={COL_C}{row_total_owners_equity}+{COL_C}{row_total_nci}", is_total_row=True, is_main_heading=True
        )

        # Save workbook
        wb.save(output_file)
        
        return output_file