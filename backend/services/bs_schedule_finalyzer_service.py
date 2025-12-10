# ============================================================================
# FILE: backend/services/bs_schedule_finalyzer_service.py
# ============================================================================
"""Service for generating BS Schedule Finalyzer with detailed line items from Balance Sheet notes across multiple periods."""

import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.config.settings import settings
# NOTE: Ensure BSScheduleGenerationResponse is now correctly integrated into
# backend.models.balance_sheet_models as per the last step.
from backend.models.bs_schedule_finalyzer import BSScheduleGenerationResponse
from backend.services.path_service import PathService


class BSScheduleFinalyzerService:
    """Service for generating BS Schedule Finalyzer with detailed breakdown across multiple periods."""

    ENTITY_BS_CONFIGS = {
        "default": {
            "asset_notes": ["3", "4", "5", "7", "8", "10", "11", "12", "13"],
            "liability_notes": ["6", "17", "18", "19", "20", "21", "22", "23", "14", "15"], 
        },
    }

    NOTE_TITLES = {
        "3": "Property, plant and equipment", "4": "Right to use assets",
        "5": "Other intangible assets", "6": "Deferred tax liabilities (net)",
        "7": "Other financial assets", "8": "Inventories",
        "10": "Cash and cash equivalents", "11": "Bank balances other than cash and cash equivalents",
        "12": "Loans", "13": "Other financial assets", "13": "Other financial assets", "14": "Equity share capital",
        "15": "Other equity", "17": "Borrowings", "18": "Lease liabilities",
        "19": "Trade payables", "20": "Other financial liabilities",
        "21": "Other current liabilities", "22": "Provisions",
        "23": "Current tax liabilities (net)",
    }

    @staticmethod
    def _get_entity_config(company_name: str) -> Dict[str, List[str]]:
        """Get BS configuration for a specific entity."""
        return BSScheduleFinalyzerService.ENTITY_BS_CONFIGS.get(
            company_name, BSScheduleFinalyzerService.ENTITY_BS_CONFIGS["default"]
        )

    @staticmethod
    def _find_latest_note_file(company_name: str, note_number: str) -> Optional[Path]:
        """Find the most recent generated note file."""
        entity_notes_dir = settings.get_entity_generated_notes_dir(company_name)
        if not entity_notes_dir.exists(): return None
        note_files = list(entity_notes_dir.glob(f"note{note_number}_*.md"))
        if not note_files:
            note_files = list(entity_notes_dir.glob(f"Note_{note_number}_*.md"))
        if not note_files: return None
        return max(note_files, key=lambda p: p.stat().st_mtime)

    @staticmethod
    def _extract_amount(amount_str: str) -> Optional[float]:
        """Extract numeric amount from string."""
        if not amount_str or amount_str == '': return None
        is_negative = '(' in amount_str and ')' in amount_str
        amount_str = amount_str.replace('₹', '').replace('RM', '').replace(',', '').replace('**', '').replace('(', '').replace(')', '').strip()
        try:
            value = float(amount_str)
            return -abs(value) if is_negative else abs(value)
        except ValueError:
            return None

    @staticmethod
    def _parse_note_details(md_content: str, note_number: str, periods: List[str]) -> Dict[str, Any]:
        """
        Parse markdown note to extract line items with amounts for multiple periods.
        (Logic remains as provided by user for multi-column extraction)
        """
        result = {"title": None, "line_items": [], "totals": {p: None for p in periods}}
        
        title_match = re.search(r'\*\*NOTE \d+:\s*([^*\n]+)\*\*', md_content, re.IGNORECASE)
        if title_match: result["title"] = title_match.group(1).strip()

        note_section_match = re.search(
            r'\| Particulars.*?\|.*?\n\|.*?---.*?\n(.*?)(?:\n\n|\Z)',
            md_content, re.DOTALL | re.IGNORECASE
        )
        if not note_section_match: return result

        table_content = note_section_match.group(1); lines = table_content.split('\n')
        
        # Determine table structure 
        header_row_match = re.search(r'\| Particulars\s*\|(.*?)\|', md_content, re.IGNORECASE)
        num_header_cols = 0
        if header_row_match:
            header_cols = [c.strip() for c in header_row_match.group(1).split('|') if c.strip()]
            num_header_cols = len(header_cols)
        
        AMOUNT_START_INDEX = 2

        for line in lines:
            line = line.strip(); 
            if '|' not in line or '---' in line: continue

            parts = [p.strip() for p in line.split('|') if p.strip()]
            
            if len(parts) < 2: continue

            label = parts[0].strip().replace('**', '')
            if not label or 'Particulars' in label: continue

            is_total_line = 'TOTAL' in label.upper() or 'TOTAL' in parts[-1].upper()
            
            # 1. Identify Consol Code 
            consol_code = ""
            if num_header_cols >= 2 and BSScheduleFinalyzerService._extract_amount(parts[1]) is None:
                 consol_code = parts[1].strip()
                 AMOUNT_START_INDEX = 2
            else:
                 AMOUNT_START_INDEX = 1
                 
            item_data = {"label": label, "consol_code": consol_code, "amounts": {p: None for p in periods}}
            
            # 2. Extract Amounts for each period
            for i, period in enumerate(periods):
                part_index = AMOUNT_START_INDEX + i 
                
                if part_index < len(parts):
                    amount_str = parts[part_index].strip()
                    amount = BSScheduleFinalyzerService._extract_amount(amount_str)
                    
                    if is_total_line:
                        result["totals"][period] = amount
                    else:
                        item_data["amounts"][period] = amount

            if not is_total_line:
                result["line_items"].append(item_data)

        return result

    @staticmethod
    def _extract_note_schedule_details(
        company_name: str, note_number: str, periods: List[str]
    ) -> Tuple[str, List[Dict], Dict]:
        """Extract note title, line items, and totals from note file."""
        note_file = BSScheduleFinalyzerService._find_latest_note_file(company_name, note_number)
        default_title = BSScheduleFinalyzerService.NOTE_TITLES.get(note_number, f"Schedule {note_number}")

        if not note_file: 
            return default_title, [], {p: None for p in periods}

        try:
            with open(note_file, "r", encoding="utf-8") as f: content = f.read()
            details = BSScheduleFinalyzerService._parse_note_details(content, note_number, periods)
            title = details["title"] if details["title"] else default_title
            
            return title, details["line_items"], details["totals"]

        except Exception as e:
            print(f"Error reading note {note_number}: {e}")
            return default_title, [], {p: None for p in periods}

    @staticmethod
    def _format_amount(value: Optional[float], prefix: str = "", show_prefix: bool = True, convert_to_lakh: bool = False, na_on_none: bool = True) -> str:
        """Format amount with optional currency prefix and lakh conversion. (Used for display only)"""
        if value is None:
            return "N/A" if na_on_none else ""
        
        is_negative = value < 0
        value = abs(value)
        
        if convert_to_lakh: value = value / 100000
        
        formatted = f"{value:,.2f}"
        
        if is_negative: formatted = f"({formatted})" # Negative in parentheses
            
        if show_prefix and prefix and not is_negative:
             return f"{prefix} {formatted}"
        elif show_prefix and prefix and is_negative:
             # Prefix inside parentheses for negative amounts
             return f"({prefix} {formatted[1:-1]})" 
        else:
            return formatted

    @staticmethod
    def generate_bs_schedule(
        company_name: str,
        period_label: str = "2025 Mar YTD",
        entity_info: str = "Entity: CPM  Book: 8937,IndAS - IndAS entities,Standalone,MYR,Actual",
        currency: str = "Malaysian Ringgit",
        scenario: str = "Actual",
        # RESOLUTION: Argument added here to match the Pydantic model/FastAPI route call
        show_currency_prefix: bool = True, 
        currency_prefix: str = "RM",
        convert_to_lakh: bool = False,
    ) -> BSScheduleGenerationResponse:
        """Generate BS Schedule Finalyzer with detailed line items across multiple periods."""
        
        try:
            match = re.search(r'(\d{4})\s*(\w+)', period_label)
            if not match:
                current_year = datetime.now().year 
                current_month = datetime.now().strftime('%b')
                ytd_label = "YTD"
            else:
                current_year = int(match.group(1))
                current_month = match.group(2)
                ytd_label = period_label.split()[-1] if len(period_label.split()) > 2 else "YTD"
            
            periods = [
                f"{current_month} {current_year} ({ytd_label})",
                f"{current_month} {current_year - 1} ({ytd_label})",
                f"{current_month} {current_year - 2} ({ytd_label})",
            ]
            
            config = BSScheduleFinalyzerService._get_entity_config(company_name)
            all_notes = config["asset_notes"] + config["liability_notes"]
            
            note_data = {}
            for note_num in all_notes:
                title, line_items, totals = BSScheduleFinalyzerService._extract_note_schedule_details(
                    company_name, note_num, periods
                )
                note_data[note_num] = {
                    "title": title,
                    "line_items": line_items,
                    "totals": totals
                }

            output_file = BSScheduleFinalyzerService._export_to_excel_schedule(
                company_name=company_name,
                period_label=period_label,
                entity_info=entity_info,
                currency=currency,
                scenario=scenario,
                periods=periods,
                note_data=note_data,
                config=config,
                currency_prefix=currency_prefix,
                convert_to_lakh=convert_to_lakh,
                # show_currency_prefix is omitted here as _export_to_excel_schedule signature does not expect it.
            )

            return BSScheduleGenerationResponse(
                success=True,
                message=f"BS Schedule generated successfully for {company_name}",
                output_file=str(output_file),
                metadata={
                    "generated_at": datetime.now().isoformat(),
                    "notes_included": list(note_data.keys()),
                    "type": "BS_Schedule_MultiPeriod",
                    "periods": periods,
                },
            )

        except Exception as e:
            # Print traceback for debugging the 500 error
            import traceback
            traceback.print_exc()
            return BSScheduleGenerationResponse(
                success=False,
                message=f"Error generating BS Schedule: {str(e)}"
            )

    @staticmethod
    def _export_to_excel_schedule(
        company_name: str, period_label: str, entity_info: str, currency: str,
        scenario: str, periods: List[str], note_data: Dict, config: Dict[str, List[str]],
        currency_prefix: str = "RM", convert_to_lakh: bool = False,
    ) -> Path:
        """Export BS Schedule to Excel matching the template, supporting multi-period columns."""
        
        path_service = PathService(company_name)
        schedule_dir = (
            path_service.get_financial_statements_dir(company_name) / "BS_Schedule"
        )
        schedule_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = schedule_dir / f"BS_Schedule_{timestamp}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "BS Schedule"

        # Define styles
        main_heading_font = Font(name="Calibri", size=11, bold=True, color="C65911") # Orange
        section_font = Font(name="Calibri", size=11, bold=True) 
        normal_font = Font(name="Calibri", size=11)
        header_font = Font(name="Calibri", size=11, bold=True)
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        light_gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        DATA_COLS = ['C', 'D', 'E'] # Fixed columns for 3 periods
        
        # Set column widths
        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 18
        for col in DATA_COLS: ws.column_dimensions[col].width = 18

        current_row = 1

        # --- Header Section (Metadata) ---
        cell = ws.cell(row=current_row, column=1); cell.value = company_name; current_row += 1
        cell = ws.cell(row=current_row, column=1); cell.value = period_label; current_row += 1
        current_row += 1
        cell = ws.cell(row=current_row, column=1); cell.value = entity_info; current_row += 1
        cell = ws.cell(row=current_row, column=1); cell.value = "BS Schedule"; cell.fill = yellow_fill; current_row += 1

        # Unit Info
        if convert_to_lakh:
            cell = ws.cell(row=current_row, column=1); cell.value = f"{currency_prefix} in Lakhs"; current_row += 1
            
        current_row += 1 # Empty row

        # Gray header section (Period Id, Currency, Scenario)
        for label, value in [("Period Id", periods[0]), ("Currency", currency), ("Scenario", scenario)]:
            ws.merge_cells(f"A{current_row}:B{current_row}")
            ws.cell(row=current_row, column=1).value = label
            ws.cell(row=current_row, column=1).fill = light_gray_fill
            ws.cell(row=current_row, column=1).border = thin_border
            
            # Merge C, D, E for the single value
            ws.merge_cells(f"C{current_row}:E{current_row}")
            ws.cell(row=current_row, column=3).value = value
            ws.cell(row=current_row, column=3).fill = light_gray_fill
            ws.cell(row=current_row, column=3).border = thin_border
            ws.cell(row=current_row, column=3).font = header_font
            current_row += 1
            
        # Column headers (Item Label, Consol Code, P1, P2, P3)
        ws.cell(row=current_row, column=1).value = "Item Label / Nature Of Report"; ws.cell(row=current_row, column=1).font = header_font
        ws.cell(row=current_row, column=2).value = "Consol Code"; ws.cell(row=current_row, column=2).font = header_font

        # Period Headers (C, D, E)
        for i, period_header in enumerate(periods):
            col_letter = DATA_COLS[i]
            cell = ws.cell(row=current_row, column=i+3)
            cell.value = period_header
            cell.font = header_font
        
        for col_idx in range(1, len(periods) + 3):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.fill = light_gray_fill
            cell.border = thin_border
            
        current_row += 1

        # Helper function
        def add_data_row(label: str, consol: str = "", amounts: Dict[str, Optional[float]] = None, formula: Optional[str] = None, is_main_heading: bool = False, is_total_row: bool = False):
            nonlocal current_row

            # Column A (Label)
            cell = ws.cell(row=current_row, column=1)
            cell.value = label
            cell.border = thin_border
            cell.font = main_heading_font if is_main_heading else (header_font if is_total_row else normal_font)
            
            # Column B (Consol Code)
            cell = ws.cell(row=current_row, column=2)
            cell.value = consol
            cell.border = thin_border
            cell.font = normal_font
            cell.alignment = Alignment(horizontal="left")

            # Columns C, D, E (Amounts/Formulas)
            for i, period in enumerate(periods):
                col_letter = DATA_COLS[i]
                cell = ws.cell(row=current_row, column=i+3)
                
                if formula:
                    cell.value = formula.replace("{COL}", col_letter) # Insert formula (e.g., =SUM(C:C))
                    cell.data_type = 'f'
                    cell.font = header_font if is_total_row else normal_font
                elif amounts:
                    # Insert raw float or N/A
                    amount = amounts.get(period)
                    if amount is not None:
                         cell.value = amount
                         cell.number_format = '#,##0.00;[RED](#,##0.00)' 
                    else:
                         cell.value = "N/A" # Display N/A if value is not found
                    cell.font = header_font if is_total_row else normal_font
                
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="right")
            
            current_row += 1
            return current_row - 1

        # --- Body Content: Schedules ---
        all_notes = config["asset_notes"] + config["liability_notes"]
        total_rows = []
        
        # Iterate through notes and format the data
        for note_num in all_notes:
            if note_num not in note_data: continue
            
            data = note_data[note_num]
            
            # --- Schedule Header (Orange Main Heading) ---
            add_data_row(f"Schedule {note_num}", is_main_heading=True)
            row_start = current_row
            
            # Note Title (Particulars)
            add_data_row(data["title"], is_main_heading=False)
            
            # Add line items
            for item in data["line_items"]:
                add_data_row(
                    label=item["label"],
                    consol=item.get("consol_code", ""),
                    amounts=item["amounts"],
                )
            
            row_end = current_row - 1
            
            # --- Total for this Note ---
            total_label = f"Total {BSScheduleFinalyzerService.NOTE_TITLES.get(note_num, f'Note {note_num}')}"
            
            total_row_number = add_data_row(
                label=total_label,
                formula=f"=SUM({{COL}}{row_start}:{{COL}}{row_end})", # Formula template
                is_total_row=True
            )
            total_rows.append(total_row_number)
            
            current_row += 1 # Empty row after total


        # --- Final Totals Section ---
        add_data_row("GRAND SCHEDULE TOTALS", is_main_heading=True)
        
        # Final Summation Row (Sums all note totals)
        final_sum_formula = "+".join([f"{{COL}}{r}" for r in total_rows])
        add_data_row(
            label="Grand Total of All Schedules",
            formula=f"={final_sum_formula}",
            is_total_row=True
        )
        
        # Save workbook
        wb.save(output_file)
        
        return output_file