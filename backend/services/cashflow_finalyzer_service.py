# # ============================================================================
# # FILE: backend/services/cashflow_finalyzer_service.py
# # ============================================================================
# """Cash Flow Statement Finalyzer service - generates Excel matching template."""

# import re
# from datetime import datetime
# from pathlib import Path
# from typing import Dict, List, Optional, Tuple

# from openpyxl import Workbook
# from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# from backend.config.settings import settings
# from backend.services.path_service import PathService


# class CashFlowFinalyzerService:
#     """Service for generating Cash Flow Statement Finalyzer matching exact Excel template."""

#     @staticmethod
#     def _find_cashflow_note_file(company_name: str) -> Optional[Path]:
#         """Find the Cash Flow Statement markdown file."""
#         entity_notes_dir = settings.get_entity_generated_notes_dir(company_name)

#         if not entity_notes_dir.exists():
#             return None

#         patterns = [
#             "Note_CASHFLOW_*.md",
#             "noteCASHFLOW_*.md",
#             "cashflow_*.md",
#             "cash_flow_*.md"
#         ]

#         for pattern in patterns:
#             note_files = list(entity_notes_dir.glob(pattern))
#             if note_files:
#                 return max(note_files, key=lambda p: p.stat().st_mtime)

#         return None

#     @staticmethod
#     def _parse_amount(amount_str: str) -> Optional[float]:
#         """Parse amount string to float."""
#         if not amount_str or amount_str.strip() == '' or amount_str == '-':
#             return None
        
#         # Remove currency symbols, commas, and spaces
#         amount_str = amount_str.replace('₹', '').replace(',', '').strip()
        
#         # Check if bracketed (negative)
#         is_negative = amount_str.startswith('(') and amount_str.endswith(')')
        
#         if is_negative:
#             amount_str = amount_str[1:-1]
        
#         try:
#             amount = float(amount_str)
#             return -amount if is_negative else amount
#         except (ValueError, TypeError):
#             return None

#     @staticmethod
#     def _extract_cashflow_items(md_content: str) -> Dict[str, float]:
#         """Extract all cash flow line items from markdown content."""
#         items = {}
        
#         # Pattern to match table rows: | Description | Amount |
#         table_pattern = r'\|\s*([^|]+?)\s*\|\s*([^|]*)\s*\|'
        
#         for match in re.finditer(table_pattern, md_content):
#             description = match.group(1).strip()
#             amount_str = match.group(2).strip()
            
#             # Skip headers and separators
#             if ('Particulars' in description or 'Amount' in description or 
#                 '---' in description or description == '' or
#                 'GL Code' in description or 'GL Description' in description or
#                 'Ind AS Minor' in description):
#                 continue
            
#             # Clean description - remove markdown formatting
#             clean_desc = description.replace('**', '').strip()
            
#             # Skip section headers that are purely descriptive
#             if clean_desc in ['Cash flow from operating activities', 
#                             'Cash flows from investing activities',
#                             'Cash flows from financing activities',
#                             'Adjustments for:',
#                             'Changes in working Capital',
#                             'Cash and cash equivalents include (refer note 13)',
#                             'Non- cash financing and investing activities:',
#                             'Balance with banks']:
#                 continue
            
#             # Parse amount
#             amount = CashFlowFinalyzerService._parse_amount(amount_str)
            
#             if amount is not None:
#                 # Store with cleaned description as key
#                 items[clean_desc] = amount
                
#                 # Also store common variations for better matching
#                 # Store simplified version without parentheses
#                 simplified = re.sub(r'\([^)]*\)', '', clean_desc).strip()
#                 if simplified != clean_desc:
#                     items[simplified] = amount
        
#         return items

#     @staticmethod
#     def _format_amount(value: Optional[float], decimal_places: int = 2) -> str:
#         """Format amount with decimal places."""
#         if value is None:
#             return ""
#         return f"{value:,.{decimal_places}f}"

#     @staticmethod
#     def generate_cashflow_finalyzer(
#         company_name: str,
#         period_label: str = "2025 Mar YTD",
#         entity_info: str = "Entity: CPM  Book: 8937,IndAS - IndAS entities,Standalone,MYR,Actual",
#         currency: str = "Malaysian Ringgit",
#         scenario: str = "Cashflow",
#     ) -> Dict:
#         """
#         Generate Cash Flow Statement Finalyzer matching exact Excel template.

#         Args:
#             company_name: Name of the company
#             period_label: Period label (e.g., "2025 Mar YTD")
#             entity_info: Entity information line
#             currency: Currency name
#             scenario: Scenario type

#         Returns:
#             Dictionary with success status and file path
#         """
#         try:
#             # Find cash flow markdown file
#             cashflow_file = CashFlowFinalyzerService._find_cashflow_note_file(company_name)
            
#             if not cashflow_file:
#                 return {
#                     "success": False,
#                     "message": "Cash Flow Statement markdown file not found. Please generate it first."
#                 }

#             # Read markdown content
#             with open(cashflow_file, 'r', encoding='utf-8') as f:
#                 md_content = f.read()

#             # Extract all line items
#             cashflow_items = CashFlowFinalyzerService._extract_cashflow_items(md_content)

#             # Generate Excel file
#             output_file = CashFlowFinalyzerService._export_to_excel_finalyzer(
#                 company_name=company_name,
#                 period_label=period_label,
#                 entity_info=entity_info,
#                 currency=currency,
#                 scenario=scenario,
#                 cashflow_items=cashflow_items,
#             )

#             return {
#                 "success": True,
#                 "message": f"Cash Flow Finalyzer generated successfully for {company_name}",
#                 "output_file": str(output_file),
#                 "company_name": company_name,
#                 "period_label": period_label,
#                 "items_extracted": len(cashflow_items)
#             }

#         except Exception as e:
#             return {
#                 "success": False,
#                 "message": f"Error generating Cash Flow Finalyzer: {str(e)}"
#             }

#     @staticmethod
#     def _export_to_excel_finalyzer(
#         company_name: str,
#         period_label: str,
#         entity_info: str,
#         currency: str,
#         scenario: str,
#         cashflow_items: Dict[str, float],
#     ) -> Path:
#         """Export Cash Flow Finalyzer to Excel with exact template formatting."""
        
#         # Create output directory
#         path_service = PathService(company_name)
#         cashflow_dir = path_service.get_financial_statements_dir(company_name) / "CashFlow_Finalyzer"
#         cashflow_dir.mkdir(parents=True, exist_ok=True)

#         timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
#         output_file = cashflow_dir / f"CashFlow_Finalyzer_{timestamp}.xlsx"

#         # Create workbook
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Cash Flow Statement"

#         # Define styles
#         # Fonts
#         company_font = Font(name="Calibri", size=11, color="C00000")  # Red
#         period_font = Font(name="Calibri", size=11, bold=True)
#         entity_font = Font(name="Calibri", size=9)
#         header_font = Font(name="Calibri", size=11, bold=True)
#         section_header_font = Font(name="Calibri", size=11, bold=True, color="C00000")  # Red
#         normal_font = Font(name="Calibri", size=11)
#         bold_font = Font(name="Calibri", size=11, bold=True)

#         # Fills
#         yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
#         light_gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

#         # Borders
#         thin_border = Border(
#             left=Side(style="thin"),
#             right=Side(style="thin"),
#             top=Side(style="thin"),
#             bottom=Side(style="thin"),
#         )

#         # Set column widths
#         ws.column_dimensions["A"].width = 65
#         ws.column_dimensions["B"].width = 15
#         ws.column_dimensions["C"].width = 18

#         current_row = 1

#         # Row 1: Company name
#         cell = ws.cell(row=current_row, column=1)
#         cell.value = company_name
#         cell.font = company_font
#         current_row += 1

#         # Row 2: Period label
#         cell = ws.cell(row=current_row, column=1)
#         cell.value = period_label
#         cell.font = period_font
#         current_row += 1

#         # Row 3: Empty
#         current_row += 1

#         # Row 4: Entity info
#         cell = ws.cell(row=current_row, column=1)
#         cell.value = entity_info
#         cell.font = entity_font
#         current_row += 1

#         # Row 5: Cashflow header with yellow background
#         cell = ws.cell(row=current_row, column=1)
#         cell.value = "Cashflow"
#         cell.font = bold_font
#         cell.fill = yellow_fill
#         ws.merge_cells(f"A{current_row}:C{current_row}")
#         current_row += 1

#         # Row 6: Empty
#         current_row += 1

#         # Row 7: Gray header row
#         headers = ["", "Entity", ""]
#         for col_idx, header in enumerate(headers, start=1):
#             cell = ws.cell(row=current_row, column=col_idx)
#             cell.value = header
#             cell.fill = light_gray_fill
#             cell.border = thin_border
#         current_row += 1

#         # Row 8: Period Id row
#         cell = ws.cell(row=current_row, column=2)
#         cell.value = "Period Id"
#         cell.fill = light_gray_fill
#         cell.border = thin_border
        
#         cell = ws.cell(row=current_row, column=3)
#         # Format: "MAR 25 (YTD)" from "2025 Mar YTD"
#         parts = period_label.split()
#         if len(parts) >= 3:
#             formatted_period = f"{parts[1].upper()} {parts[0][-2:]} ({parts[2]})"
#         else:
#             formatted_period = period_label
#         cell.value = formatted_period
#         cell.fill = light_gray_fill
#         cell.border = thin_border
#         cell.font = bold_font
#         current_row += 1

#         # Row 9: Currency row
#         cell = ws.cell(row=current_row, column=2)
#         cell.value = "Currency"
#         cell.fill = light_gray_fill
#         cell.border = thin_border
        
#         cell = ws.cell(row=current_row, column=3)
#         cell.value = currency
#         cell.fill = light_gray_fill
#         cell.border = thin_border
#         cell.font = bold_font
#         current_row += 1

#         # Row 10: Scenario row
#         cell = ws.cell(row=current_row, column=2)
#         cell.value = "Scenario"
#         cell.fill = light_gray_fill
#         cell.border = thin_border
        
#         cell = ws.cell(row=current_row, column=3)
#         cell.value = scenario
#         cell.fill = light_gray_fill
#         cell.border = thin_border
#         cell.font = bold_font
#         current_row += 1

#         # Row 11: Column headers
#         cell = ws.cell(row=current_row, column=1)
#         cell.value = "Item Label / Nature Of Report"
#         cell.fill = light_gray_fill
#         cell.border = thin_border
#         cell.font = header_font
        
#         cell = ws.cell(row=current_row, column=2)
#         cell.value = "Consol Code"
#         cell.fill = light_gray_fill
#         cell.border = thin_border
#         cell.font = header_font
        
#         cell = ws.cell(row=current_row, column=3)
#         cell.value = "Standalone"
#         cell.fill = light_gray_fill
#         cell.border = thin_border
#         cell.font = header_font
#         current_row += 1

#         # Helper function to add data row
#         def add_data_row(label: str, consol_code: str = "", amount: Optional[float] = None, 
#                         bold: bool = False, section_header: bool = False):
#             nonlocal current_row
            
#             # Column A: Label
#             cell = ws.cell(row=current_row, column=1)
#             cell.value = label
#             cell.border = thin_border
#             if section_header:
#                 cell.font = section_header_font
#             elif bold:
#                 cell.font = bold_font
#             else:
#                 cell.font = normal_font
            
#             # Column B: Consol Code
#             cell = ws.cell(row=current_row, column=2)
#             cell.value = consol_code
#             cell.border = thin_border
#             cell.font = normal_font
#             cell.alignment = Alignment(horizontal="left")
            
#             # Column C: Amount
#             cell = ws.cell(row=current_row, column=3)
#             if amount is not None:
#                 cell.value = CashFlowFinalyzerService._format_amount(amount)
#             cell.border = thin_border
#             if bold:
#                 cell.font = bold_font
#             else:
#                 cell.font = normal_font
#             cell.alignment = Alignment(horizontal="right")
            
#             current_row += 1

#         # Helper function to get amount from items
#         def get_amount(key: str, *alternative_keys: str) -> Optional[float]:
#             """Get amount by key, trying variations and alternatives."""
#             # List of all keys to try
#             all_keys = [key] + list(alternative_keys)
            
#             for search_key in all_keys:
#                 # Try exact match first
#                 if search_key in cashflow_items:
#                     return cashflow_items[search_key]
                
#                 # Try case-insensitive match
#                 key_lower = search_key.lower()
#                 for item_key, value in cashflow_items.items():
#                     if item_key.lower() == key_lower:
#                         return value
                
#                 # Try partial match (contains)
#                 for item_key, value in cashflow_items.items():
#                     if key_lower in item_key.lower():
#                         return value
            
#             return None

#         # CASH FLOW SECTIONS
#         add_data_row("Cashflow", section_header=True)
#         add_data_row("Net Value", section_header=True)
        
#         # OPERATING ACTIVITIES
#         add_data_row("CASH FLOW FROM OPERATING ACTIVITIES", section_header=True)
        
#         add_data_row("Operating Profit/ (Loss) before Working Capital Changes", 
#                     "", get_amount("Operating profit before working capital changes"))
        
#         add_data_row("Profit before taxes", 
#                     "", get_amount("Profit before tax"))
        
#         add_data_row("Finance cost reversal", 
#                     "", get_amount("Finance cost"))
        
#         add_data_row("Interest Income reversal", 
#                     "", get_amount("Interest income"))
        
#         add_data_row("Gain/ (loss) on sale of property, plant and equipment (net)", 
#                     "", get_amount("Loss on sale of property, plant and equipment"))
        
#         add_data_row("Provision for doubtful debts", 
#                     "", get_amount("Provision for bad and doubtful debts"))
        
#         add_data_row("Unrealised loss from exchange differences (net)", 
#                     "", get_amount("Unrealised loss/ gain on foreign exchange (net)"))
        
#         add_data_row("Depreciation & Amortisation Reversal", 
#                     "", get_amount("Depreciation on property, plant and equipment and right of use assets"))
        
#         add_data_row("Share-based payment expense", 
#                     "CS61204Z2Z00", get_amount("Share-based payment expense"))
        
#         add_data_row("Adjusted for Working Capital changes", section_header=True)
        
#         add_data_row("Changes in Assets and Liabilities", section_header=True)
        
#         add_data_row("Trade receivables parent", 
#                     "", get_amount("Movement in trade receivables"))
        
#         add_data_row("Other Financial Assets", 
#                     "", get_amount("Movement in financial assets"))
        
#         add_data_row("Other Financial Assets Non Current", 
#                     "", get_amount("Other Financial Assets Non Current"))
        
#         add_data_row("Other Financial Assets Current", 
#                     "", get_amount("Other Financial Assets Current"))
        
#         add_data_row("Trade payables", 
#                     "", get_amount("Movement in trade payables"))
        
#         add_data_row("Other Assets", 
#                     "", get_amount("Movement in other assets"))
        
#         add_data_row("Other Assets Non Current", 
#                     "", get_amount("Other Assets Non Current"))
        
#         add_data_row("Other Assets Current", 
#                     "", get_amount("Other Assets Current"))
        
#         add_data_row("Financial Liabilities", 
#                     "", get_amount("Movement in other financial liabilities"))
        
#         add_data_row("Financial Liabilities Current", 
#                     "", get_amount("Financial Liabilities Current"))
        
#         add_data_row("Other Liabilities", 
#                     "", get_amount("Movement in other liabilities & contact liabilities"))
        
#         add_data_row("Inventories", 
#                     "", get_amount("Movement in inventories"))
        
#         add_data_row("Taxes paid/(refund)", 
#                     "", get_amount("Income tax refund/paid (net)"))
        
#         add_data_row("Income taxes paid", 
#                     "", get_amount("Income tax paid (net)"))
        
#         add_data_row("Def Tax Movement", 
#                     "", get_amount("Deferred tax"))
        
#         # INVESTING ACTIVITIES
#         add_data_row("CASH FLOW FROM INVESTING ACTIVITIES", section_header=True)
        
#         add_data_row("Purchase of Property, plant and equipment", 
#                     "", get_amount("Purchase of property, plant and equipment"))
        
#         add_data_row("Purchase of PPE", 
#                     "", get_amount("Purchase of PPE"))
        
#         add_data_row("Purchase of Intangibles", 
#                     "", get_amount("Purchase of Intangibles"))
        
#         add_data_row("Loans given", 
#                     "", get_amount("Loan Given"))
        
#         add_data_row("Interest received", 
#                     "", get_amount("Interest received"))
        
#         add_data_row("Movement in increase in bank deposits", 
#                     "", get_amount("Investment in bank deposits (net)"))
        
#         add_data_row("Purchase of Investments", 
#                     "", get_amount("Purchase of Investments"))
        
#         # FINANCING ACTIVITIES
#         add_data_row("CASH FLOW FROM FINANCING ACTIVITIES", section_header=True)
        
#         add_data_row("Interest paid", 
#                     "", get_amount("Interest paid"))
        
#         add_data_row("Dividend Paid", 
#                     "", get_amount("Dividend Paid"))
        
#         add_data_row("Proceeds from Borrowings", 
#                     "", get_amount("Proceeds from non-current borrowings"))
        
#         add_data_row("Interest paid on lease Liabilities", 
#                     "", get_amount("Interest on lease liabilities"))
        
#         add_data_row("Payment of principal portion of lease Liabilities", 
#                     "", get_amount("Payment of principal portion of lease liabilities (including interest)"))
        
#         add_data_row("ESOPs contribution from parent entity", 
#                     "", get_amount("ESOPs contribution from parent entity"))
        
#         # NET CHANGE IN CASH
#         add_data_row("Net Increase / (Decrease) in Cash and Cash Equivalents", 
#                     "", get_amount("Net increase/(decrease) in cash and cash equivalents"), bold=True)

#         # Save workbook
#         wb.save(output_file)
        
#         return output_file

#     @staticmethod
#     def check_cashflow_finalyzer_readiness(company_name: str) -> Dict:
#         """Check if Cash Flow Statement is ready for Finalyzer generation."""
#         cashflow_file = CashFlowFinalyzerService._find_cashflow_note_file(company_name)
        
#         if cashflow_file:
#             return {
#                 "company_name": company_name,
#                 "is_ready": True,
#                 "markdown_file": str(cashflow_file),
#                 "message": "Cash Flow Statement markdown file found. Ready to generate Finalyzer."
#             }
#         else:
#             return {
#                 "company_name": company_name,
#                 "is_ready": False,
#                 "markdown_file": None,
#                 "message": "Cash Flow Statement markdown file not found. Please generate it first."
#             }

#     @staticmethod
#     def list_cashflow_finalyzer_statements(company_name: str) -> Dict:
#         """List all generated Cash Flow Finalyzer statements."""
#         try:
#             path_service = PathService(company_name)
#             cashflow_dir = path_service.get_financial_statements_dir(company_name) / "CashFlow_Finalyzer"

#             if not cashflow_dir.exists():
#                 return {
#                     "company_name": company_name,
#                     "statements": [],
#                     "count": 0
#                 }

#             statements = []
#             for file_path in cashflow_dir.glob("CashFlow_Finalyzer_*.xlsx"):
#                 statements.append({
#                     "filename": file_path.name,
#                     "path": str(file_path),
#                     "generated_at": datetime.fromtimestamp(
#                         file_path.stat().st_mtime
#                     ).isoformat(),
#                     "size": file_path.stat().st_size
#                 })

#             statements.sort(key=lambda x: x["generated_at"], reverse=True)

#             return {
#                 "company_name": company_name,
#                 "statements": statements,
#                 "count": len(statements),
#                 "latest": statements[0] if statements else None
#             }

#         except Exception as e:
#             return {
#                 "company_name": company_name,
#                 "statements": [],
#                 "count": 0,
#                 "error": str(e)
#             }

# ============================================================================
# FILE: backend/services/cashflow_finalyzer_service.py
# ============================================================================
"""Cash Flow Statement Finalyzer service - generates Excel matching template."""

import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from backend.config.settings import settings
from backend.services.path_service import PathService


class CashFlowFinalyzerService:
    """Service for generating Cash Flow Statement Finalyzer matching exact Excel template."""

    @staticmethod
    def _find_cashflow_note_file(company_name: str) -> Optional[Path]:
        """Find the Cash Flow Statement markdown file."""
        entity_notes_dir = settings.get_entity_generated_notes_dir(company_name)

        if not entity_notes_dir.exists():
            return None

        patterns = [
            "Note_CASHFLOW_*.md",
            "noteCASHFLOW_*.md",
            "cashflow_*.md",
            "cash_flow_*.md"
        ]

        for pattern in patterns:
            note_files = list(entity_notes_dir.glob(pattern))
            if note_files:
                return max(note_files, key=lambda p: p.stat().st_mtime)

        return None

    @staticmethod
    def _parse_amount(amount_str: str) -> Optional[float]:
        """Parse amount string to float."""
        if not amount_str or amount_str.strip() == '' or amount_str == '-':
            return None
        
        # Remove currency symbols, commas, and spaces
        amount_str = amount_str.replace('₹', '').replace(',', '').strip()
        
        # Check if bracketed (negative)
        is_negative = amount_str.startswith('(') and amount_str.endswith(')')
        
        if is_negative:
            amount_str = amount_str[1:-1]
        
        try:
            amount = float(amount_str)
            return -amount if is_negative else amount
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _extract_cashflow_items(md_content: str) -> Dict[str, float]:
        """Extract all cash flow line items from markdown content."""
        items = {}
        
        # Find SECTION 2: CASH FLOW STATEMENT (FINAL FORMAT)
        # This section has the correct field names we need
        section2_match = re.search(
            r'SECTION 2: CASH FLOW STATEMENT \(FINAL FORMAT\)(.*?)(?:\Z|$)',
            md_content,
            re.DOTALL | re.IGNORECASE
        )
        
        if section2_match:
            print("✅ Found SECTION 2: CASH FLOW STATEMENT (FINAL FORMAT)")
            content_to_parse = section2_match.group(1)
        else:
            # Fallback: Try to find the final format section without "SECTION 2" prefix
            final_format_match = re.search(
                r'# CASH FLOW STATEMENT \(INDIRECT METHOD\)(.*?)(?:\Z|$)',
                md_content,
                re.DOTALL | re.IGNORECASE
            )
            if final_format_match:
                print("✅ Found CASH FLOW STATEMENT (INDIRECT METHOD)")
                content_to_parse = final_format_match.group(1)
            else:
                print("⚠️  Using entire markdown content (section not found)")
                content_to_parse = md_content
        
        # Pattern to match table rows: | Description | Amount |
        table_pattern = r'\|\s*([^|]+?)\s*\|\s*([^|]*)\s*\|'
        
        for match in re.finditer(table_pattern, content_to_parse):
            description = match.group(1).strip()
            amount_str = match.group(2).strip()
            
            # Skip headers and separators
            if ('Particulars' in description or 'Amount' in description or 
                '---' in description or description == '' or ':---' in description or
                'GL Code' in description or 'GL Description' in description or
                'Ind AS Minor' in description):
                continue
            
            # Clean description - remove markdown formatting
            clean_desc = description.replace('**', '').replace('*', '').strip()
            
            # Skip section headers that are purely descriptive
            if clean_desc in ['Cash flow from operating activities', 
                            'Cash flows from investing activities',
                            'Cash flows from financing activities',
                            'Adjustments for:',
                            'Changes in working Capital',
                            'Changes in working capital',
                            'Cash and cash equivalents include (refer note 13)',
                            'Cash and cash equivalents include (refer note 11)',
                            'Non- cash financing and investing activities:',
                            'Balance with banks',
                            '']:
                continue
            
            # Parse amount
            amount = CashFlowFinalyzerService._parse_amount(amount_str)
            
            if amount is not None:
                # Store with cleaned description as key
                items[clean_desc] = amount
                
                # Also store common variations for better matching
                # Store simplified version without parentheses
                simplified = re.sub(r'\([^)]*\)', '', clean_desc).strip()
                if simplified != clean_desc and simplified:
                    items[simplified] = amount
                
                # Store version with forward slashes replaced
                if '/' in clean_desc:
                    items[clean_desc.replace('/', ' ')] = amount
                
                # Store version with special chars removed
                clean_no_special = re.sub(r'[/\(\),&]', ' ', clean_desc).strip()
                clean_no_special = re.sub(r'\s+', ' ', clean_no_special)
                if clean_no_special != clean_desc and clean_no_special:
                    items[clean_no_special] = amount
        
        # Debug: Print extracted items
        print(f"\n{'='*80}")
        print(f"EXTRACTED {len(items)} ITEMS FROM MARKDOWN:")
        print(f"{'='*80}")
        for key, value in sorted(items.items())[:30]:  # Show first 30 items sorted
            print(f"  '{key}': {value:,.2f}")
        if len(items) > 30:
            print(f"  ... and {len(items) - 30} more items")
        print(f"{'='*80}\n")
        
        return items

    @staticmethod
    def _format_amount(value: Optional[float], decimal_places: int = 2) -> str:
        """Format amount with decimal places."""
        if value is None:
            return ""
        return f"{value:,.{decimal_places}f}"

    @staticmethod
    def generate_cashflow_finalyzer(
        company_name: str,
        period_label: str = "2025 Mar YTD",
        entity_info: str = "Entity: CPM  Book: 8937,IndAS - IndAS entities,Standalone,MYR,Actual",
        currency: str = "Malaysian Ringgit",
        scenario: str = "Cashflow",
    ) -> Dict:
        """
        Generate Cash Flow Statement Finalyzer matching exact Excel template.

        Args:
            company_name: Name of the company
            period_label: Period label (e.g., "2025 Mar YTD")
            entity_info: Entity information line
            currency: Currency name
            scenario: Scenario type

        Returns:
            Dictionary with success status and file path
        """
        try:
            # Find cash flow markdown file
            cashflow_file = CashFlowFinalyzerService._find_cashflow_note_file(company_name)
            
            if not cashflow_file:
                return {
                    "success": False,
                    "message": "Cash Flow Statement markdown file not found. Please generate it first."
                }

            print(f"\n📄 Reading Cash Flow markdown from: {cashflow_file}")

            # Read markdown content
            with open(cashflow_file, 'r', encoding='utf-8') as f:
                md_content = f.read()

            print(f"📊 Markdown file size: {len(md_content)} characters")

            # Extract all line items
            cashflow_items = CashFlowFinalyzerService._extract_cashflow_items(md_content)

            if not cashflow_items:
                return {
                    "success": False,
                    "message": "No cash flow items could be extracted from the markdown file."
                }

            print(f"✅ Successfully extracted {len(cashflow_items)} items")

            # Generate Excel file
            output_file = CashFlowFinalyzerService._export_to_excel_finalyzer(
                company_name=company_name,
                period_label=period_label,
                entity_info=entity_info,
                currency=currency,
                scenario=scenario,
                cashflow_items=cashflow_items,
            )

            return {
                "success": True,
                "message": f"Cash Flow Finalyzer generated successfully for {company_name}",
                "output_file": str(output_file),
                "company_name": company_name,
                "period_label": period_label,
                "items_extracted": len(cashflow_items)
            }

        except Exception as e:
            import traceback
            print(f"\n❌ ERROR: {str(e)}")
            print(traceback.format_exc())
            return {
                "success": False,
                "message": f"Error generating Cash Flow Finalyzer: {str(e)}"
            }

    @staticmethod
    def _export_to_excel_finalyzer(
        company_name: str,
        period_label: str,
        entity_info: str,
        currency: str,
        scenario: str,
        cashflow_items: Dict[str, float],
    ) -> Path:
        """Export Cash Flow Finalyzer to Excel with exact template formatting."""
        
        # Create output directory
        path_service = PathService(company_name)
        cashflow_dir = path_service.get_financial_statements_dir(company_name) / "CashFlow_Finalyzer"
        cashflow_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = cashflow_dir / f"CashFlow_Finalyzer_{timestamp}.xlsx"

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Cash Flow Statement"

        # Define styles
        company_font = Font(name="Calibri", size=11, color="C00000")
        period_font = Font(name="Calibri", size=11, bold=True)
        entity_font = Font(name="Calibri", size=9)
        header_font = Font(name="Calibri", size=11, bold=True)
        section_header_font = Font(name="Calibri", size=11, bold=True, color="C00000")
        normal_font = Font(name="Calibri", size=11)
        bold_font = Font(name="Calibri", size=11, bold=True)

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        light_gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        ws.column_dimensions["A"].width = 65
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 18

        current_row = 1

        # Header rows
        cell = ws.cell(row=current_row, column=1)
        cell.value = company_name
        cell.font = company_font
        current_row += 1

        cell = ws.cell(row=current_row, column=1)
        cell.value = period_label
        cell.font = period_font
        current_row += 1

        current_row += 1

        cell = ws.cell(row=current_row, column=1)
        cell.value = entity_info
        cell.font = entity_font
        current_row += 1

        cell = ws.cell(row=current_row, column=1)
        cell.value = "Cashflow"
        cell.font = bold_font
        cell.fill = yellow_fill
        ws.merge_cells(f"A{current_row}:C{current_row}")
        current_row += 1

        current_row += 1

        headers = ["", "Entity", ""]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.fill = light_gray_fill
            cell.border = thin_border
        current_row += 1

        cell = ws.cell(row=current_row, column=2)
        cell.value = "Period Id"
        cell.fill = light_gray_fill
        cell.border = thin_border
        
        cell = ws.cell(row=current_row, column=3)
        parts = period_label.split()
        if len(parts) >= 3:
            formatted_period = f"{parts[1].upper()} {parts[0][-2:]} ({parts[2]})"
        else:
            formatted_period = period_label
        cell.value = formatted_period
        cell.fill = light_gray_fill
        cell.border = thin_border
        cell.font = bold_font
        current_row += 1

        cell = ws.cell(row=current_row, column=2)
        cell.value = "Currency"
        cell.fill = light_gray_fill
        cell.border = thin_border
        
        cell = ws.cell(row=current_row, column=3)
        cell.value = currency
        cell.fill = light_gray_fill
        cell.border = thin_border
        cell.font = bold_font
        current_row += 1

        cell = ws.cell(row=current_row, column=2)
        cell.value = "Scenario"
        cell.fill = light_gray_fill
        cell.border = thin_border
        
        cell = ws.cell(row=current_row, column=3)
        cell.value = scenario
        cell.fill = light_gray_fill
        cell.border = thin_border
        cell.font = bold_font
        current_row += 1

        cell = ws.cell(row=current_row, column=1)
        cell.value = "Item Label / Nature Of Report"
        cell.fill = light_gray_fill
        cell.border = thin_border
        cell.font = header_font
        
        cell = ws.cell(row=current_row, column=2)
        cell.value = "Consol Code"
        cell.fill = light_gray_fill
        cell.border = thin_border
        cell.font = header_font
        
        cell = ws.cell(row=current_row, column=3)
        cell.value = "Standalone"
        cell.fill = light_gray_fill
        cell.border = thin_border
        cell.font = header_font
        current_row += 1

        # Helper function to add data row
        def add_data_row(label: str, consol_code: str = "", amount: Optional[float] = None, 
                        bold: bool = False, section_header: bool = False):
            nonlocal current_row
            
            cell = ws.cell(row=current_row, column=1)
            cell.value = label
            cell.border = thin_border
            if section_header:
                cell.font = section_header_font
            elif bold:
                cell.font = bold_font
            else:
                cell.font = normal_font
            
            cell = ws.cell(row=current_row, column=2)
            cell.value = consol_code
            cell.border = thin_border
            cell.font = normal_font
            cell.alignment = Alignment(horizontal="left")
            
            cell = ws.cell(row=current_row, column=3)
            if amount is not None:
                cell.value = CashFlowFinalyzerService._format_amount(amount)
            cell.border = thin_border
            if bold:
                cell.font = bold_font
            else:
                cell.font = normal_font
            cell.alignment = Alignment(horizontal="right")
            
            current_row += 1

        # Helper function to get amount from items with detailed logging
        def get_amount(key: str, *alternative_keys: str) -> Optional[float]:
            """Get amount by key, trying variations and alternatives."""
            all_keys = [key] + list(alternative_keys)
            
            for search_key in all_keys:
                # Try exact match
                if search_key in cashflow_items:
                    value = cashflow_items[search_key]
                    print(f"  ✓ MATCHED '{search_key}' → {value:,.2f}")
                    return value
                
                # Try case-insensitive match
                key_lower = search_key.lower()
                for item_key, value in cashflow_items.items():
                    if item_key.lower() == key_lower:
                        print(f"  ✓ MATCHED (case-insensitive) '{search_key}' → '{item_key}' → {value:,.2f}")
                        return value
                
                # Try partial match (contains)
                for item_key, value in cashflow_items.items():
                    if key_lower in item_key.lower():
                        print(f"  ✓ MATCHED (partial) '{search_key}' → '{item_key}' → {value:,.2f}")
                        return value
            
            print(f"  ✗ NOT FOUND: '{key}' (tried {len(all_keys)} variations)")
            return None

        print(f"\n{'='*80}")
        print(f"POPULATING EXCEL ROWS:")
        print(f"{'='*80}\n")

        # CASH FLOW SECTIONS
        add_data_row("Cashflow", section_header=True)
        add_data_row("Net Value", section_header=True)
        
        add_data_row("CASH FLOW FROM OPERATING ACTIVITIES", section_header=True)
        
        print("Operating Profit before WC Changes:")
        add_data_row("Operating Profit/ (Loss) before Working Capital Changes", 
                    "", get_amount("Operating profit before working capital changes"))
        
        print("Profit before taxes:")
        add_data_row("Profit before taxes", 
                    "", get_amount("Profit before tax"))
        
        print("Finance cost:")
        add_data_row("Finance cost reversal", 
                    "", get_amount("Finance cost"))
        
        print("Interest income:")
        add_data_row("Interest Income reversal", 
                    "", get_amount("Interest income"))
        
        print("Loss on PPE sale:")
        add_data_row("Gain/ (loss) on sale of property, plant and equipment (net)", 
                    "", get_amount("Loss on sale of property, plant and equipment"))
        
        print("Provision for doubtful debts:")
        add_data_row("Provision for doubtful debts", 
                    "", get_amount("Provision for bad and doubtful debts"))
        
        print("Unrealised forex:")
        add_data_row("Unrealised loss from exchange differences (net)", 
                    "", get_amount("Unrealised loss gain on foreign exchange", 
                                 "Unrealised loss/ gain on foreign exchange (net)"))
        
        print("Depreciation:")
        add_data_row("Depreciation & Amortisation Reversal", 
                    "", get_amount("Depreciation on property, plant and equipment and right of use assets",
                                 "Depreciation on property, plant and equipment and ROU assets"))
        
        print("Share-based payment:")
        add_data_row("Share-based payment expense", 
                    "", get_amount("Share-based payment expense"))
        
        add_data_row("Adjusted for Working Capital changes", section_header=True)
        add_data_row("Changes in Assets and Liabilities", section_header=True)
        
        print("Trade receivables:")
        add_data_row("Trade receivables parent", 
                    "", get_amount("Movement in trade receivables",
                                 "Increase/Decrease in trade receivables"))
        
        print("Other Financial Assets:")
        add_data_row("Other Financial Assets", 
                    "", get_amount("Movement in financial assets",
                                 "Increase/Decrease in financial assets"))
        
        add_data_row("Other Financial Assets Non Current", "", None)
        add_data_row("Other Financial Assets Current", "", None)
        
        print("Trade payables:")
        add_data_row("Trade payables", 
                    "", get_amount("Movement in trade payables",
                                 "Increase/Decrease in trade payables"))
        
        print("Other Assets:")
        add_data_row("Other Assets", 
                    "", get_amount("Movement in other assets",
                                 "Increase/Decrease in other assets"))
        
        add_data_row("Other Assets Non Current", "", None)
        add_data_row("Other Assets Current", "", None)
        
        print("Financial Liabilities:")
        add_data_row("Financial Liabilities", 
                    "", get_amount("Movement in other financial liabilities",
                                 "Increase/Decrease in other financial liabilities"))
        
        add_data_row("Financial Liabilities Current", "", None)
        
        print("Other Liabilities:")
        add_data_row("Other Liabilities", 
                    "", get_amount("Movement in other liabilities & contact liabilities",
                                 "Movement in other liabilities & contract liabilities",
                                 "Increase/Decrease in other liabilities"))
        
        print("Inventories:")
        add_data_row("Inventories", 
                    "", get_amount("Movement in inventories",
                                 "Increase/Decrease in inventories"))
        
        print("Tax paid:")
        add_data_row("Taxes paid/(refund)", 
                    "", get_amount("Income tax refund/paid (net)",
                                 "Income tax refundpaid"))
        
        add_data_row("Income taxes paid", 
                    "", get_amount("Income tax paid (net)",
                                 "Income tax refund/paid (net)"))
        
        add_data_row("Def Tax Movement", "", None)
        
        print("Operating activities total:")
        add_data_row("Net cash flows/ (used in) operating activities (A)",
                    "", get_amount("Net cash flows used in operating activities (A)",
                                 "Net cash flows/ used in operating activities",
                                 "Net cash flows (used in) operating activities"),
                    bold=True)
        
        add_data_row("CASH FLOW FROM INVESTING ACTIVITIES", section_header=True)
        
        print("PPE purchase:")
        add_data_row("Purchase of Property, plant and equipment", 
                    "", get_amount("Purchase of property, plant and equipment and intangible assets",
                                 "Purchase of property, plant and equipment"))
        
        add_data_row("Purchase of PPE", "", None)
        add_data_row("Purchase of Intangibles", "", None)
        
        print("Proceeds from disposal:")
        add_data_row("Proceeds from disposal of property, plant and equipment",
                    "", get_amount("Proceeds from disposal of property, plant and equipment"))
        
        print("Loans given:")
        add_data_row("Loans given", "", get_amount("Loan Given"))
        
        print("Proceeds from loans:")
        add_data_row("Proceeds against loans given",
                    "", get_amount("Proceeds against loans given"))
        
        print("Interest received:")
        add_data_row("Interest received", "", get_amount("Interest received"))
        
        print("Dividend income:")
        add_data_row("Dividend income received",
                    "", get_amount("Dividend income received"))
        
        print("Bank deposits:")
        add_data_row("Movement in increase in bank deposits", 
                    "", get_amount("Investment/maturity in bank deposits",
                                 "(Investment)/maturity in bank deposits",
                                 "Investment in bank deposits (net)"))
        
        add_data_row("Purchase of Investments", 
                    "", get_amount("Investment in associates"))
        
        print("Investing activities total:")
        add_data_row("Net cash generated from /(used in) investing activities (B)",
                    "", get_amount("Net cash generated from used in investing activities (B)",
                                 "Net cash generated from /(used in) investing activities",
                                 "Net cash generated from /used in investing activities"),
                    bold=True)
        
        add_data_row("CASH FLOW FROM FINANCING ACTIVITIES", section_header=True)
        
        print("Proceeds from borrowings:")
        add_data_row("Proceeds from non-current borrowings",
                    "", get_amount("Proceeds from non-current borrowings"))
        
        print("Repayments of borrowings:")
        add_data_row("Repayments of non-current borrowings",
                    "", get_amount("Repayments of non-current borrowings"))
        
        print("Current borrowings:")
        add_data_row("(Repayments)/proceeds from current borrowings (net)",
                    "", get_amount("Repayments/proceeds from current borrowings (net)",
                                 "(Repayments)/proceeds from current borrowings",
                                 "Repaymentproceeds from current borrowings"))
        
        print("Interest paid:")
        add_data_row("Interest paid", "", get_amount("Interest paid"))
        
        add_data_row("Dividend Paid", "", None)
        
        print("Interest on lease:")
        add_data_row("Interest paid on lease Liabilities", 
                    "", get_amount("Interest on lease liabilities"))
        
        print("Lease payment:")
        add_data_row("Payment of principal portion of lease Liabilities", 
                    "", get_amount("Payment of principal portion of lease liabilities (including interest)",
                                 "Payment of principal portion of lease liabilities"))
        
        add_data_row("ESOPs contribution from parent entity", "", None)
        
        print("Financing activities total:")
        add_data_row("Net cash generated from /(used in) financing activities (C)",
                    "", get_amount("Net cash generated from used in financing activities (C)",
                                 "Net cash generated from /(used in) financing activities",
                                 "Net cash generated from /used in financing activities"),
                    bold=True)
        
        print("Net change in cash:")
        add_data_row("Net Increase / (Decrease) in Cash and Cash Equivalents", 
                    "", get_amount("Net increase/decrease in cash and cash equivalents (A+B+C+D)",
                                 "Net increase/decrease in cash and cash equivalents",
                                 "Net Increase/Decrease in Cash and Cash Equivalents"), 
                    bold=True)

        print(f"\n{'='*80}")
        print(f"EXCEL GENERATION COMPLETE")
        print(f"{'='*80}\n")

        wb.save(output_file)
        
        return output_file

    @staticmethod
    def check_cashflow_finalyzer_readiness(company_name: str) -> Dict:
        """Check if Cash Flow Statement is ready for Finalyzer generation."""
        cashflow_file = CashFlowFinalyzerService._find_cashflow_note_file(company_name)
        
        if cashflow_file:
            return {
                "company_name": company_name,
                "is_ready": True,
                "markdown_file": str(cashflow_file),
                "message": "Cash Flow Statement markdown file found. Ready to generate Finalyzer."
            }
        else:
            return {
                "company_name": company_name,
                "is_ready": False,
                "markdown_file": None,
                "message": "Cash Flow Statement markdown file not found. Please generate it first."
            }

    @staticmethod
    def list_cashflow_finalyzer_statements(company_name: str) -> Dict:
        """List all generated Cash Flow Finalyzer statements."""
        try:
            path_service = PathService(company_name)
            cashflow_dir = path_service.get_financial_statements_dir(company_name) / "CashFlow_Finalyzer"

            if not cashflow_dir.exists():
                return {
                    "company_name": company_name,
                    "statements": [],
                    "count": 0
                }

            statements = []
            for file_path in cashflow_dir.glob("CashFlow_Finalyzer_*.xlsx"):
                statements.append({
                    "filename": file_path.name,
                    "path": str(file_path),
                    "generated_at": datetime.fromtimestamp(
                        file_path.stat().st_mtime
                    ).isoformat(),
                    "size": file_path.stat().st_size
                })

            statements.sort(key=lambda x: x["generated_at"], reverse=True)

            return {
                "company_name": company_name,
                "statements": statements,
                "count": len(statements),
                "latest": statements[0] if statements else None
            }

        except Exception as e:
            return {
                "company_name": company_name,
                "statements": [],
                "count": 0,
                "error": str(e)
            }