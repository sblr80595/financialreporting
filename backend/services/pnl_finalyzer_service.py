import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from backend.config.settings import settings
from backend.models.financial_statement import (
    PLGenerationResponse,
    ProfitLossStatement,
)
from backend.services.path_service import PathService


class PNLFinalyzerService:
    """Service for generating PNL Statement Finalyzer matching the exact Excel template."""

    ENTITY_PNL_CONFIGS = {
        "default": {
            "income_notes": ["24", "25"],
            "expense_notes": ["26", "27", "28", "29", "30", "31"],
            "tax_notes": ["32"],
        },
    }

    NOTE_DESCRIPTIONS = {
        "24": "Revenue from operations",
        "25": "Other income",
        "26": "Purchase of stock-in-trade including inventory movement",
        "27": "Changes in inventories of finished goods, stock-in-trade and work-in-progress",
        "28": "Employee benefit expense",
        "29": "Finance costs",
        "30": "Depreciation and amortisation expense",
        "31": "Other expenses",
        "32": "Current tax",
        "33": "Deferred tax (Credit)/expense",
    }

    @staticmethod
    def _get_entity_config(company_name: str) -> Dict[str, List[str]]:
        """Get PNL configuration for a specific entity."""
        return PNLFinalyzerService.ENTITY_PNL_CONFIGS.get(
            company_name, PNLFinalyzerService.ENTITY_PNL_CONFIGS["default"]
        )

    @staticmethod
    def _find_latest_note_file(company_name: str, note_number: str) -> Optional[Path]:
        """Find the most recent generated note file."""
        entity_notes_dir = settings.get_entity_generated_notes_dir(company_name)

        if not entity_notes_dir.exists():
            return None

        note_files = list(entity_notes_dir.glob(f"note{note_number}_*.md"))
        if not note_files:
            note_files = list(entity_notes_dir.glob(f"Note_{note_number}_*.md"))

        if not note_files:
            return None

        return max(note_files, key=lambda p: p.stat().st_mtime)

    @staticmethod
    def _extract_total_from_markdown(md_content: str) -> Optional[float]:
        """Extract total amount from markdown note content."""
        patterns = [
            r"\|\s*\*\*TOTAL[^|]*\*\*\s*\|\s*\*\*\(₹?([\d,]+\.?\d*)\)\*\*",
            r"\|\s*\*\*TOTAL[^|]*\*\*\s*\|\s*\*\*\(([\d,]+\.?\d*)\)\*\*",
            r"\|\s*\*\*TOTAL[^|]*\*\*\s*\|\s*\*\*₹?([\d,]+\.?\d*)\*\*",
            r"\|\s*Total[^|]*\|\s*\(₹?([\d,]+\.?\d*)\)",
            r"\|\s*Total[^|]*\|\s*\(([\d,]+\.?\d*)\)",
            r"\|\s*Total[^|]*\|\s*₹?([\d,]+\.?\d*)",
            r"\*\*Total[^:]*:\*\*\s*\(₹?([\d,]+\.?\d*)\)",
            r"\*\*Total[^:]*:\*\*\s*\(([\d,]+\.?\d*)\)",
            r"\*\*Total[^:]*:\*\*\s*₹?([\d,]+\.?\d*)",
        ]

        for pattern in patterns:
            matches = re.findall(pattern, md_content, re.IGNORECASE)
            if matches:
                amount_str = matches[-1].replace(",", "")
                try:
                    amount = abs(float(amount_str))
                    return amount
                except ValueError:
                    continue

        return None

    @staticmethod
    def _extract_note_details(
        company_name: str, note_number: str
    ) -> Tuple[Optional[float], Optional[str]]:
        """Extract amount and description from a note file."""
        note_file = PNLFinalyzerService._find_latest_note_file(
            company_name, note_number
        )

        if not note_file:
            return None, PNLFinalyzerService.NOTE_DESCRIPTIONS.get(note_number)

        try:
            with open(note_file, "r", encoding="utf-8") as f:
                content = f.read()

            amount = PNLFinalyzerService._extract_total_from_markdown(content)
            description = PNLFinalyzerService.NOTE_DESCRIPTIONS.get(note_number)

            return amount, description

        except Exception as e:
            print(f"Error reading note {note_number}: {e}")
            return None, PNLFinalyzerService.NOTE_DESCRIPTIONS.get(note_number)

    @staticmethod
    def _format_amount(value: Optional[float], decimal_places: int = 2) -> str:
        """Format amount with decimal places."""
        if value is None:
            return ""
        
        value = abs(value)
        return f"{value:,.{decimal_places}f}"

    @staticmethod
    def generate_pnl_finalyzer(
        company_name: str,
        period_label: str = "2025 Mar YTD",
        entity_info: str = "Entity: CPM  Book: 8937,IndAS - IndAS entities,Standalone,MYR,Actual",
        currency: str = "Malaysian Ringgit",
        scenario: str = "Actual",
    ) -> PLGenerationResponse:
        """
        Generate PNL Statement Finalyzer matching exact Excel template.

        Args:
            company_name: Name of the company
            period_label: Period label (e.g., "2025 Mar YTD")
            entity_info: Entity information line
            currency: Currency name
            scenario: Scenario type

        Returns:
            PLGenerationResponse with generated statement and file path
        """
        try:
            config = PNLFinalyzerService._get_entity_config(company_name)

            # Extract all note amounts
            note_amounts = {}
            
            for note_num in config["income_notes"] + config["expense_notes"] + config["tax_notes"]:
                amount, _ = PNLFinalyzerService._extract_note_details(
                    company_name, note_num
                )
                if amount is not None:
                    note_amounts[note_num] = amount

            # Calculate totals
            income_total = sum(note_amounts.get(n, 0) for n in config["income_notes"])
            expense_total = sum(note_amounts.get(n, 0) for n in config["expense_notes"])
            tax_total = sum(note_amounts.get(n, 0) for n in config["tax_notes"])
            
            profit_before_tax = income_total - expense_total
            net_profit = profit_before_tax - tax_total

            # Generate Excel file
            output_file = PNLFinalyzerService._export_to_excel_finalyzer(
                company_name=company_name,
                period_label=period_label,
                entity_info=entity_info,
                currency=currency,
                scenario=scenario,
                note_amounts=note_amounts,
                config=config,
                income_total=income_total,
                expense_total=expense_total,
                tax_total=tax_total,
                profit_before_tax=profit_before_tax,
                net_profit=net_profit,
            )

            statement = ProfitLossStatement(
                company_name=company_name,
                period_ended=period_label,
                sections=[],
                metadata={
                    "generated_at": datetime.now().isoformat(),
                    "notes_used": list(note_amounts.keys()),
                    "type": "PNL_Finalyzer",
                },
            )

            return PLGenerationResponse(
                success=True,
                message=f"PNL Finalyzer generated successfully for {company_name}",
                statement=statement,
                output_file=str(output_file),
            )

        except Exception as e:
            return PLGenerationResponse(
                success=False, message=f"Error generating PNL Finalyzer: {str(e)}"
            )

    @staticmethod
    def _export_to_excel_finalyzer(
        company_name: str,
        period_label: str,
        entity_info: str,
        currency: str,
        scenario: str,
        note_amounts: Dict[str, float],
        config: Dict[str, List[str]],
        income_total: float,
        expense_total: float,
        tax_total: float,
        profit_before_tax: float,
        net_profit: float,
    ) -> Path:
        """Export PNL Finalyzer to Excel with exact template formatting."""
        
        # Create output directory
        path_service = PathService(company_name)
        pnl_dir = path_service.get_financial_statements_dir(company_name) / "PNL_Finalyzer"
        pnl_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = pnl_dir / f"PNL_Finalyzer_{timestamp}.xlsx"

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "P&L Statement"

        # Define styles
        # Fonts
        company_font = Font(name="Calibri", size=11, color="C00000")  # Red
        period_font = Font(name="Calibri", size=11, bold=True)
        entity_font = Font(name="Calibri", size=9)
        header_font = Font(name="Calibri", size=11, bold=True)
        section_header_font = Font(name="Calibri", size=11, bold=True, color="C00000")  # Red
        normal_font = Font(name="Calibri", size=11)
        bold_font = Font(name="Calibri", size=11, bold=True)

        # Fills
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        light_gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

        # Borders
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Set column widths
        ws.column_dimensions["A"].width = 65
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 18

        current_row = 1

        # Row 1: Company name
        cell = ws.cell(row=current_row, column=1)
        cell.value = company_name
        cell.font = company_font
        current_row += 1

        # Row 2: Period label
        cell = ws.cell(row=current_row, column=1)
        cell.value = period_label
        cell.font = period_font
        current_row += 1

        # Row 3: Empty
        current_row += 1

        # Row 4: Entity info
        cell = ws.cell(row=current_row, column=1)
        cell.value = entity_info
        cell.font = entity_font
        current_row += 1

        # Row 5: PL header with yellow background
        cell = ws.cell(row=current_row, column=1)
        cell.value = "PL"
        cell.font = bold_font
        cell.fill = yellow_fill
        ws.merge_cells(f"A{current_row}:C{current_row}")
        current_row += 1

        # Row 6: Empty
        current_row += 1

        # Row 7: Gray header row
        headers = ["", "Entity", ""]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.fill = light_gray_fill
            cell.border = thin_border
        current_row += 1

        # Row 8: Period Id row
        cell = ws.cell(row=current_row, column=2)
        cell.value = "Period Id"
        cell.fill = light_gray_fill
        cell.border = thin_border
        
        cell = ws.cell(row=current_row, column=3)
        cell.value = period_label.split()[1].upper() + " " + period_label.split()[0]  # "MAR 25 (YTD)"
        cell.fill = light_gray_fill
        cell.border = thin_border
        cell.font = bold_font
        current_row += 1

        # Row 9: Currency row
        cell = ws.cell(row=current_row, column=2)
        cell.value = "Currency"
        cell.fill = light_gray_fill
        cell.border = thin_border
        
        cell = ws.cell(row=current_row, column=3)
        cell.value = currency
        cell.fill = light_gray_fill
        cell.border = thin_border
        cell.font = bold_font
        current_row += 1

        # Row 10: Scenario row
        cell = ws.cell(row=current_row, column=2)
        cell.value = "Scenario"
        cell.fill = light_gray_fill
        cell.border = thin_border
        
        cell = ws.cell(row=current_row, column=3)
        cell.value = scenario
        cell.fill = light_gray_fill
        cell.border = thin_border
        cell.font = bold_font
        current_row += 1

        # Row 11: Column headers
        cell = ws.cell(row=current_row, column=1)
        cell.value = "Item Label / Nature Of Report"
        cell.fill = light_gray_fill
        cell.border = thin_border
        cell.font = header_font
        
        cell = ws.cell(row=current_row, column=2)
        cell.value = "Consol Code"
        cell.fill = light_gray_fill
        cell.border = thin_border
        cell.font = header_font
        
        cell = ws.cell(row=current_row, column=3)
        cell.value = "Standalone"
        cell.fill = light_gray_fill
        cell.border = thin_border
        cell.font = header_font
        current_row += 1

        # Helper function to add data row
        def add_data_row(label: str, consol_code: str = "", amount: Optional[float] = None, 
                        bold: bool = False, section_header: bool = False, orange_text: bool = False):
            nonlocal current_row
            
            # Column A: Label
            cell = ws.cell(row=current_row, column=1)
            cell.value = label
            cell.border = thin_border
            if section_header:
                cell.font = section_header_font  # Red/Orange text, NO background
            elif orange_text:
                cell.font = Font(name="Calibri", size=11, bold=True, color="C65911")  # Orange text
            elif bold:
                cell.font = bold_font
            else:
                cell.font = normal_font
            
            # Column B: Consol Code
            cell = ws.cell(row=current_row, column=2)
            cell.value = consol_code
            cell.border = thin_border
            cell.font = normal_font
            cell.alignment = Alignment(horizontal="left")
            
            # Column C: Amount
            cell = ws.cell(row=current_row, column=3)
            if amount is not None:
                cell.value = PNLFinalyzerService._format_amount(amount)
            cell.border = thin_border
            if orange_text:
                cell.font = Font(name="Calibri", size=11, bold=True, color="C65911")  # Orange text
            elif bold:
                cell.font = bold_font
            else:
                cell.font = normal_font
            cell.alignment = Alignment(horizontal="right")
            
            current_row += 1

        # INCOME SECTION
        add_data_row("Profit and loss", section_header=True)
        add_data_row("Income", section_header=True)
        
        add_data_row("Revenue from operations", "", note_amounts.get("24"))
        add_data_row("Other income", "", note_amounts.get("25"))
        add_data_row("Total Income", "", income_total, bold=True)
        
        # EXPENSES SECTION
        add_data_row("Expenses", section_header=True)
        add_data_row("Cost of Materail Consumed", section_header=True)
        
        add_data_row("Purchase of stock-in-trade including inventory movement", "", note_amounts.get("26"))
        add_data_row("Changes in inventories of finished goods, stock-in-trade and work-in-progress", "", note_amounts.get("27"))
        add_data_row("Employee benefit expense", "", note_amounts.get("28"))
        add_data_row("Finance costs", "", note_amounts.get("29"))
        add_data_row("Depreciation and amortisation expense", "", note_amounts.get("30"))
        add_data_row("Other expenses", "", note_amounts.get("31"))
        add_data_row("Total Expenses", "", expense_total, bold=True)
        
        # PROFIT BEFORE TAX
        add_data_row("Profit / (loss) before tax and exceptional items", "", profit_before_tax, bold=True)
        add_data_row("Profit / (loss) before tax", "", profit_before_tax, bold=True)
        
        # TAX EXPENSES SECTION
        add_data_row("Tax Expenses", section_header=True)
        
        # Get tax amounts (Note 32 = Current tax, Note 33 = Deferred tax)
        current_tax = note_amounts.get("32", 0)
        deferred_tax = note_amounts.get("33", 0)
        
        add_data_row("Current tax", "", current_tax)
        add_data_row("Deferred tax (Credit)/expense", "", deferred_tax)
        add_data_row("Total Tax expense", "", tax_total, bold=True)
        
        # NET PROFIT
        add_data_row("Profit / (loss) for the period", "", net_profit, bold=True)
        add_data_row("Total comprehensive income for the period", "", net_profit, bold=True)

        # Save workbook
        wb.save(output_file)
        
        return output_file