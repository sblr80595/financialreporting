# ============================================================================
# FILE: backend/services/pl_statement_service.py
# ============================================================================
"""Service for generating Profit & Loss statements from notes."""

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from backend.config.settings import settings
from backend.models.financial_statement import (
    PLGenerationResponse,
    PLLineItem,
    PLSection,
    ProfitLossStatement,
)

from backend.services.path_service import PathService
from backend.services.period_discovery_service import PeriodDiscoveryService
from backend.services.currency_service import CurrencyService
from backend.config.period_config import PeriodConfig

class PLStatementService:
    """Service for generating P&L statements from note markdown files."""

    # Entity-specific P&L structure configuration
    ENTITY_PL_CONFIGS = {
        # Default configuration for all entities
        "default": {
            "income_notes": ["24", "25"],
            "expense_notes": ["26", "27", "28", "29", "30", "31"],
            "tax_notes": ["32"],
        },
        # Add entity-specific overrides if needed
        # "XYZ_Company": {
        #     "income_notes": ["24", "25"],
        #     "expense_notes": ["26", "27", "28", "29", "30", "31"],
        #     "tax_notes": ["33"],  # Different tax note
        # },
    }

    # Note number to description mapping
    NOTE_DESCRIPTIONS = {
        "24": "Revenue from operations",
        "25": "Other income",
        "26": "Purchases of traded goods",
        "27": "Changes in inventories of traded goods",
        "28": "Employee benefits expense",
        "29": "Finance costs",
        "30": "Depreciation and amortisation expenses",
        "31": "Other expenses",
        "32": "Tax expense",
        "33": "Tax expense",  # Alternative tax note
    }

    @staticmethod
    def _get_entity_currency(company_name: str) -> dict:
        """
        Get currency information for an entity from the currency mapping config.
        
        Args:
            company_name: Name of the company/entity
            
        Returns:
            Dictionary with currency information (symbol, code, format)
        """
        try:
            currency = CurrencyService.get_entity_currency(company_name)
            return currency.model_dump()
        except Exception as e:
            # Default to INR on error
            return {
                "entity_name": company_name,
                "default_currency": "INR",
                "currency_symbol": "₹",
                "currency_name": "Indian Rupee",
                "decimal_places": 2,
                "format": "₹#,##,##0.00"
            }

    @staticmethod
    def _get_entity_config(company_name: str) -> Dict[str, List[str]]:
        """Get P&L configuration for a specific entity."""
        return PLStatementService.ENTITY_PL_CONFIGS.get(
            company_name, PLStatementService.ENTITY_PL_CONFIGS["default"]
        )

    @staticmethod
    def _find_latest_note_file(company_name: str, note_number: str) -> Optional[Path]:
        """
        Find the most recent generated note file.

        Args:
            company_name: Name of the company
            note_number: Note number to find

        Returns:
            Path to latest note file or None
        """
        # Use standard path: data/{entity}/output/generated_notes
        entity_notes_dir = settings.get_entity_generated_notes_dir(company_name)

        if not entity_notes_dir.exists():
            return None

        # Find all files for this note (case-insensitive for 'note' vs 'Note')
        # Try both patterns: note{number}_*.md and Note_{number}_*.md
        note_files = list(entity_notes_dir.glob(f"note{note_number}_*.md"))
        if not note_files:
            note_files = list(entity_notes_dir.glob(f"Note_{note_number}_*.md"))

        if not note_files:
            return None

        # Return the most recent file
        return max(note_files, key=lambda p: p.stat().st_mtime)

    @staticmethod
    def _extract_total_from_markdown(md_content: str) -> Optional[float]:
        """
        Extract total amount from markdown note content.

        Args:
            md_content: Markdown content of the note

        Returns:
            Total amount as float (always positive, sign handled by context)
        """
        # Updated patterns to support multiple currency symbols: ₹, ₱, $, RM, €, S$, etc.
        # Currency symbol is optional and can be: ₹ ₱ $ € or text like RM, S$, USD
        patterns = [
            # Pattern 1: Bold total with bold amount in parentheses (negative) - with currency
            r"\|\s*\*\*TOTAL[^|]*\*\*\s*\|\s*\*\*\([₹₱$€RM]*\s*([\d,]+(?:\.\d{2})?)\)\*\*",
            # Pattern 2: Bold total with bold amount (positive) - with currency
            r"\|\s*\*\*TOTAL[^|]*\*\*\s*\|\s*\*\*[₹₱$€RM]*\s*([\d,]+(?:\.\d{2})?)\*\*",
            # Pattern 3: Regular total with amount in parentheses (negative) - with currency
            r"\|\s*Total[^|]*\|\s*\([₹₱$€RM]*\s*([\d,]+(?:\.\d{2})?)\)",
            # Pattern 4: Regular total with amount (positive) - with currency
            r"\|\s*Total[^|]*\|\s*[₹₱$€RM]*\s*([\d,]+(?:\.\d{2})?)",
            # Pattern 5: Standalone bold total - with currency
            r"\*\*Total[^:]*:\*\*\s*\([₹₱$€RM]*\s*([\d,]+(?:\.\d{2})?)\)",
            r"\*\*Total[^:]*:\*\*\s*[₹₱$€RM]*\s*([\d,]+(?:\.\d{2})?)",
            # Pattern 6: Bold total without currency symbols
            r"\|\s*\*\*TOTAL[^|]*\*\*\s*\|\s*\*\*\(([\d,]+(?:\.\d{2})?)\)\*\*",
            r"\|\s*\*\*TOTAL[^|]*\*\*\s*\|\s*\*\*([\d,]+(?:\.\d{2})?)\*\*",
            # Pattern 7: Regular total without currency symbols
            r"\|\s*Total[^|]*\|\s*\(([\d,]+(?:\.\d{2})?)\)",
            r"\|\s*Total[^|]*\|\s*([\d,]+(?:\.\d{2})?)",
        ]

        for pattern in patterns:
            matches = re.findall(pattern, md_content, re.IGNORECASE)
            if matches:
                # Get the last match (usually the grand total)
                amount_str = matches[-1].replace(",", "")
                try:
                    amount = abs(float(amount_str))  # Always return positive
                    return amount
                except ValueError:
                    continue

        return None

    @staticmethod
    def _extract_note_details(
        company_name: str, note_number: str
    ) -> Tuple[Optional[float], Optional[str]]:
        """
        Extract amount and description from a note file.

        Args:
            company_name: Name of the company
            note_number: Note number

        Returns:
            Tuple of (amount, description)
        """
        note_file = PLStatementService._find_latest_note_file(company_name, note_number)

        if not note_file:
            return None, PLStatementService.NOTE_DESCRIPTIONS.get(note_number)

        try:
            with open(note_file, "r", encoding="utf-8") as f:
                content = f.read()

            amount = PLStatementService._extract_total_from_markdown(content)
            description = PLStatementService.NOTE_DESCRIPTIONS.get(note_number)

            return amount, description

        except Exception as e:
            print(f"Error reading note {note_number}: {e}")
            return None, PLStatementService.NOTE_DESCRIPTIONS.get(note_number)

    @staticmethod
    def check_pl_readiness(company_name: str) -> Dict:
        """
        Check if all required notes are available for P&L generation.

        Args:
            company_name: Name of the company

        Returns:
            Dictionary with readiness status and details
        """
        config = PLStatementService._get_entity_config(company_name)

        required_notes = (
            config["income_notes"] + config["expense_notes"] + config["tax_notes"]
        )

        found_notes = []
        missing_notes = []
        note_details = {}

        for note_num in required_notes:
            note_file = PLStatementService._find_latest_note_file(company_name, note_num)
            if note_file:
                found_notes.append(note_num)
                amount, description = PLStatementService._extract_note_details(
                    company_name, note_num
                )
                note_details[note_num] = {
                    "description": description,
                    "amount": amount,
                    "file_path": str(note_file),
                    "generated_at": datetime.fromtimestamp(
                        note_file.stat().st_mtime
                    ).isoformat(),
                }
            else:
                missing_notes.append(note_num)

        is_ready = len(missing_notes) == 0

        return {
            "company_name": company_name,
            "is_ready": is_ready,
            "found_notes": found_notes,
            "missing_notes": missing_notes,
            "note_details": note_details,
            "total_required": len(required_notes),
            "total_found": len(found_notes),
            "completeness_percentage": round(
                (len(found_notes) / len(required_notes)) * 100, 2
            ),
            "config": {
                "income_notes": config["income_notes"],
                "expense_notes": config["expense_notes"],
                "tax_notes": config["tax_notes"],
            },
        }

    @staticmethod
    def _build_section(
        section_name: str,
        note_list: List[str],
        note_amounts: Dict[str, float],
        company_name: str,
        total_label: Optional[str] = None,
    ) -> PLSection:
        """
        Build a single P&L section.

        Args:
            section_name: Name of the section
            note_list: List of note numbers
            note_amounts: Dictionary to store extracted amounts
            company_name: Company name
            total_label: Optional label for section total

        Returns:
            PLSection object
        """
        line_items = []
        section_total = 0.0

        for note_num in note_list:
            amount, description = PLStatementService._extract_note_details(
                company_name, note_num
            )

            if amount is not None:
                note_amounts[note_num] = amount
                section_total += amount

            line_items.append(
                PLLineItem(
                    particulars=description or f"Note {note_num}",
                    note=note_num,
                    amount=amount,
                    indent_level=0,
                )
            )

        # Add section total if label provided
        if total_label:
            line_items.append(
                PLLineItem(
                    particulars=total_label,
                    amount=section_total,
                    is_subtotal=True,
                    indent_level=0,
                )
            )

        return PLSection(section_name=section_name, line_items=line_items)

    @staticmethod
    def generate_pl_statement(
        company_name: str, period_ended: str, note_numbers: Optional[List[str]] = None
    ) -> PLGenerationResponse:
        """
        Generate complete P&L statement from note files.

        Args:
            company_name: Name of the company
            period_ended: Period ending date
            note_numbers: Optional (ignored, uses entity config)

        Returns:
            PLGenerationResponse with generated statement
        """
        try:
            # Check readiness first
            readiness = PLStatementService.check_pl_readiness(company_name)

            if not readiness["is_ready"]:
                return PLGenerationResponse(
                    success=False,
                    message=f"Cannot generate P&L statement. Missing notes: {', '.join(readiness['missing_notes'])}",
                )

            # Get entity configuration
            config = PLStatementService._get_entity_config(company_name)

            # Store extracted amounts
            note_amounts = {}
            sections = []

            # Build Income section
            income_section = PLStatementService._build_section(
                "I. Income",
                config["income_notes"],
                note_amounts,
                company_name,
                "Total income (I)",
            )
            sections.append(income_section)
            total_income = sum(
                item.amount or 0
                for item in income_section.line_items
                if not item.is_subtotal
            )

            # Build Expenses section
            expenses_section = PLStatementService._build_section(
                "II. Expenses",
                config["expense_notes"],
                note_amounts,
                company_name,
                "Total expenses (II)",
            )
            sections.append(expenses_section)
            total_expenses = sum(
                item.amount or 0
                for item in expenses_section.line_items
                if not item.is_subtotal
            )

            # Calculate Profit before tax
            profit_before_tax = total_income - total_expenses
            sections.append(
                PLSection(
                    section_name="III. Profit before tax (I-II)",
                    line_items=[
                        PLLineItem(
                            particulars="Profit before tax (I-II)",
                            amount=profit_before_tax,
                            is_total=True,
                            indent_level=0,
                        )
                    ],
                )
            )

            # Build Tax section
            tax_section = PLStatementService._build_section(
                "IV. Tax expense",
                config["tax_notes"],
                note_amounts,
                company_name,
                "Total tax expense (IV)",
            )
            sections.append(tax_section)
            total_tax = sum(
                item.amount or 0
                for item in tax_section.line_items
                if not item.is_subtotal
            )

            # Calculate Net Profit
            net_profit = profit_before_tax - total_tax
            sections.append(
                PLSection(
                    section_name="V. Net Profit after tax for the period (III-IV)",
                    line_items=[
                        PLLineItem(
                            particulars="Net Profit after tax for the period (III-IV)",
                            amount=net_profit,
                            is_total=True,
                            indent_level=0,
                        )
                    ],
                )
            )

            # Total comprehensive income
            sections.append(
                PLSection(
                    section_name="Total comprehensive income for the period",
                    line_items=[
                        PLLineItem(
                            particulars="Total comprehensive income for the period",
                            amount=net_profit,
                            is_total=True,
                            indent_level=0,
                        )
                    ],
                )
            )

            # Create statement
            statement = ProfitLossStatement(
                company_name=company_name,
                period_ended=period_ended,
                sections=sections,
                metadata={
                    "generated_at": datetime.now().isoformat(),
                    "notes_used": list(note_amounts.keys()),
                    "config_used": config,
                },
            )

            # Generate Excel file
            output_file = PLStatementService._export_to_excel(statement)

            return PLGenerationResponse(
                success=True,
                message=f"P&L statement generated successfully for {company_name}",
                statement=statement,
                output_file=str(output_file),
            )

        except Exception as e:
            return PLGenerationResponse(
                success=False, message=f"Error generating P&L statement: {str(e)}"
            )

    @staticmethod
    def _format_indian_number(value: float) -> str:
        """
        Format number in Indian numbering system without negative signs.

        Args:
            value: Number to format

        Returns:
            Formatted string (always positive)
        """
        value = abs(value)  # Always positive

        if value == 0:
            return "0"

        # Convert to string and split
        s = f"{value:,.2f}"
        parts = s.split(".")

        # Indian numbering system
        whole = parts[0].replace(",", "")
        if len(whole) > 3:
            last_three = whole[-3:]
            remaining = whole[:-3]

            # Add commas every 2 digits for Indian system
            result = ""
            while len(remaining) > 2:
                result = "," + remaining[-2:] + result
                remaining = remaining[:-2]

            if remaining:
                result = remaining + result

            formatted = result + "," + last_three
        else:
            formatted = whole

        # Remove decimals if .00
        if parts[1] == "00":
            return formatted

        return formatted + "." + parts[1]

    @staticmethod
    def _export_to_excel(statement: ProfitLossStatement) -> Path:
        """
        Export P&L statement to Excel file with formatting.

        Args:
            statement: ProfitLossStatement object

        Returns:
            Path to generated Excel file
        """
        # Create reports directory
        path_service = PathService(statement.company_name)
        pl_dir = path_service.get_financial_statements_dir(statement.company_name) / "PL"
        pl_dir.mkdir(parents=True, exist_ok=True)
        reports_dir = pl_dir

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = reports_dir / f"PL_Statement_{timestamp}.xlsx"

        # Get currency information for the entity
        currency_info = PLStatementService._get_entity_currency(statement.company_name)
        currency_symbol = currency_info['currency_symbol']
        
        print(f"[PL Statement] Using currency: {currency_symbol} ({currency_info['currency_name']}) for {statement.company_name}")

        # Prepare data
        data = []

        # Title rows
        data.append([statement.company_name.upper()])
        data.append(["STATEMENT OF PROFIT AND LOSS"])
        
        # Format period name: convert period key to MMM-YYYY format
        period_key = PeriodConfig.get_current_period()
        period_display = PeriodDiscoveryService.get_period_display_name(period_key) if period_key else statement.period_ended
        data.append([f"FOR THE PERIOD ENDED {period_display.upper()}"])
        data.append([])  # Empty row

        # Header row
        data.append(["Particulars", "Note", f"Amount ({currency_symbol})"])

        # Add sections
        for section in statement.sections:
            # Section might have header in first item
            for idx, item in enumerate(section.line_items):
                particulars = "  " * item.indent_level + item.particulars
                note = item.note if item.note else ""

                # Format amount
                amount_str = ""
                if item.amount is not None:
                    amount_str = PLStatementService._format_indian_number(item.amount)

                data.append([particulars, note, amount_str])

            # Add empty row after section
            data.append([])

        # Create DataFrame
        df = pd.DataFrame(data)

        # Export with styling
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="P&L Statement", index=False, header=False)

            # workbook = writer.book  # Reference to workbook (not used but available)
            worksheet = writer.sheets["P&L Statement"]

            # Apply formatting
            title_font = Font(name="Arial", size=14, bold=True)
            header_font = Font(name="Arial", size=11, bold=True)
            normal_font = Font(name="Arial", size=10)
            bold_font = Font(name="Arial", size=10, bold=True)

            header_fill = PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
            )

            # Format title rows (rows 1-3)
            for row in range(1, 4):
                cell = worksheet.cell(row, 1)
                cell.font = title_font
                cell.alignment = Alignment(horizontal="center")
                worksheet.merge_cells(f"A{row}:C{row}")

            # Format header row (row 5)
            for col in range(1, 4):
                cell = worksheet.cell(5, col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center" if col > 1 else "left")

            # Format data rows
            for row in range(6, len(data) + 1):
                for col in range(1, 4):
                    cell = worksheet.cell(row, col)
                    cell.font = normal_font

                    # Check if it's a bold row (total/subtotal)
                    cell_value = str(cell.value) if cell.value else ""
                    if (
                        "Total" in cell_value
                        or "TOTAL" in cell_value
                        or "Profit" in cell_value
                        or "comprehensive" in cell_value
                    ):
                        cell.font = bold_font

                    # Alignment
                    if col == 1:  # Particulars
                        cell.alignment = Alignment(horizontal="left")
                    elif col == 2:  # Note
                        cell.alignment = Alignment(horizontal="center")
                    else:  # Amount
                        cell.alignment = Alignment(horizontal="right")

            # Set column widths
            worksheet.column_dimensions["A"].width = 60
            worksheet.column_dimensions["B"].width = 10
            worksheet.column_dimensions["C"].width = 20

        return output_file
