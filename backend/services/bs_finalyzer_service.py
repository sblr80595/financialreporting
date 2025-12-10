# ============================================================================
# FILE: backend/services/bs_finalyzer_service.py
# ============================================================================
"""Service for generating BS Statement Finalyzer matching exact Excel template."""

import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Assuming imports from PNL service setup are correct
from backend.config.settings import settings
from backend.models.balance_sheet_models import (
    BSGenerationResponse,
    BalanceSheetStatement,
)
from backend.services.path_service import PathService


class BSFinalyzerService:
    """Service for generating BS Statement Finalyzer matching the exact Excel template."""

    ENTITY_BS_CONFIGS = {
        "default": {
            "non_current_assets": {
                "ppe": "3", "rou_assets": "4", "intangible": None,
                "investments": "5", "other_financial_nc": "7", "deferred_tax_assets": "6",
            },
            "current_assets": {
                "inventories": "8", "trade_receivables": "9", "cash": "10", 
                "bank_balances": "11", "loans": "12", "other_financial_c": "13", 
                "income_tax_assets": None, "other_current": "7",
            },
            "equity": {
                "share_capital": "14", "other_equity": "15",
            },
            "non_current_liabilities": {
                "lease_liability_nc": "18", "borrowings_nc": "17", "deferred_tax_liab": "6",
            },
            "current_liabilities": {
                "contract_liabilities": "16", "lease_liability_c": "18", "borrowings_c": "17",
                "trade_payables_msme": "19", "trade_payables_non_msme": "19",
                "other_financial_c": "20", "other_current_liab": "21",
                "provisions": "22", "current_tax_liab": "23",
            },
        },
    }

    NOTE_DESCRIPTIONS = {
        "3": "Property, plant and equipment", "4": "Right-of-use assets",
        "5": "Investments, Non current", "6": "Deferred tax liabilities (net)",
        "7": "Other financial assets/Other assets", "8": "Inventories",
        "9": "Trade receivables", "10": "Cash and cash equivalents",
        "11": "Bank balances other than cash and cash equivalents", "12": "Loans, Current",
        "13": "Other financial assets, Current", "14": "Share capital",
        "15": "Other equity", "16": "Contract Liabilities",
        "17": "Borrowings", "18": "Lease liability", "19": "Trade Payables",
        "20": "Other financial liabilities", "21": "Other Current liabilities",
        "22": "Provisions", "23": "Current tax liabilities (net)",
    }

    @staticmethod
    def _get_entity_config(company_name: str) -> Dict:
        """Get BS configuration for a specific entity."""
        return BSFinalyzerService.ENTITY_BS_CONFIGS.get(
            company_name, BSFinalyzerService.ENTITY_BS_CONFIGS["default"]
        )

    @staticmethod
    def _find_latest_note_file(company_name: str, note_number: str) -> Optional[Path]:
        """Find the most recent generated note file."""
        entity_notes_dir = settings.get_entity_generated_notes_dir(company_name)

        if not entity_notes_dir.exists():
            return None

        note_files = list(entity_notes_dir.glob(f"note{note_number}_*.md"))
        if not note_files:
            note_files = list(entity_notes_dir.glob(f"Note_{note_number}_*.md"))

        if not note_files:
            return None

        return max(note_files, key=lambda p: p.stat().st_mtime)

    @staticmethod
    def _extract_total_from_markdown(md_content: str) -> Optional[float]:
        """Extract total amount from markdown note content (Robust version)."""
        patterns = [
            r"\|\s*\*\*TOTAL[^|]*\*\*\s*\|\s*\*\*\(.*?([\d,]+\.?\d*)\)\*\*", 
            r"\|\s*\*\*TOTAL[^|]*\*\*\s*\|\s*\*\*.*?([\d,]+\.?\d*)\*\*",     
            r"\|\s*Total[^|]*\|\s*\(.*?([\d,]+\.?\d*)\)",                    
            r"\|\s*Total[^|]*\|\s*.*?([\d,]+\.?\d*)",                       
            r"\*\*Total[^:]*:\*\*\s*\(.*?([\d,]+\.?\d*)\)",                  
            r"\*\*Total[^:]*:\*\*\s*.*?([\d,]+\.?\d*)",                      
        ]

        for pattern in patterns:
            matches = re.findall(pattern, md_content, re.IGNORECASE)
            if matches:
                amount_str = matches[-1].replace(",", "")
                try:
                    amount = abs(float(amount_str))
                    return amount
                except ValueError:
                    continue
        
        simple_patterns = [r"Final Total\s*:\s*.*?([\d,]+\.?\d*)"]
        for pattern in simple_patterns:
            matches = re.findall(pattern, md_content, re.IGNORECASE)
            if matches:
                 amount_str = matches[-1].replace(",", "")
                 try:
                    amount = abs(float(amount_str))
                    return amount
                 except ValueError:
                    continue
        
        return None

    @staticmethod
    def _extract_note_details(
        company_name: str, note_number: str
    ) -> Tuple[Optional[float], Optional[str]]:
        """Extract amount and description from a note file by searching for the latest markdown note."""
        
        note_file = BSFinalyzerService._find_latest_note_file(company_name, note_number)
        description = BSFinalyzerService.NOTE_DESCRIPTIONS.get(note_number)

        if not note_file: return None, description

        try:
            with open(note_file, "r", encoding="utf-8") as f: content = f.read()
            amount = BSFinalyzerService._extract_total_from_markdown(content)
            return amount, description
        except Exception as e:
            print(f"Error reading note {note_number} file {note_file.name}: {e}")
            return None, description

    @staticmethod
    def _format_amount(value: Optional[float], show_zero: bool = True) -> str:
        """Format amount in Indian numbering system. (Function remains for legacy/metadata only)"""
        if value is None: return "" if not show_zero else "0"
        value = abs(value)
        if value == 0: return "0" if show_zero else ""

        s = f"{value:,.0f}"
        parts = s.replace(",", "")
        if len(parts) > 3:
            last_three = parts[-3:]
            remaining = parts[:-3]
            result = ""
            while len(remaining) > 2:
                result = "," + remaining[-2:] + result
                remaining = remaining[:-2]
            if remaining: result = remaining + result
            formatted = result + "," + last_three
        else: formatted = parts
        return formatted

    @staticmethod
    def generate_bs_finalyzer(
        company_name: str, period_label: str = "2025 Mar YTD",
        entity_info: str = "Entity: CPM  Book: 8937,IndAS - IndAS entities,Standalone,MYR,Actual",
        currency: str = "Malaysian Ringgit", scenario: str = "Actual",
    ) -> BSGenerationResponse:
        """Generate BS Statement Finalyzer matching exact Excel template."""
        try:
            config = BSFinalyzerService._get_entity_config(company_name); note_amounts = {}
            for section_config in config.values():
                if isinstance(section_config, dict):
                    for note_num in section_config.values():
                        if note_num:
                            amount, _ = BSFinalyzerService._extract_note_details(company_name, note_num)
                            if amount is not None: note_amounts[note_num] = amount
            
            # Recalculate totals in Python
            total_nca = sum(note_amounts.get(config["non_current_assets"][key], 0.0) for key in config["non_current_assets"].keys() if config["non_current_assets"][key])
            total_ca = sum(note_amounts.get(config["current_assets"][key], 0.0) for key in config["current_assets"].keys() if config["current_assets"][key])
            total_assets = total_nca + total_ca
            total_equity = sum(note_amounts.get(config["equity"][key], 0.0) for key in config["equity"].keys() if config["equity"][key])
            total_ncl = sum(note_amounts.get(config["non_current_liabilities"][key], 0.0) for key in config["non_current_liabilities"].keys() if config["non_current_liabilities"][key])
            total_cl = sum(note_amounts.get(config["current_liabilities"][key], 0.0) for key in config["current_liabilities"].keys() if config["current_liabilities"][key])
            total_liabilities = total_ncl + total_cl
            total_equity_liabilities = total_equity + total_liabilities

            output_file = BSFinalyzerService._export_to_excel_finalyzer(
                company_name=company_name, period_label=period_label, entity_info=entity_info, currency=currency, scenario=scenario,
                note_amounts=note_amounts, config=config,
                total_nca=total_nca, total_ca=total_ca, total_assets=total_assets, total_equity=total_equity,
                total_ncl=total_ncl, total_cl=total_cl, total_liabilities=total_liabilities, total_equity_liabilities=total_equity_liabilities,
            )

            statement = BalanceSheetStatement(
                company_name=company_name, as_at_date=period_label, sections=[],
                metadata={
                    "generated_at": datetime.now().isoformat(), 
                    "notes_used": list(set(note for section in config.values() for note in section.values() if note)), 
                    "type": "BS_Finalyzer",
                },
            )

            return BSGenerationResponse(success=True, message=f"BS Finalyzer generated successfully for {company_name}", statement=statement, output_file=str(output_file))

        except Exception as e:
            print(f"CRITICAL ERROR in generate_bs_finalyzer: {e}")
            return BSGenerationResponse(success=False, message=f"Error generating BS Finalyzer: {str(e)}")

    @staticmethod
    def _export_to_excel_finalyzer(
        company_name: str, period_label: str, entity_info: str, currency: str,
        scenario: str, note_amounts: Dict[str, float], config: Dict,
        total_nca: float, total_ca: float, total_assets: float, total_equity: float,
        total_ncl: float, total_cl: float, total_liabilities: float, total_equity_liabilities: float,
    ) -> Path:
        """Export BS Finalyzer to Excel with exact template formatting, inserting SUM formulas."""
        
        path_service = PathService(company_name); bs_dir = path_service.get_financial_statements_dir(company_name) / "BS_Finalyzer"
        bs_dir.mkdir(parents=True, exist_ok=True); timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = bs_dir / f"BS_Finalyzer_{timestamp}.xlsx"; wb = Workbook(); ws = wb.active
        ws.title = "Balance Sheet"

        # Define Styles
        company_font = Font(name="Calibri", size=11, color="C00000") 
        period_font = Font(name="Calibri", size=11, bold=True)
        entity_font = Font(name="Calibri", size=9)
        header_font = Font(name="Calibri", size=11, bold=True)
        section_header_font = Font(name="Calibri", size=11, bold=True, color="C00000")
        total_orange_font = Font(name="Calibri", size=11, bold=True, color="C65911")
        normal_font = Font(name="Calibri", size=11)
        bold_font = Font(name="Calibri", size=11, bold=True)
        
        # Indian Number Format for Excel (key to resolving the 0 sum issue and comma display)
        indian_number_format = r'[>9999999]##\,##\,##\,##0;[>99999]##\,##\,##0;##,##0'

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        light_gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        ws.column_dimensions["A"].width = 65; ws.column_dimensions["B"].width = 15; ws.column_dimensions["C"].width = 18

        current_row = 1

        # --- Header Rows 1-11 ---
        cell = ws.cell(row=current_row, column=1); cell.value = company_name; cell.font = company_font; current_row += 1
        cell = ws.cell(row=current_row, column=1); cell.value = period_label; cell.font = period_font; current_row += 1
        current_row += 1
        cell = ws.cell(row=current_row, column=1); cell.value = entity_info; cell.font = entity_font; current_row += 1
        cell = ws.cell(row=current_row, column=1); cell.value = "BS"; cell.font = bold_font; cell.fill = yellow_fill; ws.merge_cells(f"A{current_row}:C{current_row}"); current_row += 1
        cell = ws.cell(row=current_row, column=1); cell.value = "RM in Lakhs"; cell.font = entity_font; current_row += 1
        headers = ["", "Entity", ""]; 
        for col_idx, header in enumerate(headers, start=1): cell = ws.cell(row=current_row, column=col_idx); cell.value = header; cell.fill = light_gray_fill; cell.border = thin_border
        current_row += 1
        cell = ws.cell(row=current_row, column=2); cell.value = "Period Id"; cell.fill = light_gray_fill; cell.border = thin_border
        cell = ws.cell(row=current_row, column=3); cell.value = period_label.split()[1].upper() + " " + period_label.split()[0]; cell.fill = light_gray_fill; cell.border = thin_border; cell.font = bold_font; current_row += 1
        cell = ws.cell(row=current_row, column=2); cell.value = "Currency"; cell.fill = light_gray_fill; cell.border = thin_border
        cell = ws.cell(row=current_row, column=3); cell.value = currency; cell.fill = light_gray_fill; cell.border = thin_border; cell.font = bold_font; current_row += 1
        cell = ws.cell(row=current_row, column=2); cell.value = "Scenario"; cell.fill = light_gray_fill; cell.border = thin_border
        cell = ws.cell(row=current_row, column=3); cell.value = scenario; cell.fill = light_gray_fill; cell.border = thin_border; cell.font = bold_font; current_row += 1
        for col_idx, header in enumerate(["Item Label / Nature Of Report", "Consol Code", "Standalone"], start=1):
            cell = ws.cell(row=current_row, column=col_idx); cell.value = header; cell.fill = light_gray_fill; cell.border = thin_border; cell.font = header_font
        current_row += 1

        # Helper function to add data row (Lookups by Note Number)
        def add_data_row(label: str, note_num: Optional[str] = None, formula: Optional[str] = None, is_main_heading: bool = False, is_total_row: bool = False, section_header: bool = False):
            nonlocal current_row
            
            amount = note_amounts.get(note_num) if note_num else None
            
            # Column A: Label
            cell = ws.cell(row=current_row, column=1)
            cell.value = label
            cell.border = thin_border
            if is_main_heading: cell.font = total_orange_font
            elif section_header: cell.font = section_header_font
            elif is_total_row: cell.font = bold_font
            else: cell.font = normal_font
            
            # Column B: Consol Code (Note Number)
            cell = ws.cell(row=current_row, column=2)
            cell.value = note_num if note_num else ""
            cell.border = thin_border
            cell.font = normal_font
            cell.alignment = Alignment(horizontal="left")
            
            # Column C: Amount/Formula (FIXED for both 0 issue and comma display)
            cell = ws.cell(row=current_row, column=3)
            
            if formula:
                cell.value = formula  # Insert formula string
                cell.data_type = 'f' # Set data type to formula
                cell.number_format = indian_number_format # Apply formatting to formula result
                if is_main_heading: cell.font = total_orange_font
                elif is_total_row: cell.font = bold_font
            elif amount is not None:
                cell.value = amount # Insert raw float value (Required for SUM function)
                cell.number_format = indian_number_format # Apply formatting style
            
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right")
            
            current_row += 1
            return current_row - 1 # Return the row number of the inserted data

        # Define Column C constant
        COL_C = 'C'
        
        # --- BODY CONTENT ---
        
        # Balance Sheet Net (Placeholder Row for Final Calculation)
        add_data_row("Balance sheet", is_main_heading=True)
        row_bs_net = add_data_row("Balance sheet net", formula="", is_total_row=True, is_main_heading=True) 
        
        # ASSETS SECTION (Main Heading)
        add_data_row("Assets", is_main_heading=True)
        
        # Non-Current Assets (Subheading)
        add_data_row("Non-Current assets", section_header=True)
        nca_sum_start_row = current_row
        
        add_data_row("Property, plant and equipment", config["non_current_assets"].get("ppe"))
        add_data_row("Right-of-use assets", config["non_current_assets"].get("rou_assets"))
        add_data_row("Other intangible assets", config["non_current_assets"].get("intangible"))
        
        add_data_row("Financial Assets", section_header=True)
        add_data_row("Investments, Non current", config["non_current_assets"].get("investments"))
        add_data_row("Other financial assets, Non current", config["non_current_assets"].get("other_financial_nc"))
        add_data_row("Deferred tax assets (net)", config["non_current_assets"].get("deferred_tax_assets"))
        nca_sum_end_row = current_row - 1
        
        # Total Non Current Assets
        row_nca_total = add_data_row(
            "Total Non Current Assets", formula=f"=SUM({COL_C}{nca_sum_start_row}:{COL_C}{nca_sum_end_row})", is_total_row=True
        )
        current_row += 1 # Empty row

        # Current Assets (Subheading)
        add_data_row("Current Assets", section_header=True)
        ca_sum_start_row = current_row
        
        add_data_row("Inventories", config["current_assets"].get("inventories"))
        add_data_row("Financial Assets", section_header=True)
        add_data_row("Trade receivables", config["current_assets"].get("trade_receivables"))
        add_data_row("Cash and cash equivalents", config["current_assets"].get("cash"))
        add_data_row("Bank balances other than cash and cash equivalents", config["current_assets"].get("bank_balances"))
        add_data_row("Loans, Current", config["current_assets"].get("loans"))
        add_data_row("Other financial assets, Current", config["current_assets"].get("other_financial_c"))
        
        add_data_row("Income tax assets, Current", config["current_assets"].get("income_tax_assets")) 
        add_data_row("Other Current assets", config["current_assets"].get("other_current"))
        ca_sum_end_row = current_row - 1
        
        # Total Current Assets
        row_ca_total = add_data_row(
            "Total Current assets", formula=f"=SUM({COL_C}{ca_sum_start_row}:{COL_C}{ca_sum_end_row})", is_total_row=True
        )
        
        # Total Assets
        row_total_assets = add_data_row(
            "Total Assets", formula=f"={COL_C}{row_nca_total}+{COL_C}{row_ca_total}", is_total_row=True, is_main_heading=True
        )
        current_row += 1
        current_row += 1
        
        # EQUITY AND LIABILITIES SECTION (Main Heading)
        add_data_row("Equity And Liabilities", is_main_heading=True)

        # Equity (Subheading)
        add_data_row("Equity", section_header=True)
        eq_sum_start_row = current_row
        add_data_row("Share capital", config["equity"].get("share_capital"))
        add_data_row("Other equity", config["equity"].get("other_equity"))
        eq_sum_end_row = current_row - 1
        
        # Total Equity
        row_equity_total = add_data_row(
            "Total Equity", formula=f"=SUM({COL_C}{eq_sum_start_row}:{COL_C}{eq_sum_end_row})", is_total_row=True
        )
        current_row += 1
        
        # Liabilities (Subheading)
        add_data_row("Liabilities", section_header=True)
        
        # Non-Current Liabilities (Subheading)
        add_data_row("Non Current liabilities", section_header=True)
        ncl_sum_start_row = current_row
        add_data_row("Financial liabilities", section_header=True)
        add_data_row("Lease liability", config["non_current_liabilities"].get("lease_liability_nc"))
        add_data_row("Borrowings", config["non_current_liabilities"].get("borrowings_nc"))
        add_data_row("Deferred tax liabilities", config["non_current_liabilities"].get("deferred_tax_liab"))
        ncl_sum_end_row = current_row - 1
        
        # Total Non current liabilities
        row_ncl_total = add_data_row(
            "Total Non current liabilities", formula=f"=SUM({COL_C}{ncl_sum_start_row}:{COL_C}{ncl_sum_end_row})", is_total_row=True
        )
        current_row += 1

        # Current Liabilities (Subheading)
        add_data_row("Current Liabilities", section_header=True)
        cl_sum_start_row = current_row
        add_data_row("Contract Liabilities", config["current_liabilities"].get("contract_liabilities"))
        add_data_row("Financial liabilities", section_header=True)
        add_data_row("Lease liability", config["current_liabilities"].get("lease_liability_c"))
        add_data_row("Borrowings", config["current_liabilities"].get("borrowings_c"))
        
        add_data_row("Trade Payables", section_header=True)
        add_data_row("Trade payables - Non MSME", config["current_liabilities"].get("trade_payables_non_msme"))
        add_data_row("Other financial liabilities", config["current_liabilities"].get("other_financial_c"))
        
        add_data_row("Other Current liabilities", config["current_liabilities"].get("other_current_liab"))
        add_data_row("Provisions", config["current_liabilities"].get("provisions"))
        add_data_row("Current tax liabilities (net)", config["current_liabilities"].get("current_tax_liab"))
        cl_sum_end_row = current_row - 1
        
        # Total Current liabilities
        row_cl_total = add_data_row(
            "Total Current liabilities", formula=f"=SUM({COL_C}{cl_sum_start_row}:{COL_C}{cl_sum_end_row})", is_total_row=True
        )
        
        # Total Equity and Liabilities
        row_total_equity_liabilities = add_data_row(
            "Total Equity and liabilities", formula=f"={COL_C}{row_equity_total}+{COL_C}{row_ncl_total}+{COL_C}{row_cl_total}", is_total_row=True, is_main_heading=True
        )
        current_row += 1 
        
        # FINAL Balance Sheet Net (Assets - E&L)
        ws.cell(row=row_bs_net, column=3, value=f"={COL_C}{row_total_assets}-{COL_C}{row_total_equity_liabilities}")
        ws.cell(row=row_bs_net, column=3).data_type = 'f'
        ws.cell(row=row_bs_net, column=3).number_format = indian_number_format
        
        # Save workbook
        wb.save(output_file)
        
        return output_file