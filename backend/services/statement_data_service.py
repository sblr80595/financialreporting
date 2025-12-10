"""
Service for extracting statement data from generated Excel files.
Maps Excel rows to template row IDs for the viewer.
"""

from pathlib import Path
from typing import Dict, Optional, Any, List
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from backend.services.path_service import PathService
import re


class StatementDataService:
    """Extract data from generated statement Excel files for viewer."""

    @staticmethod
    def get_pl_statement_data(company_name: str) -> Optional[Dict[str, Any]]:
        """
        Extract P&L statement data from latest Excel file.

        Args:
            company_name: Name of the company

        Returns:
            Dict with row_id -> {current, previous} mapping, or None if no file
        """
        path_service = PathService(company_name)
        pl_dir = path_service.get_financial_statements_dir(company_name) / "PL"

        print(f"[StatementData] Looking for PL files in: {pl_dir}")
        
        if not pl_dir.exists():
            print(f"[StatementData] PL directory does not exist: {pl_dir}")
            return None

        # Find most recent file
        pl_files = list(pl_dir.glob("PL_Statement_*.xlsx"))
        print(f"[StatementData] Found {len(pl_files)} PL files")
        
        if not pl_files:
            return None

        latest_file = max(pl_files, key=lambda p: p.stat().st_mtime)
        print(f"[StatementData] Using latest PL file: {latest_file.name}")

        return StatementDataService._extract_pl_data(latest_file)

    @staticmethod
    def get_bs_statement_data(company_name: str) -> Optional[Dict[str, Any]]:
        """
        Extract Balance Sheet data from latest Excel file.

        Args:
            company_name: Name of the company

        Returns:
            Dict with row_id -> {current, previous} mapping, or None if no file
        """
        path_service = PathService(company_name)
        bs_dir = path_service.get_financial_statements_dir(company_name) / "BS"

        print(f"[StatementData] Looking for BS files in: {bs_dir}")
        
        if not bs_dir.exists():
            print(f"[StatementData] BS directory does not exist: {bs_dir}")
            return None

        # Find most recent file - check both naming patterns
        bs_files = list(bs_dir.glob("BS_Statement_*.xlsx"))
        if not bs_files:
            bs_files = list(bs_dir.glob("BalanceSheet_*.xlsx"))
        
        print(f"[StatementData] Found {len(bs_files)} BS files")
        
        if not bs_files:
            return None

        latest_file = max(bs_files, key=lambda p: p.stat().st_mtime)
        print(f"[StatementData] Using latest BS file: {latest_file.name}")

        return StatementDataService._extract_bs_data(latest_file)

    @staticmethod
    def get_cf_statement_data(company_name: str) -> Optional[Dict[str, Any]]:
        """
        Extract Cash Flow statement data from latest Excel file.

        Args:
            company_name: Name of the company

        Returns:
            Dict with row_id -> {current, previous} mapping, or None if no file
        """
        path_service = PathService(company_name)
        cf_dir = path_service.get_financial_statements_dir(company_name) / "CashFlow"

        print(f"[StatementData] Looking for CF files in: {cf_dir}")
        
        if not cf_dir.exists():
            print(f"[StatementData] CF directory does not exist: {cf_dir}")
            return None

        # Find most recent file - check multiple patterns
        cf_files = list(cf_dir.glob("CashFlow_Statement_*.xlsx"))
        if not cf_files:
            cf_files = list(cf_dir.glob("CashFlow_*.xlsx"))
        
        print(f"[StatementData] Found {len(cf_files)} CF files")
        
        if not cf_files:
            return None

        latest_file = max(cf_files, key=lambda p: p.stat().st_mtime)
        print(f"[StatementData] Using latest CF file: {latest_file.name}")

        return StatementDataService._extract_cf_data(latest_file)

    @staticmethod
    def _extract_pl_data(file_path: Path) -> Dict[str, Any]:
        """
        Parse P&L Excel file and extract row data.

        Expected structure:
        - Row 1: Company name
        - Row 2: "STATEMENT OF PROFIT AND LOSS"
        - Row 3: Period
        - Row 4: Empty
        - Row 5: Headers ["Particulars", "Note", "Amount (₹)"]
        - Row 6+: Data rows

        Returns:
            {
                "revenue": {"current": 1000000, "previous": null},
                "operating_income": {"current": 950000, "previous": null},
                ...
            }
        """
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        print(f"[StatementData] Parsing PL Excel: {file_path.name}")
        print(f"[StatementData] Excel has {ws.max_row} rows")
        
        data = {}
        row_mapping = StatementDataService._get_pl_row_mapping()
        matched_count = 0
        unmatched_rows = []

        # Start reading from row 6 (after headers)
        for row_idx in range(6, ws.max_row + 1):
            particulars = ws.cell(row_idx, 1).value  # Column A
            note = ws.cell(row_idx, 2).value  # Column B
            amount_str = ws.cell(row_idx, 3).value  # Column C

            if not particulars or particulars.strip() == "":
                continue

            # Clean particulars (remove indentation)
            clean_particulars = particulars.strip()

            # Parse amount
            amount = StatementDataService._parse_amount(amount_str)

            # Try to match to template row
            row_id = StatementDataService._match_pl_row(clean_particulars, note, row_mapping)

            if row_id:
                data[row_id] = {
                    "current": amount,
                    "previous": None  # Generated files don't have previous period yet
                }
                matched_count += 1
            else:
                unmatched_rows.append(clean_particulars)

        print(f"[StatementData] PL: Matched {matched_count} rows, {len(unmatched_rows)} unmatched")
        if unmatched_rows and len(unmatched_rows) <= 10:
            print(f"[StatementData] Unmatched rows: {unmatched_rows}")
        
        wb.close()
        return data

    @staticmethod
    def _extract_bs_data(file_path: Path) -> Dict[str, Any]:
        """
        Parse Balance Sheet Excel file and extract row data.

        Expected structure:
        - Row 1: Company name
        - Row 2: "BALANCE SHEET"
        - Row 3: "AS AT {DATE}"
        - Row 4: Empty
        - Row 5: Headers ["Particulars", "Note", "Amount (₹)"]
        - Row 6+: Data rows

        Returns:
            {
                "assets_total": {"current": 5000000, "previous": null},
                "non_current_assets": {"current": 3000000, "previous": null},
                ...
            }
        """
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        print(f"[StatementData] Parsing BS Excel: {file_path.name}")
        print(f"[StatementData] Excel has {ws.max_row} rows")
        
        data = {}
        row_mapping = StatementDataService._get_bs_row_mapping()
        matched_count = 0
        unmatched_rows = []

        # Start reading from row 6 (after headers)
        for row_idx in range(6, ws.max_row + 1):
            particulars = ws.cell(row_idx, 1).value  # Column A
            note = ws.cell(row_idx, 2).value  # Column B
            amount_str = ws.cell(row_idx, 3).value  # Column C

            if not particulars or particulars.strip() == "":
                continue

            # Clean particulars
            clean_particulars = particulars.strip()

            # Parse amount
            amount = StatementDataService._parse_amount(amount_str)

            # Try to match to template row
            row_id = StatementDataService._match_bs_row(clean_particulars, note, row_mapping)

            if row_id:
                data[row_id] = {
                    "current": amount,
                    "previous": None
                }
                matched_count += 1
            else:
                unmatched_rows.append(clean_particulars)

        print(f"[StatementData] PL: Matched {matched_count} rows, {len(unmatched_rows)} unmatched")
        if unmatched_rows and len(unmatched_rows) <= 10:
            print(f"[StatementData] Unmatched rows: {unmatched_rows}")
        
        wb.close()
        return data

    @staticmethod
    def _extract_cf_data(file_path: Path) -> Dict[str, Any]:
        """
        Parse Cash Flow Excel file and extract row data.

        Returns:
            {
                "net_cash_operating": {"current": 500000, "previous": null},
                "net_cash_investing": {"current": -200000, "previous": null},
                ...
            }
        """
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        data = {}
        row_mapping = StatementDataService._get_cf_row_mapping()

        # Start reading from row 6 (after headers)
        for row_idx in range(6, ws.max_row + 1):
            particulars = ws.cell(row_idx, 1).value  # Column A
            note = ws.cell(row_idx, 2).value  # Column B
            amount_str = ws.cell(row_idx, 3).value  # Column C

            if not particulars or particulars.strip() == "":
                continue

            # Clean particulars
            clean_particulars = particulars.strip()

            # Parse amount
            amount = StatementDataService._parse_amount(amount_str)

            # Try to match to template row
            row_id = StatementDataService._match_cf_row(clean_particulars, note, row_mapping)

            if row_id:
                data[row_id] = {
                    "current": amount,
                    "previous": None
                }

        wb.close()
        return data

    @staticmethod
    def _parse_amount(amount_str: Any) -> Optional[float]:
        """
        Parse amount string to float, handling Indian number format.

        Examples:
            "1,23,45,678.50" -> 12345678.50
            "12,345.00" -> 12345.00
            "(1,234.50)" -> -1234.50

        Args:
            amount_str: Amount string from Excel

        Returns:
            Float value or None
        """
        if amount_str is None or amount_str == "" or amount_str == "-":
            return None

        # If already a number
        if isinstance(amount_str, (int, float)):
            return float(amount_str)

        # Convert to string
        amount_str = str(amount_str).strip()

        if amount_str == "" or amount_str == "-":
            return None

        # Check for negative (parentheses)
        is_negative = amount_str.startswith("(") and amount_str.endswith(")")
        if is_negative:
            amount_str = amount_str[1:-1]

        # Remove commas
        amount_str = amount_str.replace(",", "")

        # Try to convert
        try:
            value = float(amount_str)
            return -value if is_negative else value
        except ValueError:
            return None

    @staticmethod
    def _get_pl_row_mapping() -> Dict[str, List[str]]:
        """
        Get mapping of P&L row IDs to possible text variations.
        
        Row IDs match the frontend template (pl_indas.json, pl_ifrs.json)

        Returns:
            Dict mapping row_id to list of possible text matches
        """
        return {
            # Income section
            "REV_OPS": ["Revenue from operations", "Sales", "Revenue"],
            "OTHER_INC": ["Other income", "Other Income"],
            "TOTAL_INCOME": ["Total income (I)", "Total Income", "Total Revenue"],

            # Expenses section
            "MAT_CONS": ["Cost of materials consumed", "Cost of Materials"],
            "STOCK_IN_TRADE": [
                "Purchases of stock-in-trade",
                "Purchases of Stock-in-Trade",
            ],
            "CHG_INVENTORY": [
                "Changes in inventories of finished goods, work-in-progress and stock-in-trade",
                "Changes in Inventories",
                "Change in Inventory",
            ],
            "EMP_BEN": [
                "Employee benefits expense",
                "Employee Benefits",
                "Employee Benefit Expense",
            ],
            "FIN_COST": ["Finance costs", "Finance Costs", "Interest Expense"],
            "DEP_AMORT": [
                "Depreciation and amortisation expense",
                "Depreciation and Amortization",
                "Depreciation",
            ],
            "OTHER_EXP": ["Other expenses", "Other Expenses"],
            "TOTAL_EXPENSES": ["Total expenses (II)", "Total Expenses"],

            # Profit lines
            "PBT_BEFORE_EXC": [
                "Profit before exceptional items and tax (I - II)",
                "Profit before exceptional items and tax",
            ],
            "EXCEPTIONAL": ["Exceptional items", "Exceptional Items"],
            "PBT": [
                "Profit before tax (I-II)",
                "Profit before tax (III)",
                "Profit Before Tax",
                "PBT",
            ],

            # Tax section
            "CURRENT_TAX": ["Current tax", "Current Tax"],
            "DEFERRED_TAX": ["Deferred tax", "Deferred Tax"],
            "TAX_ADJ": ["Tax adjustments for earlier periods", "Tax Adjustments"],
            "TOTAL_TAX": ["Total tax expense (IV)", "Total tax expense", "Total Tax", "Tax Expense"],

            # Net profit
            "PROFIT_FOR_YEAR": [
                "Net Profit after tax for the period (III-IV)",
                "Profit for the year (IV)",
                "Net Profit",
                "Profit After Tax",
                "PAT",
            ],
            
            # Other comprehensive income
            "OCI_TAX": ["Income tax relating to items of OCI", "Income tax relating to OCI"],
            "TOTAL_OCI": ["Total other comprehensive income", "Total OCI"],
            "TOTAL_CI": ["Total comprehensive income for the year (V)", "Total comprehensive income"],
            
            # EPS
            "EPS_BASIC": ["Earnings per equity share - Basic", "EPS - Basic"],
            "EPS_DILUTED": ["Earnings per equity share - Diluted", "EPS - Diluted"],
        }

    @staticmethod
    def _get_bs_row_mapping() -> Dict[str, List[str]]:
        """
        Get mapping of Balance Sheet row IDs to possible text variations.
        
        Row IDs match the frontend template (bs_indas.json)

        Returns:
            Dict mapping row_id to list of possible text matches
        """
        return {
            # Main sections
            "ASSETS_SECTION": ["ASSETS"],
            "EQUITY_LIAB_SECTION": ["EQUITY AND LIABILITIES"],
            
            # Assets
            "TOTAL_ASSETS": ["TOTAL ASSETS", "Total Assets"],
            
            # Non-current assets
            "PPE": [
                "Property, plant and equipment",
                "Property, Plant and Equipment",
                "Fixed Assets",
            ],
            "CWIP": [
                "Capital work-in-progress",
                "Capital Work-in-Progress",
                "CWIP",
            ],
            "INV_PROP": ["Investment property", "Investment Property"],
            "GOODWILL": ["Goodwill"],
            "INTANGIBLES": [
                "Other intangible assets",
                "Other Intangible Assets",
                "Intangible Assets",
            ],
            "INTANGIBLES_UNDER_DEV": [
                "Intangible assets under development",
                "Intangible Assets under Development",
            ],
            "ROU_ASSETS": ["Right-of-use assets", "ROU Assets"],
            "INVESTMENTS_NC": ["Investments", "(i) Investments"],
            "TRADE_REC_NC": ["Trade receivables", "(i) Trade receivables"],
            "LOANS_NC": ["Loans", "(ii) Loans", "(i) Loans"],
            "OTHER_FIN_NC": [
                "Others",
                "Other financial assets",
                "(iii) Other financial assets",
                "(ii) Others",
            ],
            "DEF_TAX_ASSETS": [
                "Deferred tax assets (net)",
                "Deferred Tax Assets",
            ],
            "OTHER_NC_ASSETS": [
                "Other non-current assets",
                "Other Non-Current Assets",
            ],

            # Current assets
            "INVENTORIES": ["Inventories"],
            "INVESTMENTS_C": ["Investments", "(i) Investments"],
            "TRADE_REC_C": ["Trade receivables", "(ii) Trade receivables"],
            "CASH_EQ": [
                "Cash and cash equivalents",
                "(iii) Cash and cash equivalents",
            ],
            "BANK_BAL": [
                "Bank balances other than above",
                "(iv) Bank balances other than above",
            ],
            "LOANS_C": ["Loans", "(v) Loans"],
            "OTHER_FIN_C": [
                "Others",
                "Other financial assets",
                "(vi) Other financial assets",
            ],
            "CURR_TAX_ASSETS": [
                "Current tax assets (net)",
                "Current Tax Assets",
            ],
            "OTHER_C_ASSETS": ["Other current assets", "Other Current Assets"],

            # Equity
            "SHARE_CAPITAL": [
                "Equity share capital",
                "Share Capital",
                "Equity Share Capital",
            ],
            "OTHER_EQUITY": ["Other equity", "Other Equity", "Reserves and Surplus"],

            # Non-current liabilities
            "BORROWINGS_NC": ["Borrowings", "(i) Borrowings"],
            "LEASE_LIAB_NC": ["Lease liabilities", "(i) Lease liabilities"],
            "OTHER_FIN_LIAB_NC": [
                "Others",
                "Other financial liabilities",
                "(ii) Other financial liabilities",
                "(iii) Others",
            ],
            "PROVISIONS_NC": ["Provisions", "Non-Current Provisions"],
            "DEF_TAX_LIAB": [
                "Deferred tax liabilities (net)",
                "Deferred Tax Liabilities",
            ],
            "OTHER_NC_LIAB": [
                "Other non-current liabilities",
                "Other Non-Current Liabilities",
            ],

            # Current liabilities
            "BORROWINGS_C": ["Borrowings", "(i) Borrowings"],
            "TRADE_PAY_MSME": [
                "Trade payables - MSME",
                "Trade Payables - MSME",
            ],
            "TRADE_PAY_OTHER": [
                "Trade payables - Others",
                "Trade Payables - Others",
                "Trade payables",
                "(ii) Trade payables",
            ],
            "LEASE_LIAB_C": ["Lease liabilities", "(ii) Lease liabilities"],
            "OTHER_FIN_LIAB_C": [
                "Others",
                "Other financial liabilities",
                "(iii) Other financial liabilities",
            ],
            "OTHER_CURR_LIAB": [
                "Other current liabilities",
                "Other Current Liabilities",
            ],
            "PROVISIONS_C": ["Provisions", "Current Provisions"],
            "CURR_TAX_LIAB": [
                "Current tax liabilities (net)",
                "Current Tax Liabilities",
            ],
            
            # Totals
            "TOTAL_EQUITY_LIAB": [
                "TOTAL EQUITY AND LIABILITIES",
                "Total Equity and Liabilities",
            ],
        }

    @staticmethod
    def _get_cf_row_mapping() -> Dict[str, List[str]]:
        """
        Get mapping of Cash Flow row IDs to possible text variations.

        Returns:
            Dict mapping row_id to list of possible text matches
        """
        return {
            # Operating activities
            "profit_before_tax_cf": ["Profit before tax", "PBT"],
            "depreciation_amortization": [
                "Depreciation and amortisation",
                "Depreciation and Amortization",
            ],
            "interest_income": ["Interest income", "Interest Income"],
            "interest_expense": ["Interest expense", "Interest Expense"],
            "operating_profit": [
                "Operating profit before working capital changes",
                "Operating Profit",
            ],
            "trade_receivables_change": [
                "Trade receivables",
                "Change in Trade Receivables",
            ],
            "inventories_change": ["Inventories", "Change in Inventories"],
            "trade_payables_change": ["Trade payables", "Change in Trade Payables"],
            "cash_from_operations": [
                "Cash generated from operations",
                "Cash from Operations",
            ],
            "taxes_paid": ["Income taxes paid", "Taxes Paid"],
            "net_cash_operating": [
                "Net cash from operating activities",
                "Net Cash from Operating Activities",
            ],

            # Investing activities
            "ppe_purchase": [
                "Purchase of property, plant and equipment",
                "Purchase of PPE",
                "Capital Expenditure",
            ],
            "ppe_sale": [
                "Sale of property, plant and equipment",
                "Sale of PPE",
            ],
            "investments_purchase": ["Purchase of investments", "Investments Made"],
            "investments_sale": ["Sale of investments", "Investments Sold"],
            "interest_received": ["Interest received", "Interest Received"],
            "net_cash_investing": [
                "Net cash used in investing activities",
                "Net Cash from Investing Activities",
            ],

            # Financing activities
            "proceeds_borrowings": [
                "Proceeds from borrowings",
                "Borrowings Raised",
            ],
            "repayment_borrowings": [
                "Repayment of borrowings",
                "Borrowings Repaid",
            ],
            "interest_paid": ["Interest paid", "Interest Paid"],
            "dividends_paid": ["Dividends paid", "Dividends Paid"],
            "net_cash_financing": [
                "Net cash from financing activities",
                "Net Cash from Financing Activities",
            ],

            # Summary
            "net_change_cash": [
                "Net increase/(decrease) in cash and cash equivalents",
                "Net Change in Cash",
            ],
            "cash_beginning": [
                "Cash and cash equivalents at the beginning of the period",
                "Opening Cash Balance",
            ],
            "cash_ending": [
                "Cash and cash equivalents at the end of the period",
                "Closing Cash Balance",
            ],
        }

    @staticmethod
    def _match_pl_row(
        particulars: str, note: Optional[str], mapping: Dict[str, List[str]]
    ) -> Optional[str]:
        """
        Match P&L particulars text to template row ID.

        Args:
            particulars: The particulars text from Excel
            note: The note reference (if any)
            mapping: Row mapping dictionary

        Returns:
            Matched row_id or None
        """
        particulars_lower = particulars.lower()

        for row_id, variations in mapping.items():
            for variation in variations:
                if variation.lower() == particulars_lower:
                    return row_id

        return None

    @staticmethod
    def _match_bs_row(
        particulars: str, note: Optional[str], mapping: Dict[str, List[str]]
    ) -> Optional[str]:
        """
        Match Balance Sheet particulars text to template row ID.

        Args:
            particulars: The particulars text from Excel
            note: The note reference (if any)
            mapping: Row mapping dictionary

        Returns:
            Matched row_id or None
        """
        particulars_lower = particulars.lower()

        for row_id, variations in mapping.items():
            for variation in variations:
                if variation.lower() == particulars_lower:
                    return row_id

        return None

    @staticmethod
    def _match_cf_row(
        particulars: str, note: Optional[str], mapping: Dict[str, List[str]]
    ) -> Optional[str]:
        """
        Match Cash Flow particulars text to template row ID.

        Args:
            particulars: The particulars text from Excel
            note: The note reference (if any)
            mapping: Row mapping dictionary

        Returns:
            Matched row_id or None
        """
        particulars_lower = particulars.lower()

        for row_id, variations in mapping.items():
            for variation in variations:
                if variation.lower() == particulars_lower:
                    return row_id

        return None
