# ============================================================================
# FILE: backend/services/cashflow_statement_service.py
# ============================================================================
"""Cash Flow Statement generation service - matches reference format exactly."""

import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.config.settings import settings
from backend.services.path_service import PathService


class CashFlowStatementService:
    """Service for generating Cash Flow Statement from markdown files."""

    @staticmethod
    def check_cashflow_readiness(company_name: str) -> Dict:
        """Check if Cash Flow Statement can be generated (always ready)."""
        try:
            # Cash Flow can always be generated
            return {
                "company_name": company_name,
                "is_ready": True,
                "message": "Ready to Generate"
            }

        except Exception as e:
            return {
                "company_name": company_name,
                "is_ready": False,
                "error": str(e),
                "message": f"Error checking readiness: {str(e)}"
            }

    @staticmethod
    def _find_cashflow_markdown(company_name: str) -> Optional[Path]:
        """Find the Cash Flow Statement markdown file."""
        entity_notes_dir = settings.get_entity_generated_notes_dir(company_name)

        if not entity_notes_dir.exists():
            return None

        patterns = [
            "Note_CASHFLOW_*.md",
            "noteCASHFLOW_*.md",
            "cashflow_*.md",
            "cash_flow_*.md"
        ]

        for pattern in patterns:
            md_files = list(entity_notes_dir.glob(pattern))
            if md_files:
                return max(md_files, key=lambda p: p.stat().st_mtime)

        return None

    @staticmethod
    def generate_cashflow_excel(
        company_name: str,
        period_ended: str = None
    ) -> Dict:
        """Generate Cash Flow Statement Excel template from markdown file."""
        try:
            # Try to find the markdown file directly
            md_file = CashFlowStatementService._find_cashflow_markdown(company_name)
            
            if not md_file:
                return {
                    "success": False,
                    "message": "Cash Flow markdown file not found. Please generate the Cash Flow note first."
                }

            with open(md_file, 'r', encoding='utf-8') as f:
                md_content = f.read()

            if not period_ended:
                period_match = re.search(
                    r'For the (?:period ending|year ended):\s*([^\n]+)', md_content
                )
                period_ended = period_match.group(1).strip() if period_match else "Unknown Period"

            # Parse markdown content
            cashflow_data = CashFlowStatementService._parse_markdown_content(md_content)

            # Create Excel file
            output_file = CashFlowStatementService._create_excel_file(
                company_name,
                period_ended,
                cashflow_data
            )

            return {
                "success": True,
                "message": "Cash Flow Statement Excel template generated successfully",
                "output_file": str(output_file),
                "company_name": company_name,
                "period_ended": period_ended,
                "source_markdown": str(md_file)
            }

        except Exception as e:
            return {
                "success": False,
                "message": f"Error generating Cash Flow Excel: {str(e)}",
                "error": str(e)
            }

    @staticmethod
    def _parse_markdown_content(content: str) -> Dict:
        """Parse markdown to extract structured cash flow data."""
        data = {
            "operating": [],
            "investing": [],
            "financing": [],
            "reconciliation": [],
            "cash_breakdown": [],
            "non_cash": []
        }

        # Extract Operating Activities Section
        operating_match = re.search(
            r'\*\*(?:A\.|A\s)?Cash flows? from operating activities\*\*.*?\n(.*?)(?=\*\*(?:B\.|B\s)?Cash flows? from investing|\Z)',
            content,
            re.DOTALL | re.IGNORECASE
        )
        if operating_match:
            data["operating"] = CashFlowStatementService._extract_line_items(
                operating_match.group(1), "operating"
            )

        # Extract Investing Activities Section
        investing_match = re.search(
            r'\*\*(?:B\.|B\s)?Cash flows? from investing activities\*\*.*?\n(.*?)(?=\*\*(?:C\.|C\s)?Cash flows? from financing|\Z)',
            content,
            re.DOTALL | re.IGNORECASE
        )
        if investing_match:
            data["investing"] = CashFlowStatementService._extract_line_items(
                investing_match.group(1), "investing"
            )

        # Extract Financing Activities Section
        financing_match = re.search(
            r'\*\*(?:C\.|C\s)?Cash flows? from financing activities\*\*.*?\n(.*?)(?=\*\*Net \(decrease\)|\*\*Net increase|Cash and cash equivalents|\Z)',
            content,
            re.DOTALL | re.IGNORECASE
        )
        if financing_match:
            data["financing"] = CashFlowStatementService._extract_line_items(
                financing_match.group(1), "financing"
            )

        # Extract Reconciliation Section
        reconciliation_match = re.search(
            r'\*\*Net \(decrease\)/\s*increase in cash.*?\n(.*?)(?=\*\*Cash and cash equivalents include|\Z)',
            content,
            re.DOTALL | re.IGNORECASE
        )
        if reconciliation_match:
            data["reconciliation"] = CashFlowStatementService._extract_line_items(
                reconciliation_match.group(0), "reconciliation"
            )

        # Extract Cash Breakdown
        cash_breakdown_match = re.search(
            r'\*\*Cash and cash equivalents include.*?\n(.*?)(?=\*\*Non-\s*cash|\Z)',
            content,
            re.DOTALL | re.IGNORECASE
        )
        if cash_breakdown_match:
            data["cash_breakdown"] = CashFlowStatementService._extract_line_items(
                cash_breakdown_match.group(1), "cash_breakdown"
            )

        # Extract Non-cash Activities
        non_cash_match = re.search(
            r'\*\*Non-\s*cash financing and investing activities.*?\n(.*?)(?=\Z)',
            content,
            re.DOTALL | re.IGNORECASE
        )
        if non_cash_match:
            data["non_cash"] = CashFlowStatementService._extract_line_items(
                non_cash_match.group(1), "non_cash"
            )

        return data

    @staticmethod
    def _extract_line_items(section_content: str, section_type: str) -> List[Dict]:
        """Extract line items from markdown section."""
        line_items = []
        
        # Match table rows: | Description | Amount |
        table_pattern = r'\|\s*([^|]+?)\s*\|\s*([^|]*)\s*\|'
        
        lines = section_content.split('\n')
        
        for line in lines:
            match = re.match(table_pattern, line)
            if not match:
                continue
                
            description = match.group(1).strip()
            amount_str = match.group(2).strip()
            
            # Skip headers and separators
            if ('Particulars' in description or 'Amount' in description or 
                '---' in description or description == ''):
                continue
            
            # Parse amount
            amount = CashFlowStatementService._parse_amount(amount_str)
            
            # Determine formatting
            is_bold = '**' in description
            is_section_header = section_type in description.lower() or 'activities' in description.lower()
            is_subtotal = any(keyword in description for keyword in [
                'Operating profit before working capital',
                'Cash generated from',
                'Net cash flow',
                'Net cash flows',
                'Net cash used',
                'Net cash generated',
                'Net (decrease)',
                'Net increase',
                'Cash and cash equivalents at the end'
            ])
            
            # Clean description
            clean_desc = description.replace('**', '').strip()
            
            # Determine indent level
            indent_level = 0
            if clean_desc.startswith('- '):
                indent_level = 1
                clean_desc = clean_desc[2:].strip()
            elif section_type == "cash_breakdown" and "Balance with banks" not in clean_desc:
                if "On current accounts" in clean_desc or "Deposits with" in clean_desc:
                    indent_level = 1
            
            line_items.append({
                "description": clean_desc,
                "amount": amount,
                "amount_str": amount_str,
                "is_bold": is_bold or is_subtotal,
                "is_section_header": is_section_header,
                "is_subtotal": is_subtotal,
                "indent_level": indent_level
            })
        
        return line_items

    @staticmethod
    def _parse_amount(amount_str: str) -> Optional[float]:
        """Parse amount string to float."""
        if not amount_str or amount_str.strip() == '' or amount_str == '-':
            return None
        
        # Remove multiple currency symbols and spaces
        # Support: ₹, ₱, $, €, RM, S$, USD, etc.
        import re
        amount_str = re.sub(r'[₹₱$€RM\s]+', '', amount_str)
        amount_str = amount_str.replace(',', '').strip()
        
        # Check if bracketed (negative)
        is_negative = amount_str.startswith('(') and amount_str.endswith(')')
        
        if is_negative:
            amount_str = amount_str[1:-1]
        
        try:
            amount = float(amount_str)
            return -amount if is_negative else amount
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _create_excel_file(
        company_name: str,
        period_ended: str,
        cashflow_data: Dict
    ) -> Path:
        """Create Excel file matching the reference format exactly."""
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Cash Flow Statement"

        # Define styles matching reference image
        title_font = Font(name='Arial', size=11, bold=True)
        section_font = Font(name='Arial', size=10, bold=True)
        bold_font = Font(name='Arial', size=10, bold=True)
        normal_font = Font(name='Arial', size=10)
        small_font = Font(name='Arial', size=9)
        
        # Borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        bottom_border = Border(bottom=Side(style='thin'))
        top_bottom_border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Column widths
        ws.column_dimensions['A'].width = 65
        ws.column_dimensions['B'].width = 18

        row = 1

        # Header Section (no content, just spacing)
        row += 1

        # Section A: Cash flows from operating activities
        ws[f'A{row}'] = "A  Cash flows from operating activities"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].alignment = Alignment(horizontal='left')
        ws[f'B{row}'] = period_ended
        ws[f'B{row}'].font = small_font
        ws[f'B{row}'].alignment = Alignment(horizontal='right')
        row += 1

        # Add operating activities items
        for item in cashflow_data["operating"]:
            if item["is_section_header"]:
                continue
                
            # Set description with proper indentation
            indent = "    " * item["indent_level"]
            ws[f'A{row}'] = indent + item["description"]
            
            # Set font
            if item["is_bold"] or item["is_subtotal"]:
                ws[f'A{row}'].font = bold_font
            else:
                ws[f'A{row}'].font = normal_font
            
            ws[f'A{row}'].alignment = Alignment(horizontal='left', wrap_text=False)
            
            # Set amount
            if item["amount"] is not None:
                ws[f'B{row}'] = item["amount"]
                ws[f'B{row}'].number_format = '#,##0'
                
                if item["is_bold"] or item["is_subtotal"]:
                    ws[f'B{row}'].font = bold_font
                else:
                    ws[f'B{row}'].font = normal_font
                    
                ws[f'B{row}'].alignment = Alignment(horizontal='right')
            
            # Add borders for subtotals
            if item["is_subtotal"]:
                ws[f'A{row}'].border = top_bottom_border
                ws[f'B{row}'].border = top_bottom_border
            
            row += 1

        # Blank row
        row += 1

        # Section B: Cash flows from investing activities
        ws[f'A{row}'] = "B  Cash flows from investing activities"
        ws[f'A{row}'].font = section_font
        row += 1

        for item in cashflow_data["investing"]:
            if item["is_section_header"]:
                continue
                
            indent = "    " * item["indent_level"]
            ws[f'A{row}'] = indent + item["description"]
            
            if item["is_bold"] or item["is_subtotal"]:
                ws[f'A{row}'].font = bold_font
            else:
                ws[f'A{row}'].font = normal_font
            
            if item["amount"] is not None:
                ws[f'B{row}'] = item["amount"]
                ws[f'B{row}'].number_format = '#,##0'
                
                if item["is_bold"] or item["is_subtotal"]:
                    ws[f'B{row}'].font = bold_font
                else:
                    ws[f'B{row}'].font = normal_font
                    
                ws[f'B{row}'].alignment = Alignment(horizontal='right')
            
            if item["is_subtotal"]:
                ws[f'A{row}'].border = top_bottom_border
                ws[f'B{row}'].border = top_bottom_border
            
            row += 1

        # Blank row
        row += 1

        # Section C: Cash flows from financing activities
        ws[f'A{row}'] = "C  Cash flows from financing activities"
        ws[f'A{row}'].font = section_font
        row += 1

        for item in cashflow_data["financing"]:
            if item["is_section_header"]:
                continue
                
            indent = "    " * item["indent_level"]
            ws[f'A{row}'] = indent + item["description"]
            
            if item["is_bold"] or item["is_subtotal"]:
                ws[f'A{row}'].font = bold_font
            else:
                ws[f'A{row}'].font = normal_font
            
            if item["amount"] is not None:
                ws[f'B{row}'] = item["amount"]
                ws[f'B{row}'].number_format = '#,##0'
                
                if item["is_bold"] or item["is_subtotal"]:
                    ws[f'B{row}'].font = bold_font
                else:
                    ws[f'B{row}'].font = normal_font
                    
                ws[f'B{row}'].alignment = Alignment(horizontal='right')
            
            if item["is_subtotal"]:
                ws[f'A{row}'].border = top_bottom_border
                ws[f'B{row}'].border = top_bottom_border
            
            row += 1

        # Blank row
        row += 1

        # Reconciliation section
        for item in cashflow_data["reconciliation"]:
            ws[f'A{row}'] = item["description"]
            
            if item["is_bold"] or item["is_subtotal"]:
                ws[f'A{row}'].font = bold_font
            else:
                ws[f'A{row}'].font = normal_font
            
            if item["amount"] is not None:
                ws[f'B{row}'] = item["amount"]
                ws[f'B{row}'].number_format = '#,##0'
                
                if item["is_bold"] or item["is_subtotal"]:
                    ws[f'B{row}'].font = bold_font
                else:
                    ws[f'B{row}'].font = normal_font
                    
                ws[f'B{row}'].alignment = Alignment(horizontal='right')
            
            # Double underline for final total
            if "at the end" in item["description"].lower():
                ws[f'B{row}'].border = Border(
                    bottom=Side(style='double')
                )
            
            row += 1

        # Blank row
        row += 1

        # Cash and cash equivalents breakdown
        if cashflow_data["cash_breakdown"]:
            ws[f'A{row}'] = "Cash and cash equivalents include (refer note 11)"
            ws[f'A{row}'].font = bold_font
            row += 1
            
            for item in cashflow_data["cash_breakdown"]:
                indent = "    " * item["indent_level"]
                ws[f'A{row}'] = indent + item["description"]
                ws[f'A{row}'].font = normal_font
                
                if item["amount"] is not None:
                    ws[f'B{row}'] = item["amount"]
                    ws[f'B{row}'].number_format = '#,##0'
                    ws[f'B{row}'].font = normal_font
                    ws[f'B{row}'].alignment = Alignment(horizontal='right')
                
                row += 1

        # Blank row
        row += 1

        # Non-cash activities
        if cashflow_data["non_cash"]:
            ws[f'A{row}'] = "Non- cash financing and investing activities:"
            ws[f'A{row}'].font = bold_font
            row += 1
            
            for item in cashflow_data["non_cash"]:
                ws[f'A{row}'] = item["description"]
                ws[f'A{row}'].font = normal_font
                
                if item["amount"] is not None:
                    ws[f'B{row}'] = item["amount"]
                    ws[f'B{row}'].number_format = '#,##0'
                    ws[f'B{row}'].font = normal_font
                    ws[f'B{row}'].alignment = Alignment(horizontal='right')
                
                row += 1

        # Save file
        path_service = PathService(company_name)
        output_dir = path_service.get_financial_statements_dir(company_name) / "CashFlow"
        output_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = output_dir / f"CashFlow_Statement_{timestamp}.xlsx"
        
        wb.save(output_file)
        
        return output_file

    @staticmethod
    def list_cashflow_statements(company_name: str) -> Dict:
        """List all generated Cash Flow statements."""
        try:
            path_service = PathService(company_name)
            cashflow_dir = path_service.get_financial_statements_dir(company_name) / "CashFlow"

            if not cashflow_dir.exists():
                return {
                    "company_name": company_name,
                    "files": [],
                    "statements": [],
                    "count": 0
                }

            statements = []
            for file_path in cashflow_dir.glob("CashFlow_Statement_*.xlsx"):
                statements.append({
                    "filename": file_path.name,
                    "file_path": str(file_path),
                    "size_bytes": file_path.stat().st_size,
                    "generated_at": datetime.fromtimestamp(
                        file_path.stat().st_mtime
                    ).isoformat(),
                    "download_url": f"/api/cashflow-statement/{company_name}/download/{file_path.name}"
                })

            statements.sort(key=lambda x: x["generated_at"], reverse=True)

            return {
                "company_name": company_name,
                "files": statements,  # For consistency with other endpoints
                "statements": statements,  # For backwards compatibility
                "count": len(statements),
                "latest": statements[0] if statements else None
            }

        except Exception as e:
            return {
                "company_name": company_name,
                "files": [],
                "statements": [],
                "count": 0,
                "error": str(e)
            }