# ============================================================================
# FILE: backend/services/bs_statement_service.py
# ============================================================================
"""Service for generating Balance Sheet statements from notes."""

import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from backend.config.settings import settings
from backend.models.balance_sheet_models import (
    BSGenerationResponse,
    BSLineItem,
    BSSection,
    BalanceSheetStatement,
)


from backend.services.path_service import PathService
from backend.services.period_discovery_service import PeriodDiscoveryService
from backend.config.period_config import PeriodConfig

class BSStatementService:
    """Service for generating Balance Sheet statements from note markdown files."""

    # Entity-specific Balance Sheet structure configuration
    ENTITY_BS_CONFIGS = {
        # Default configuration for all entities
        "default": {
            "non_current_assets": ["3", "4", "5", "7"],  # PPE, ROU, Investments, Other non-current
            "current_assets": ["8", "9", "10", "11", "12", "13", "7"],  # Inventories, Trade receivables, Cash, Bank, Loans, Other FA, Other current
            "equity": ["14", "15"],  # Equity share capital, Other equity
            "non_current_liabilities": ["17", "18", "6"],  # Borrowings, Lease liabilities, Deferred tax
            "current_liabilities": ["16", "17", "18", "19", "20", "21", "22", "23"],  # Contract, Borrowings, Lease, Trade payables, Other FA, Other current, Provisions, Current tax
        },
    }

    # Note number to description mapping
    NOTE_DESCRIPTIONS = {
        # Assets
        "3": "Property, plant and equipment",
        "4": "Right of use assets",
        "5": "Investments",
        "7": "Other non-current assets",
        "8": "Inventories",
        "9": "Trade receivables",
        "10": "Cash and cash equivalents",
        "11": "Bank balances other than (ii) above",
        "12": "Loans",
        "13": "Other financial assets",
        # Equity
        "14": "Equity share capital",
        "15": "Other equity",
        # Liabilities
        "6": "Deferred tax liabilities (net)",
        "16": "Contract liabilities",
        "17": "Borrowings",
        "18": "Lease liabilities",
        "19": "Trade payables",
        "20": "Other financial liabilities",
        "21": "Other current liabilities",
        "22": "Provisions",
        "23": "Current tax liabilities (net)",
    }

    @staticmethod
    def _get_entity_config(company_name: str) -> Dict[str, List[str]]:
        """Get Balance Sheet configuration for a specific entity."""
        return BSStatementService.ENTITY_BS_CONFIGS.get(
            company_name, BSStatementService.ENTITY_BS_CONFIGS["default"]
        )

    @staticmethod
    def _find_latest_note_file(company_name: str, note_number: str) -> Optional[Path]:
        """
        Find the most recent generated note file.

        Args:
            company_name: Name of the company
            note_number: Note number to find

        Returns:
            Path to latest note file or None
        """
        entity_notes_dir = settings.get_entity_generated_notes_dir(company_name)

        if not entity_notes_dir.exists():
            return None

        # Try both patterns
        note_files = list(entity_notes_dir.glob(f"note{note_number}_*.md"))
        if not note_files:
            note_files = list(entity_notes_dir.glob(f"Note_{note_number}_*.md"))

        if not note_files:
            return None

        return max(note_files, key=lambda p: p.stat().st_mtime)

    @staticmethod
    def _extract_total_from_markdown(md_content: str) -> Optional[float]:
        """
        Extract total amount from markdown note content.

        Args:
            md_content: Markdown content of the note

        Returns:
            Total amount as float (always positive)
        """
        # Updated patterns to support multiple currency symbols: ₹, ₱, $, RM, €, S$, etc.
        patterns = [
            # Bold total with bold amount - with currency
            r"\|\s*\*\*TOTAL[^|]*\*\*\s*\|\s*\*\*[₹₱$€RM]*\s*([\d,]+(?:\.\d{2})?)\*\*",
            r"\|\s*\*\*Total[^|]*\*\*\s*\|\s*\*\*[₹₱$€RM]*\s*([\d,]+(?:\.\d{2})?)\*\*",
            # Regular total - with currency
            r"\|\s*Total[^|]*\|\s*[₹₱$€RM]*\s*([\d,]+(?:\.\d{2})?)",
            # Standalone bold total - with currency
            r"\*\*Total[^:]*:\*\*\s*[₹₱$€RM]*\s*([\d,]+(?:\.\d{2})?)",
            # Without currency symbols
            r"\|\s*\*\*TOTAL[^|]*\*\*\s*\|\s*\*\*([\d,]+(?:\.\d{2})?)\*\*",
            r"\|\s*\*\*Total[^|]*\*\*\s*\|\s*\*\*([\d,]+(?:\.\d{2})?)\*\*",
            r"\|\s*Total[^|]*\|\s*([\d,]+(?:\.\d{2})?)",
            r"\*\*Total[^:]*:\*\*\s*([\d,]+(?:\.\d{2})?)",
        ]

        for pattern in patterns:
            matches = re.findall(pattern, md_content, re.IGNORECASE)
            if matches:
                amount_str = matches[-1].replace(",", "")
                try:
                    amount = abs(float(amount_str))
                    return amount
                except ValueError:
                    continue

        return None

    @staticmethod
    def _extract_note_details(
        company_name: str, note_number: str
    ) -> Tuple[Optional[float], Optional[str]]:
        """
        Extract amount and description from a note file.

        Args:
            company_name: Name of the company
            note_number: Note number

        Returns:
            Tuple of (amount, description)
        """
        note_file = BSStatementService._find_latest_note_file(company_name, note_number)

        if not note_file:
            return None, BSStatementService.NOTE_DESCRIPTIONS.get(note_number)

        try:
            with open(note_file, "r", encoding="utf-8") as f:
                content = f.read()

            amount = BSStatementService._extract_total_from_markdown(content)
            description = BSStatementService.NOTE_DESCRIPTIONS.get(note_number)

            return amount, description

        except Exception as e:
            print(f"Error reading note {note_number}: {e}")
            return None, BSStatementService.NOTE_DESCRIPTIONS.get(note_number)

    @staticmethod
    def check_bs_readiness(company_name: str) -> Dict:
        """
        Check if all required notes are available for Balance Sheet generation.

        Args:
            company_name: Name of the company

        Returns:
            Dictionary with readiness status and details
        """
        config = BSStatementService._get_entity_config(company_name)

        # Get all unique note numbers
        all_notes = set()
        for category_notes in config.values():
            all_notes.update(category_notes)

        required_notes = sorted(list(all_notes))

        found_notes = []
        missing_notes = []
        note_details = {}

        for note_num in required_notes:
            note_file = BSStatementService._find_latest_note_file(company_name, note_num)
            if note_file:
                found_notes.append(note_num)
                amount, description = BSStatementService._extract_note_details(
                    company_name, note_num
                )
                note_details[note_num] = {
                    "description": description,
                    "amount": amount,
                    "file_path": str(note_file),
                    "generated_at": datetime.fromtimestamp(
                        note_file.stat().st_mtime
                    ).isoformat(),
                }
            else:
                missing_notes.append(note_num)

        is_ready = len(missing_notes) == 0

        return {
            "company_name": company_name,
            "is_ready": is_ready,
            "found_notes": found_notes,
            "missing_notes": missing_notes,
            "note_details": note_details,
            "total_required": len(required_notes),
            "total_found": len(found_notes),
            "completeness_percentage": round(
                (len(found_notes) / len(required_notes)) * 100, 2
            ) if len(required_notes) > 0 else 0,
            "config": config,
        }

    @staticmethod
    def _build_asset_section(
        section_name: str,
        section_label: str,
        note_list: List[str],
        note_amounts: Dict[str, float],
        company_name: str,
        total_label: Optional[str] = None,
    ) -> Tuple[BSSection, float]:
        """Build an asset section with sub-items."""
        line_items = []
        section_total = 0.0

        # Add section header
        line_items.append(
            BSLineItem(
                particulars=section_label,
                is_subtotal=False,
                indent_level=0,
            )
        )

        # Add each note as a line item
        for note_num in note_list:
            amount, description = BSStatementService._extract_note_details(
                company_name, note_num
            )

            if amount is not None:
                note_amounts[note_num] = amount
                section_total += amount

            # Determine indent level based on line type
            indent = 0
            if description and ("(i)" in description or "(ii)" in description or "(iii)" in description or "(iv)" in description):
                indent = 1
            elif description and description.startswith("    "):
                indent = 2

            line_items.append(
                BSLineItem(
                    particulars=description or f"Note {note_num}",
                    note=note_num,
                    amount=amount,
                    indent_level=indent,
                )
            )

        # Add section total
        if total_label:
            line_items.append(
                BSLineItem(
                    particulars=total_label,
                    amount=section_total,
                    is_subtotal=True,
                    indent_level=0,
                )
            )

        return BSSection(section_name=section_name, line_items=line_items), section_total

    @staticmethod
    def generate_bs_statement(
        company_name: str, as_at_date: str, note_numbers: Optional[List[str]] = None
    ) -> BSGenerationResponse:
        """
        Generate complete Balance Sheet statement from note files.

        Args:
            company_name: Name of the company
            as_at_date: Balance sheet date
            note_numbers: Optional (ignored, uses entity config)

        Returns:
            BSGenerationResponse with generated statement
        """
        try:
            # Check readiness first
            readiness = BSStatementService.check_bs_readiness(company_name)

            if not readiness["is_ready"]:
                return BSGenerationResponse(
                    success=False,
                    message=f"Cannot generate Balance Sheet. Missing notes: {', '.join(readiness['missing_notes'])}",
                )

            # Get entity configuration
            config = BSStatementService._get_entity_config(company_name)

            # Store extracted amounts
            note_amounts = {}
            sections = []

            # === ASSETS ===
            assets_sections = []

            # A. Non-current assets
            nca_section, nca_total = BSStatementService._build_asset_section(
                "A.Non-current assets",
                "A.Non-current assets",
                config["non_current_assets"],
                note_amounts,
                company_name,
                "Total non-current assets (A)",
            )
            assets_sections.append(nca_section)

            # B. Current assets
            ca_section, ca_total = BSStatementService._build_asset_section(
                "B.Current assets",
                "B.Current assets",
                config["current_assets"],
                note_amounts,
                company_name,
                "Total current assets (B)",
            )
            assets_sections.append(ca_section)

            # Total assets
            total_assets = nca_total + ca_total
            assets_sections.append(
                BSSection(
                    section_name="Total assets (A+B)",
                    line_items=[
                        BSLineItem(
                            particulars="Total assets (A+B)",
                            amount=total_assets,
                            is_total=True,
                            indent_level=0,
                        )
                    ],
                )
            )

            sections.extend(assets_sections)

            # === EQUITY AND LIABILITIES ===
            el_sections = []

            # C. Equity
            equity_items = []
            equity_total = 0.0

            equity_items.append(
                BSLineItem(
                    particulars="C.Equity",
                    is_subtotal=False,
                    indent_level=0,
                )
            )

            for note_num in config["equity"]:
                amount, description = BSStatementService._extract_note_details(
                    company_name, note_num
                )
                if amount is not None:
                    note_amounts[note_num] = amount
                    equity_total += amount

                equity_items.append(
                    BSLineItem(
                        particulars=description or f"Note {note_num}",
                        note=note_num,
                        amount=amount,
                        indent_level=0,
                    )
                )

            equity_items.append(
                BSLineItem(
                    particulars="Total equity (C)",
                    amount=equity_total,
                    is_subtotal=True,
                    indent_level=0,
                )
            )

            el_sections.append(
                BSSection(section_name="C.Equity", line_items=equity_items)
            )

            # LIABILITIES header
            el_sections.append(
                BSSection(
                    section_name="LIABILITIES",
                    line_items=[
                        BSLineItem(
                            particulars="LIABILITIES",
                            is_subtotal=False,
                            indent_level=0,
                        )
                    ],
                )
            )

            # D. Non-current liabilities
            ncl_items = []
            ncl_total = 0.0

            ncl_items.append(
                BSLineItem(
                    particulars="D.Non-current liabilities",
                    is_subtotal=False,
                    indent_level=0,
                )
            )

            # Financial liabilities sub-header
            ncl_items.append(
                BSLineItem(
                    particulars="Financial liabilities",
                    indent_level=0,
                )
            )

            for note_num in config["non_current_liabilities"]:
                amount, description = BSStatementService._extract_note_details(
                    company_name, note_num
                )
                if amount is not None:
                    note_amounts[note_num] = amount
                    ncl_total += amount

                indent = 1 if "(i)" in (description or "") or "(ii)" in (description or "") else 0

                ncl_items.append(
                    BSLineItem(
                        particulars=description or f"Note {note_num}",
                        note=note_num,
                        amount=amount,
                        indent_level=indent,
                    )
                )

            ncl_items.append(
                BSLineItem(
                    particulars="Total non-current liabilities (D)",
                    amount=ncl_total,
                    is_subtotal=True,
                    indent_level=0,
                )
            )

            el_sections.append(
                BSSection(section_name="D.Non-current liabilities", line_items=ncl_items)
            )

            # E. Current liabilities
            cl_items = []
            cl_total = 0.0

            cl_items.append(
                BSLineItem(
                    particulars="E.Current liabilities",
                    is_subtotal=False,
                    indent_level=0,
                )
            )

            # Add contract liabilities first
            cl_items.append(
                BSLineItem(
                    particulars="Contract liabilities",
                    indent_level=0,
                )
            )

            # Financial liabilities sub-header
            cl_items.append(
                BSLineItem(
                    particulars="Financial liabilities",
                    indent_level=0,
                )
            )

            for note_num in config["current_liabilities"]:
                amount, description = BSStatementService._extract_note_details(
                    company_name, note_num
                )
                if amount is not None:
                    note_amounts[note_num] = amount
                    cl_total += amount

                indent = 1 if "(i)" in (description or "") or "(ii)" in (description or "") or "(iii)" in (description or "") or "(iv)" in (description or "") else 0

                cl_items.append(
                    BSLineItem(
                        particulars=description or f"Note {note_num}",
                        note=note_num,
                        amount=amount,
                        indent_level=indent,
                    )
                )

            cl_items.append(
                BSLineItem(
                    particulars="Total current liabilities (E)",
                    amount=cl_total,
                    is_subtotal=True,
                    indent_level=0,
                )
            )

            el_sections.append(
                BSSection(section_name="E.Current liabilities", line_items=cl_items)
            )

            # Total liabilities
            total_liabilities = ncl_total + cl_total
            el_sections.append(
                BSSection(
                    section_name="Total liabilities (F= D+E)",
                    line_items=[
                        BSLineItem(
                            particulars="Total liabilities (F= D+E)",
                            amount=total_liabilities,
                            is_total=True,
                            indent_level=0,
                        )
                    ],
                )
            )

            # Total equity and liabilities
            total_el = equity_total + total_liabilities
            el_sections.append(
                BSSection(
                    section_name="Total equity and liabilities (C+F)",
                    line_items=[
                        BSLineItem(
                            particulars="Total equity and liabilities (C+F)",
                            amount=total_el,
                            is_total=True,
                            indent_level=0,
                        )
                    ],
                )
            )

            sections.extend(el_sections)

            # Create statement
            statement = BalanceSheetStatement(
                company_name=company_name,
                as_at_date=as_at_date,
                sections=sections,
                metadata={
                    "generated_at": datetime.now().isoformat(),
                    "notes_used": list(note_amounts.keys()),
                    "config_used": config,
                    "total_assets": total_assets,
                    "total_equity_liabilities": total_el,
                },
            )

            # Generate Excel file
            output_file = BSStatementService._export_to_excel(statement)

            return BSGenerationResponse(
                success=True,
                message=f"Balance Sheet generated successfully for {company_name}",
                statement=statement,
                output_file=str(output_file),
            )

        except Exception as e:
            import traceback
            traceback.print_exc()
            return BSGenerationResponse(
                success=False, message=f"Error generating Balance Sheet: {str(e)}"
            )

    @staticmethod
    def _format_indian_number(value: float) -> str:
        """Format number in Indian numbering system."""
        if value is None:
            return ""
        
        value = abs(value)

        if value == 0:
            return "0"

        s = f"{value:,.2f}"
        parts = s.split(".")

        whole = parts[0].replace(",", "")
        if len(whole) > 3:
            last_three = whole[-3:]
            remaining = whole[:-3]

            result = ""
            while len(remaining) > 2:
                result = "," + remaining[-2:] + result
                remaining = remaining[:-2]

            if remaining:
                result = remaining + result

            formatted = result + "," + last_three
        else:
            formatted = whole

        if parts[1] == "00":
            return formatted

        return formatted + "." + parts[1]

    @staticmethod
    def _export_to_excel(statement: BalanceSheetStatement) -> Path:
        """Export Balance Sheet to Excel with formatting."""
        path_service = PathService(statement.company_name)
        bs_dir = path_service.get_financial_statements_dir(statement.company_name) / "BS"
        bs_dir.mkdir(parents=True, exist_ok=True)
        reports_dir = bs_dir

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = reports_dir / f"BS_Statement_{timestamp}.xlsx"

        data = []

        # Title rows
        data.append([statement.company_name.upper()])
        data.append(["BALANCE SHEET"])
        
        # Format period name: convert period key to MMM-YY format (e.g., "Jun-25")
        period_key = PeriodConfig.get_current_period()
        if period_key:
            period_display_full = PeriodDiscoveryService.get_period_display_name(period_key)
            # Convert "Jun-2025" to "Jun-25"
            parts = period_display_full.split('-')
            if len(parts) == 2 and len(parts[1]) == 4:
                period_display = f"{parts[0]}-{parts[1][-2:]}"  # Take last 2 digits of year
            else:
                period_display = period_display_full
        else:
            period_display = statement.as_at_date
        
        data.append([f"AS AT {period_display.upper()}"])
        data.append([])

        # Header row
        data.append(["Particulars", "Notes", f"As at\n{period_display}"])

        # Add sections
        for section in statement.sections:
            for item in section.line_items:
                particulars = "  " * item.indent_level + item.particulars
                note = item.note if item.note else ""
                amount_str = ""
                if item.amount is not None:
                    amount_str = BSStatementService._format_indian_number(item.amount)

                data.append([particulars, note, amount_str])

        # Create DataFrame
        df = pd.DataFrame(data)

        # Export with styling
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Balance Sheet", index=False, header=False)

            worksheet = writer.sheets["Balance Sheet"]

            # Formatting
            title_font = Font(name="Arial", size=14, bold=True)
            header_font = Font(name="Arial", size=11, bold=True)
            normal_font = Font(name="Arial", size=10)
            bold_font = Font(name="Arial", size=10, bold=True)

            header_fill = PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
            )

            # Title rows
            for row in range(1, 4):
                cell = worksheet.cell(row, 1)
                cell.font = title_font
                cell.alignment = Alignment(horizontal="center")
                worksheet.merge_cells(f"A{row}:C{row}")

            # Header row
            for col in range(1, 4):
                cell = worksheet.cell(5, col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center" if col > 1 else "left", wrap_text=True)

            # Data rows
            for row in range(6, len(data) + 1):
                for col in range(1, 4):
                    cell = worksheet.cell(row, col)
                    cell.font = normal_font

                    cell_value = str(cell.value) if cell.value else ""
                    if (
                        "Total" in cell_value
                        or "ASSETS" in cell_value
                        or "EQUITY" in cell_value
                        or "LIABILITIES" in cell_value
                    ):
                        cell.font = bold_font

                    if col == 1:
                        cell.alignment = Alignment(horizontal="left")
                    elif col == 2:
                        cell.alignment = Alignment(horizontal="center")
                    else:
                        cell.alignment = Alignment(horizontal="right")

            # Column widths
            worksheet.column_dimensions["A"].width = 70
            worksheet.column_dimensions["B"].width = 10
            worksheet.column_dimensions["C"].width = 20

        return output_file