"""
Data Extractor Module
Extracts financial data from both SQL Server and API sources
Supports Trial Balance, P&L, Balance Sheet, Cash Flow, and Sales Reports
"""

import pandas as pd
from datetime import datetime
from typing import Dict, List, Optional
from backend.sap_connect.connectivity_manager import ConnectivityManager


class DataExtractor:
    """
    Extract financial data from SAP B1 via SQL or API
    """
    
    def __init__(self, connectivity_manager: ConnectivityManager):
        """
        Initialize data extractor
        
        Args:
            connectivity_manager: ConnectivityManager instance
        """
        self.conn_mgr = connectivity_manager
        self.config = connectivity_manager.config
    
    # ========== SQL Server Data Extraction ==========
    
    def _get_trial_balance_sql(self, entity: Dict, start_date: str, end_date: str) -> pd.DataFrame:
        """
        Extract Trial Balance from SQL Server
        
        Args:
            entity: Entity configuration
            start_date: Start date (YYYY-MM-DD)
            end_date: End date (YYYY-MM-DD)
        """
        query = f"""
        SELECT 
            t0.AcctCode AS AccountCode,
            t0.AcctName AS AccountName,
            t0.GroupMask AS AccountType,
            t0.Levels AS Level,
            t0.FatherNum AS ParentAccount,
            COALESCE(SUM(CASE WHEN t1.TransType IN ('13','15','18','30','162','204') 
                              THEN t1.Debit ELSE 0 END), 0) AS Debit,
            COALESCE(SUM(CASE WHEN t1.TransType IN ('13','15','18','30','162','204') 
                              THEN t1.Credit ELSE 0 END), 0) AS Credit
        FROM OACT t0
        LEFT JOIN JDT1 t1 ON t0.AcctCode = t1.Account 
            AND t1.RefDate BETWEEN '{start_date}' AND '{end_date}'
        WHERE t0.Postable = 'Y'
        GROUP BY t0.AcctCode, t0.AcctName, t0.GroupMask, t0.Levels, t0.FatherNum
        ORDER BY t0.AcctCode
        """
        
        results = self.conn_mgr.execute_sql_query(entity, query)
        df = pd.DataFrame(results)
        
        if not df.empty:
            # Convert numeric columns to proper types
            df['Debit'] = pd.to_numeric(df['Debit'], errors='coerce').fillna(0)
            df['Credit'] = pd.to_numeric(df['Credit'], errors='coerce').fillna(0)
            df['Balance'] = df['Debit'] - df['Credit']
        
        return df
    
    def _get_profit_loss_sql(self, entity: Dict, start_date: str, end_date: str) -> pd.DataFrame:
        """
        Extract Profit & Loss from SQL Server
        
        Args:
            entity: Entity configuration
            start_date: Start date (YYYY-MM-DD)
            end_date: End date (YYYY-MM-DD)
        """
        query = f"""
        SELECT 
            t0.AcctCode AS AccountCode,
            t0.AcctName AS AccountName,
            t0.GroupMask AS AccountType,
            COALESCE(SUM(t1.Debit), 0) AS Debit,
            COALESCE(SUM(t1.Credit), 0) AS Credit,
            COALESCE(SUM(t1.Credit - t1.Debit), 0) AS Amount
        FROM OACT t0
        LEFT JOIN JDT1 t1 ON t0.AcctCode = t1.Account 
            AND t1.RefDate BETWEEN '{start_date}' AND '{end_date}'
        WHERE (t0.AcctCode LIKE '4%' OR t0.AcctCode LIKE '5%' OR t0.AcctCode LIKE '6%')
        GROUP BY t0.AcctCode, t0.AcctName, t0.GroupMask
        HAVING SUM(ISNULL(t1.Debit, 0)) != 0 OR SUM(ISNULL(t1.Credit, 0)) != 0
        ORDER BY t0.AcctCode
        """
        
        results = self.conn_mgr.execute_sql_query(entity, query)
        df = pd.DataFrame(results)
        
        if not df.empty:
            # Convert numeric columns to proper types
            df['Debit'] = pd.to_numeric(df['Debit'], errors='coerce').fillna(0)
            df['Credit'] = pd.to_numeric(df['Credit'], errors='coerce').fillna(0)
            df['Amount'] = pd.to_numeric(df['Amount'], errors='coerce').fillna(0)
        
        return df
    
    def _get_balance_sheet_sql(self, entity: Dict, as_of_date: str) -> pd.DataFrame:
        """
        Extract Balance Sheet from SQL Server
        
        Args:
            entity: Entity configuration
            as_of_date: As of date (YYYY-MM-DD)
        """
        query = f"""
        SELECT 
            t0.AcctCode AS AccountCode,
            t0.AcctName AS AccountName,
            t0.GroupMask AS AccountType,
            COALESCE(SUM(t1.Debit), 0) AS Debit,
            COALESCE(SUM(t1.Credit), 0) AS Credit,
            COALESCE(SUM(t1.Debit - t1.Credit), 0) AS Balance
        FROM OACT t0
        LEFT JOIN JDT1 t1 ON t0.AcctCode = t1.Account 
            AND t1.RefDate <= '{as_of_date}'
        WHERE (t0.AcctCode LIKE '1%' OR t0.AcctCode LIKE '2%' OR t0.AcctCode LIKE '3%')
        GROUP BY t0.AcctCode, t0.AcctName, t0.GroupMask
        HAVING SUM(ISNULL(t1.Debit, 0)) != 0 OR SUM(ISNULL(t1.Credit, 0)) != 0
        ORDER BY t0.AcctCode
        """
        
        results = self.conn_mgr.execute_sql_query(entity, query)
        df = pd.DataFrame(results)
        
        if not df.empty:
            # Convert numeric columns to proper types
            df['Debit'] = pd.to_numeric(df['Debit'], errors='coerce').fillna(0)
            df['Credit'] = pd.to_numeric(df['Credit'], errors='coerce').fillna(0)
            df['Balance'] = pd.to_numeric(df['Balance'], errors='coerce').fillna(0)
        
        return df
    
    def _get_sales_report_sql(self, entity: Dict, start_date: str, end_date: str) -> pd.DataFrame:
        """
        Extract Sales Report from SQL Server
        
        Args:
            entity: Entity configuration
            start_date: Start date (YYYY-MM-DD)
            end_date: End date (YYYY-MM-DD)
        """
        query = f"""
        SELECT 
            t0.DocNum AS InvoiceNumber,
            t0.DocDate AS InvoiceDate,
            t0.CardCode AS CustomerCode,
            t0.CardName AS CustomerName,
            t1.ItemCode,
            t1.Dscription AS ItemDescription,
            t1.Quantity,
            t1.Price AS UnitPrice,
            t1.LineTotal AS LineTotal,
            t0.DocTotal AS InvoiceTotal,
            t0.DocCur AS Currency
        FROM OINV t0
        INNER JOIN INV1 t1 ON t0.DocEntry = t1.DocEntry
        WHERE t0.DocDate BETWEEN '{start_date}' AND '{end_date}'
        ORDER BY t0.DocDate DESC, t0.DocNum
        """
        
        results = self.conn_mgr.execute_sql_query(entity, query)
        df = pd.DataFrame(results)
        
        if not df.empty:
            # Convert numeric columns to proper types
            df['Quantity'] = pd.to_numeric(df['Quantity'], errors='coerce').fillna(0)
            df['UnitPrice'] = pd.to_numeric(df['UnitPrice'], errors='coerce').fillna(0)
            df['LineTotal'] = pd.to_numeric(df['LineTotal'], errors='coerce').fillna(0)
            df['InvoiceTotal'] = pd.to_numeric(df['InvoiceTotal'], errors='coerce').fillna(0)
        
        return df
    
    # ========== SQL Master Data Extraction ==========
    
    def _get_chart_of_accounts_sql(self, entity: Dict) -> pd.DataFrame:
        """Get Chart of Accounts from SQL"""
        query = """
        SELECT 
            AcctCode AS Code,
            AcctName AS Name,
            CurrTotal AS Balance,
            ActCurr AS Currency,
            Finanse AS FinanceType,
            GroupMask,
            FrozenFor AS Frozen,
            ActType AS AccountType,
            Postable,
            FormatCode
        FROM OACT
        ORDER BY AcctCode
        """
        results = self.conn_mgr.execute_sql_query(entity, query)
        return pd.DataFrame(results)
    
    def _get_business_partners_sql(self, entity: Dict) -> pd.DataFrame:
        """Get Business Partners from SQL"""
        query = """
        SELECT 
            CardCode AS Code,
            CardName AS Name,
            CardType AS Type,
            GroupCode,
            Currency,
            Phone1 AS Phone,
            E_Mail AS Email,
            Address AS BillingAddress,
            MailAddres AS MailingAddress,
            Notes,
            Balance,
            DebtLine AS CreditLimit,
            CreditLine
        FROM OCRD
        ORDER BY CardCode
        """
        results = self.conn_mgr.execute_sql_query(entity, query)
        return pd.DataFrame(results)
    
    def _get_business_partner_groups_sql(self, entity: Dict) -> pd.DataFrame:
        """Get Business Partner Groups from SQL"""
        query = """
        SELECT 
            GroupCode AS Code,
            GroupName AS Name,
            GroupType AS Type
        FROM OCRG
        ORDER BY GroupCode
        """
        results = self.conn_mgr.execute_sql_query(entity, query)
        return pd.DataFrame(results)
    
    def _get_payment_terms_sql(self, entity: Dict) -> pd.DataFrame:
        """Get Payment Terms from SQL"""
        query = """
        SELECT 
            GroupNum AS Code,
            PymntGroup AS Name,
            ExtraDays AS ExtraDays,
            ExtraMonth AS ExtraMonths
        FROM OCTG
        ORDER BY GroupNum
        """
        results = self.conn_mgr.execute_sql_query(entity, query)
        return pd.DataFrame(results)
    
    def _get_withholding_tax_sql(self, entity: Dict) -> pd.DataFrame:
        """Get Withholding Tax Codes from SQL"""
        query = """
        SELECT 
            WTCode AS Code,
            WTName AS Name,
            Rate,
            Category,
            BaseType,
            Account AS GLAccount
        FROM OWHT
        ORDER BY WTCode
        """
        results = self.conn_mgr.execute_sql_query(entity, query)
        return pd.DataFrame(results)
    
    def _get_banks_sql(self, entity: Dict) -> pd.DataFrame:
        """Get Bank Master from SQL"""
        query = """
        SELECT 
            BankCode AS Code,
            BankName AS Name,
            CountryCod AS CountryCode,
            DfltBranch AS Branch,
            DfltAcct AS DefaultAccount,
            SwiftNum AS SwiftCode,
            IBAN
        FROM ODSC
        ORDER BY BankCode
        """
        results = self.conn_mgr.execute_sql_query(entity, query)
        return pd.DataFrame(results)
    
    def _get_item_groups_sql(self, entity: Dict) -> pd.DataFrame:
        """Get Item Groups from SQL"""
        query = """
        SELECT 
            ItmsGrpCod AS Code,
            ItmsGrpNam AS Name,
            Locked
        FROM OITB
        ORDER BY ItmsGrpCod
        """
        results = self.conn_mgr.execute_sql_query(entity, query)
        return pd.DataFrame(results)
    
    def _get_currencies_sql(self, entity: Dict) -> pd.DataFrame:
        """Get Currencies from SQL"""
        query = """
        SELECT 
            CurrCode AS Code,
            CurrName AS Name,
            DocCurrCod AS DocumentCode,
            Decimals
        FROM OCRN
        ORDER BY CurrCode
        """
        results = self.conn_mgr.execute_sql_query(entity, query)
        return pd.DataFrame(results)
    
    def _get_tax_codes_sql(self, entity: Dict) -> pd.DataFrame:
        """Get Sales Tax Codes from SQL"""
        query = """
        SELECT 
            Code,
            Name,
            Rate,
            Category,
            Locked
        FROM OVTG
        ORDER BY Code
        """
        results = self.conn_mgr.execute_sql_query(entity, query)
        return pd.DataFrame(results)
    
    def _get_cost_centers_sql(self, entity: Dict) -> pd.DataFrame:
        """Get Cost Centers (Profit Centers) from SQL"""
        query = """
        SELECT 
            PrcCode AS Code,
            PrcName AS Name,
            GrpCode AS GroupCode,
            Active,
            Locked
        FROM OPRC
        ORDER BY PrcCode
        """
        results = self.conn_mgr.execute_sql_query(entity, query)
        return pd.DataFrame(results)
    
    def _get_warehouse_locations_sql(self, entity: Dict) -> pd.DataFrame:
        """Get Warehouse Locations from SQL"""
        query = """
        SELECT 
            Code,
            Location,
            PanCirNo,
            PanWardNo,
            TanNo
        FROM OLCT
        ORDER BY Code
        """
        results = self.conn_mgr.execute_sql_query(entity, query)
        return pd.DataFrame(results)
    
    def _get_journal_entries_sql(self, entity: Dict, start_date: str, end_date: str) -> pd.DataFrame:
        """Get Journal Entries from SQL Server (JDT1)"""
        query = f"""
        SELECT 
            t0.TransId AS TransactionID,
            t0.Line_ID AS LineNumber,
            t0.RefDate AS ReferenceDate,
            t0.DueDate AS DueDate,
            t0.Ref1 AS Reference1,
            t0.Ref2 AS Reference2,
            t0.TransType AS TransactionType,
            t0.BaseRef AS BaseReference,
            t0.Account AS AccountCode,
            t1.AcctName AS AccountName,
            t0.Debit,
            t0.Credit,
            t0.FCDebit AS FCDebit,
            t0.FCCredit AS FCCredit,
            t0.FCCurrency AS Currency,
            t0.LineMemo AS Memo,
            t0.Project AS ProjectCode,
            t0.ProfitCode AS ProfitCenter,
            t0.DebCred AS DebitCredit
        FROM JDT1 t0
        LEFT JOIN OACT t1 ON t0.Account = t1.AcctCode
        WHERE t0.RefDate BETWEEN '{start_date}' AND '{end_date}'
        ORDER BY t0.TransId, t0.Line_ID
        """
        results = self.conn_mgr.execute_sql_query(entity, query)
        df = pd.DataFrame(results)
        
        if not df.empty:
            # Convert numeric columns to proper types
            df['Debit'] = pd.to_numeric(df['Debit'], errors='coerce').fillna(0)
            df['Credit'] = pd.to_numeric(df['Credit'], errors='coerce').fillna(0)
            df['FCDebit'] = pd.to_numeric(df['FCDebit'], errors='coerce').fillna(0)
            df['FCCredit'] = pd.to_numeric(df['FCCredit'], errors='coerce').fillna(0)
        
        return df
    
    # ========== API Data Extraction ==========
    
    def _get_chart_of_accounts_api(self, entity: Dict) -> pd.DataFrame:
        """Get Chart of Accounts from API"""
        # Check if entity has custom field configuration for ChartOfAccounts
        default_fields = 'Code,Name,AccountType,Levels,FatherAccountKey'
        
        if 'api_field_config' in entity and 'ChartOfAccounts' in entity['api_field_config']:
            select_fields = entity['api_field_config']['ChartOfAccounts']['select_fields']
            print(f"✓ Using custom ChartOfAccounts fields for {entity.get('id')}: {select_fields}")
        else:
            select_fields = default_fields
        
        try:
            try:
                data = self.conn_mgr.fetch_api_data(
                    entity,
                    'ChartOfAccounts',
                    select_fields=select_fields
                )
            except Exception as e:
                # If default fields fail, try without Levels field
                if 'Levels' in str(e) and select_fields == default_fields:
                    print(f"⚠️  ChartOfAccounts API failed with Levels field for {entity.get('id')}, retrying without it")
                    select_fields = 'Code,Name,AccountType,FatherAccountKey'
                    data = self.conn_mgr.fetch_api_data(
                        entity,
                        'ChartOfAccounts',
                        select_fields=select_fields
                    )
                else:
                    raise
            
            df = pd.DataFrame(data)
            
            # Ensure we have the required columns
            if not df.empty:
                # Add Levels column if it doesn't exist (set to None/0)
                if 'Levels' not in df.columns:
                    df['Levels'] = 0
                    print(f"✓ Added default Levels column for {entity.get('id')}")
                
                # Don't request Balance from API, we'll calculate it
                if 'Code' not in df.columns or 'Name' not in df.columns:
                    raise ValueError(f"Chart of Accounts missing required columns. Got: {df.columns.tolist()}")
            
            return df
        finally:
            # Always disconnect after the operation
            self.conn_mgr.disconnect_api(entity)
    
    def _get_journal_entries_api(self, entity: Dict, start_date: str, end_date: str) -> pd.DataFrame:
        """Get Journal Entries from API"""
        filter_query = f"ReferenceDate ge '{start_date}' and ReferenceDate le '{end_date}'"
        
        try:
            data = self.conn_mgr.fetch_api_data(
                entity,
                'JournalEntries',
                filter_query=filter_query
            )
            return pd.DataFrame(data)
        finally:
            # Always disconnect after the operation
            self.conn_mgr.disconnect_api(entity)
    
    def _get_trial_balance_api(self, entity: Dict, start_date: str, end_date: str) -> pd.DataFrame:
        """
        Generate Trial Balance from API data
        
        Args:
            entity: Entity configuration
            start_date: Start date (YYYY-MM-DD)
            end_date: End date (YYYY-MM-DD)
        """
        # Get chart of accounts
        coa_df = self._get_chart_of_accounts_api(entity)
        
        # Get journal entries
        je_df = self._get_journal_entries_api(entity, start_date, end_date)
        
        # Process journal entries to calculate debits and credits
        if je_df.empty:
            # Return COA with zero balances
            coa_df['Debit'] = 0
            coa_df['Credit'] = 0
            coa_df['Balance'] = 0
            return coa_df[['Code', 'Name', 'AccountType', 'Debit', 'Credit', 'Balance']]
        
        # Flatten journal entry lines
        all_lines = []
        for _, je in je_df.iterrows():
            for line in je.get('JournalEntryLines', []):
                all_lines.append({
                    'AccountCode': line.get('AccountCode'),
                    'Debit': line.get('Debit', 0),
                    'Credit': line.get('Credit', 0)
                })
        
        lines_df = pd.DataFrame(all_lines)
        
        # Aggregate by account
        summary = lines_df.groupby('AccountCode').agg({
            'Debit': 'sum',
            'Credit': 'sum'
        }).reset_index()
        
        # Merge with chart of accounts
        trial_balance = coa_df.merge(
            summary,
            left_on='Code',
            right_on='AccountCode',
            how='left'
        )
        
        # Fill NA values and calculate balance
        trial_balance['Debit'] = trial_balance['Debit'].fillna(0)
        trial_balance['Credit'] = trial_balance['Credit'].fillna(0)
        trial_balance['Balance'] = trial_balance['Debit'] - trial_balance['Credit']
        
        # Include Levels and Parent Account if available
        columns = ['Code', 'Name', 'AccountType']
        if 'Levels' in trial_balance.columns:
            columns.append('Levels')
        if 'FatherAccountKey' in trial_balance.columns:
            trial_balance.rename(columns={'FatherAccountKey': 'ParentAccount'}, inplace=True)
            columns.append('ParentAccount')
        
        columns.extend(['Debit', 'Credit', 'Balance'])
        
        return trial_balance[columns]
    
    def _get_profit_loss_api(self, entity: Dict, start_date: str, end_date: str) -> pd.DataFrame:
        """Generate P&L from API data"""
        trial_balance = self._get_trial_balance_api(entity, start_date, end_date)
        
        # Filter for P&L accounts (Revenue and Expenses)
        pl_accounts = trial_balance[
            trial_balance['Code'].str.startswith(('4', '5', '6'), na=False)
        ].copy()
        
        pl_accounts['Amount'] = pl_accounts['Credit'] - pl_accounts['Debit']
        
        return pl_accounts
    
    def _get_balance_sheet_api(self, entity: Dict, as_of_date: str) -> pd.DataFrame:
        """Generate Balance Sheet from API data"""
        start_date = "1900-01-01"  # Get all transactions up to as_of_date
        trial_balance = self._get_trial_balance_api(entity, start_date, as_of_date)
        
        # Filter for BS accounts (Assets, Liabilities, Equity)
        bs_accounts = trial_balance[
            trial_balance['Code'].str.startswith(('1', '2', '3'), na=False)
        ].copy()
        
        return bs_accounts
    
    def _get_sales_report_api(self, entity: Dict, start_date: str, end_date: str) -> pd.DataFrame:
        """Get Sales Report from API"""
        filter_query = f"DocDate ge '{start_date}' and DocDate le '{end_date}'"
        
        try:
            invoices = self.conn_mgr.fetch_api_data(
                entity,
                'Invoices',
                filter_query=filter_query
            )
            
            # Flatten invoice lines
            all_lines = []
            for invoice in invoices:
                for line in invoice.get('DocumentLines', []):
                    all_lines.append({
                        'InvoiceNumber': invoice.get('DocNum'),
                    'InvoiceDate': invoice.get('DocDate'),
                    'CustomerCode': invoice.get('CardCode'),
                    'CustomerName': invoice.get('CardName'),
                    'ItemCode': line.get('ItemCode'),
                    'ItemDescription': line.get('ItemDescription'),
                    'Quantity': line.get('Quantity'),
                    'UnitPrice': line.get('UnitPrice'),
                    'LineTotal': line.get('LineTotal'),
                    'InvoiceTotal': invoice.get('DocTotal'),
                    'Currency': invoice.get('DocCurrency')
                })
        
            return pd.DataFrame(all_lines)
        finally:
            # Always disconnect after the operation
            self.conn_mgr.disconnect_api(entity)
    
    # ========== Universal Extract Methods ==========
    
    def extract_trial_balance(self, entity_id: str, start_date: str = None, 
                             end_date: str = None) -> pd.DataFrame:
        """
        Extract Trial Balance for any entity
        
        Args:
            entity_id: Entity ID
            start_date: Start date (YYYY-MM-DD), defaults to config
            end_date: End date (YYYY-MM-DD), defaults to config
        """
        entity = self.conn_mgr.get_entity_by_id(entity_id)
        if not entity:
            raise ValueError(f"Entity '{entity_id}' not found")
        
        # Use default dates if not provided
        if not start_date:
            start_date = self.config['date_range']['start_date']
        if not end_date:
            end_date = self.config['date_range']['end_date']
        
        if entity['connection_type'] == 'sql':
            return self._get_trial_balance_sql(entity, start_date, end_date)
        elif entity['connection_type'] == 'api':
            return self._get_trial_balance_api(entity, start_date, end_date)
        else:
            raise ValueError(f"Unknown connection type: {entity['connection_type']}")
    
    def extract_profit_loss(self, entity_id: str, start_date: str = None, 
                           end_date: str = None) -> pd.DataFrame:
        """Extract Profit & Loss Statement"""
        entity = self.conn_mgr.get_entity_by_id(entity_id)
        if not entity:
            raise ValueError(f"Entity '{entity_id}' not found")
        
        if not start_date:
            start_date = self.config['date_range']['start_date']
        if not end_date:
            end_date = self.config['date_range']['end_date']
        
        if entity['connection_type'] == 'sql':
            return self._get_profit_loss_sql(entity, start_date, end_date)
        elif entity['connection_type'] == 'api':
            return self._get_profit_loss_api(entity, start_date, end_date)
        else:
            raise ValueError(f"Unknown connection type: {entity['connection_type']}")
    
    def extract_balance_sheet(self, entity_id: str, as_of_date: str = None) -> pd.DataFrame:
        """Extract Balance Sheet"""
        entity = self.conn_mgr.get_entity_by_id(entity_id)
        if not entity:
            raise ValueError(f"Entity '{entity_id}' not found")
        
        if not as_of_date:
            as_of_date = self.config['date_range']['end_date']
        
        if entity['connection_type'] == 'sql':
            return self._get_balance_sheet_sql(entity, as_of_date)
        elif entity['connection_type'] == 'api':
            return self._get_balance_sheet_api(entity, as_of_date)
        else:
            raise ValueError(f"Unknown connection type: {entity['connection_type']}")
    
    def extract_sales_report(self, entity_id: str, start_date: str = None, 
                            end_date: str = None) -> pd.DataFrame:
        """Extract Sales Report"""
        entity = self.conn_mgr.get_entity_by_id(entity_id)
        if not entity:
            raise ValueError(f"Entity '{entity_id}' not found")
        
        if not start_date:
            start_date = self.config['date_range']['start_date']
        if not end_date:
            end_date = self.config['date_range']['end_date']
        
        if 'sales_report' not in entity['reports_available']:
            raise ValueError(f"Sales report not available for entity '{entity_id}'")
        
        if entity['connection_type'] == 'sql':
            return self._get_sales_report_sql(entity, start_date, end_date)
        elif entity['connection_type'] == 'api':
            return self._get_sales_report_api(entity, start_date, end_date)
        else:
            raise ValueError(f"Unknown connection type: {entity['connection_type']}")
    
    def save_to_excel(self, df: pd.DataFrame, filename: str, sheet_name: str = "Data"):
        """
        Save DataFrame to Excel with formatting
        
        Args:
            df: DataFrame to save
            filename: Output filename
            sheet_name: Sheet name
        """
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Save to Excel
        df.to_excel(filename, sheet_name=sheet_name, index=False)
        
        # Format the Excel file
        wb = load_workbook(filename)
        ws = wb[sheet_name]
        
        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filename)
    
    # ========== Master Data Extraction (SQL & API) ==========
    
    def extract_chart_of_accounts(self, entity_id: str) -> pd.DataFrame:
        """Extract Chart of Accounts"""
        entity = self.conn_mgr.get_entity_by_id(entity_id)
        if not entity:
            raise ValueError(f"Entity '{entity_id}' not found")
        
        if entity['connection_type'] == 'sql':
            return self._get_chart_of_accounts_sql(entity)
        elif entity['connection_type'] == 'api':
            # Use the same method that handles entity-specific configurations
            return self._get_chart_of_accounts_api(entity)
        else:
            raise ValueError(f"Unknown connection type: {entity['connection_type']}")
    
    def extract_business_partners(self, entity_id: str) -> pd.DataFrame:
        """Extract Business Partners (Vendors/Customers)"""
        entity = self.conn_mgr.get_entity_by_id(entity_id)
        if not entity:
            raise ValueError(f"Entity '{entity_id}' not found")
        
        if entity['connection_type'] == 'sql':
            return self._get_business_partners_sql(entity)
        elif entity['connection_type'] == 'api':
            try:
                data = self.conn_mgr.fetch_api_data(entity, 'BusinessPartners')
                return pd.DataFrame(data)
            finally:
                self.conn_mgr.disconnect_api(entity)
        else:
            raise ValueError(f"Unknown connection type: {entity['connection_type']}")
    
    def extract_business_partner_groups(self, entity_id: str) -> pd.DataFrame:
        """Extract Business Partner Groups"""
        entity = self.conn_mgr.get_entity_by_id(entity_id)
        if not entity:
            raise ValueError(f"Entity '{entity_id}' not found")
        
        if entity['connection_type'] == 'sql':
            return self._get_business_partner_groups_sql(entity)
        elif entity['connection_type'] == 'api':
            try:
                data = self.conn_mgr.fetch_api_data(entity, 'BusinessPartnerGroups')
                return pd.DataFrame(data)
            finally:
                self.conn_mgr.disconnect_api(entity)
        else:
            raise ValueError(f"Unknown connection type: {entity['connection_type']}")
    
    def extract_payment_terms(self, entity_id: str) -> pd.DataFrame:
        """Extract Payment Terms"""
        entity = self.conn_mgr.get_entity_by_id(entity_id)
        if not entity:
            raise ValueError(f"Entity '{entity_id}' not found")
        
        if entity['connection_type'] == 'sql':
            return self._get_payment_terms_sql(entity)
        elif entity['connection_type'] == 'api':
            try:
                data = self.conn_mgr.fetch_api_data(entity, 'PaymentTermsTypes')
                return pd.DataFrame(data)
            finally:
                self.conn_mgr.disconnect_api(entity)
        else:
            raise ValueError(f"Unknown connection type: {entity['connection_type']}")
    
    def extract_withholding_tax(self, entity_id: str) -> pd.DataFrame:
        """Extract Withholding Tax Codes"""
        entity = self.conn_mgr.get_entity_by_id(entity_id)
        if not entity:
            raise ValueError(f"Entity '{entity_id}' not found")
        
        if entity['connection_type'] == 'sql':
            return self._get_withholding_tax_sql(entity)
        elif entity['connection_type'] == 'api':
            try:
                data = self.conn_mgr.fetch_api_data(entity, 'WithholdingTaxCodes')
                return pd.DataFrame(data)
            finally:
                self.conn_mgr.disconnect_api(entity)
        else:
            raise ValueError(f"Unknown connection type: {entity['connection_type']}")
    
    def extract_banks(self, entity_id: str) -> pd.DataFrame:
        """Extract Bank Master"""
        entity = self.conn_mgr.get_entity_by_id(entity_id)
        if not entity:
            raise ValueError(f"Entity '{entity_id}' not found")
        
        if entity['connection_type'] == 'sql':
            return self._get_banks_sql(entity)
        elif entity['connection_type'] == 'api':
            try:
                data = self.conn_mgr.fetch_api_data(entity, 'Banks')
                return pd.DataFrame(data)
            finally:
                self.conn_mgr.disconnect_api(entity)
        else:
            raise ValueError(f"Unknown connection type: {entity['connection_type']}")
    
    def extract_item_groups(self, entity_id: str) -> pd.DataFrame:
        """Extract Item Groups"""
        entity = self.conn_mgr.get_entity_by_id(entity_id)
        if not entity:
            raise ValueError(f"Entity '{entity_id}' not found")
        
        if entity['connection_type'] == 'sql':
            return self._get_item_groups_sql(entity)
        elif entity['connection_type'] == 'api':
            try:
                data = self.conn_mgr.fetch_api_data(entity, 'ItemGroups')
                return pd.DataFrame(data)
            finally:
                self.conn_mgr.disconnect_api(entity)
        else:
            raise ValueError(f"Unknown connection type: {entity['connection_type']}")
    
    def extract_hsn_sac_codes(self, entity_id: str) -> pd.DataFrame:
        """Extract HSN/SAC Codes for India (API Only)"""
        entity = self.conn_mgr.get_entity_by_id(entity_id)
        if not entity:
            raise ValueError(f"Entity '{entity_id}' not found")
        
        if entity['connection_type'] == 'sql':
            raise ValueError(f"HSN/SAC codes extraction not available for SQL entities")
        elif entity['connection_type'] == 'api':
            try:
                data = self.conn_mgr.fetch_api_data(entity, 'IndiaSacCode')
                return pd.DataFrame(data)
            finally:
                self.conn_mgr.disconnect_api(entity)
        else:
            raise ValueError(f"Unknown connection type: {entity['connection_type']}")
    
    def extract_currencies(self, entity_id: str) -> pd.DataFrame:
        """Extract Currencies"""
        entity = self.conn_mgr.get_entity_by_id(entity_id)
        if not entity:
            raise ValueError(f"Entity '{entity_id}' not found")
        
        if entity['connection_type'] == 'sql':
            return self._get_currencies_sql(entity)
        elif entity['connection_type'] == 'api':
            try:
                data = self.conn_mgr.fetch_api_data(entity, 'Currencies')
                return pd.DataFrame(data)
            finally:
                self.conn_mgr.disconnect_api(entity)
        else:
            raise ValueError(f"Unknown connection type: {entity['connection_type']}")
    
    def extract_tax_codes(self, entity_id: str) -> pd.DataFrame:
        """Extract Sales Tax Codes"""
        entity = self.conn_mgr.get_entity_by_id(entity_id)
        if not entity:
            raise ValueError(f"Entity '{entity_id}' not found")
        
        if entity['connection_type'] == 'sql':
            return self._get_tax_codes_sql(entity)
        elif entity['connection_type'] == 'api':
            try:
                data = self.conn_mgr.fetch_api_data(entity, 'SalesTaxCodes')
                return pd.DataFrame(data)
            finally:
                self.conn_mgr.disconnect_api(entity)
        else:
            raise ValueError(f"Unknown connection type: {entity['connection_type']}")
    
    def extract_cost_centers(self, entity_id: str) -> pd.DataFrame:
        """Extract Cost Centers (Profit Centers)"""
        entity = self.conn_mgr.get_entity_by_id(entity_id)
        if not entity:
            raise ValueError(f"Entity '{entity_id}' not found")
        
        if entity['connection_type'] == 'sql':
            return self._get_cost_centers_sql(entity)
        elif entity['connection_type'] == 'api':
            try:
                data = self.conn_mgr.fetch_api_data(entity, 'ProfitCenters')
                return pd.DataFrame(data)
            finally:
                self.conn_mgr.disconnect_api(entity)
        else:
            raise ValueError(f"Unknown connection type: {entity['connection_type']}")
    
    def extract_warehouse_locations(self, entity_id: str) -> pd.DataFrame:
        """Extract Warehouse Locations"""
        entity = self.conn_mgr.get_entity_by_id(entity_id)
        if not entity:
            raise ValueError(f"Entity '{entity_id}' not found")
        
        if entity['connection_type'] == 'sql':
            return self._get_warehouse_locations_sql(entity)
        elif entity['connection_type'] == 'api':
            try:
                data = self.conn_mgr.fetch_api_data(entity, 'WarehouseLocations')
                return pd.DataFrame(data)
            finally:
                self.conn_mgr.disconnect_api(entity)
        else:
            raise ValueError(f"Unknown connection type: {entity['connection_type']}")
    
    def extract_journal_entries(self, entity_id: str, start_date: str = None, 
                                 end_date: str = None) -> pd.DataFrame:
        """
        Extract General Ledger Journal Entries
        
        Args:
            entity_id: Entity ID
            start_date: Start date (YYYY-MM-DD), defaults to config
            end_date: End date (YYYY-MM-DD), defaults to config
        
        Returns:
            DataFrame with journal entries
        """
        # Use default dates if not provided
        if not start_date:
            start_date = self.config['date_range']['start_date']
        if not end_date:
            end_date = self.config['date_range']['end_date']
        
        entity = self.conn_mgr.get_entity_by_id(entity_id)
        if not entity:
            raise ValueError(f"Entity '{entity_id}' not found")
        
        if entity['connection_type'] == 'sql':
            return self._get_journal_entries_sql(entity, start_date, end_date)
        elif entity['connection_type'] == 'api':
            return self._get_journal_entries_api(entity, start_date, end_date)
        else:
            raise ValueError(f"Unknown connection type: {entity['connection_type']}")
        
    def save_to_csv(self, df: pd.DataFrame, filename: str):
        """Save DataFrame to CSV"""
        df.to_csv(filename, index=False)
    
    # ========== Schedule III Financial Statements ==========
    
    def extract_schedule3_balance_sheet(self, entity_id: str, as_at_date: str = None) -> pd.DataFrame:
        """
        Extract Balance Sheet in Schedule III format (Companies Act, 2013)
        
        Args:
            entity_id: Entity ID
            as_at_date: Balance Sheet date (YYYY-MM-DD), defaults to config end date
        
        Returns:
            DataFrame formatted as per Schedule III Balance Sheet
        """
        from backend.sap_connect.schedule3_formatter import Schedule3BalanceSheet
        
        # Use default date if not provided
        if not as_at_date:
            as_at_date = self.config['date_range']['end_date']
        
        # Get Trial Balance data
        start_date = self.config['date_range']['start_date']
        tb_df = self.extract_trial_balance(entity_id, start_date, as_at_date)
        
        # Get entity details
        entity = self.conn_mgr.get_entity_by_id(entity_id)
        entity_name = entity['name'] if entity else entity_id
        
        # Format as Schedule III (pass entity_id for custom mapping)
        formatter = Schedule3BalanceSheet(tb_df, entity_name, as_at_date, entity_id)
        return formatter.generate_balance_sheet()
    
    def extract_schedule3_profit_loss(self, entity_id: str, start_date: str = None, 
                                      end_date: str = None) -> pd.DataFrame:
        """
        Extract Profit & Loss Statement in Schedule III format (Companies Act, 2013)
        
        Args:
            entity_id: Entity ID
            start_date: Period start date (YYYY-MM-DD), defaults to config
            end_date: Period end date (YYYY-MM-DD), defaults to config
        
        Returns:
            DataFrame formatted as per Schedule III P&L Statement
        """
        from backend.sap_connect.schedule3_formatter import Schedule3ProfitLoss
        
        # Use default dates if not provided
        if not start_date:
            start_date = self.config['date_range']['start_date']
        if not end_date:
            end_date = self.config['date_range']['end_date']
        
        # Get Profit & Loss data
        pl_df = self.extract_profit_loss(entity_id, start_date, end_date)
        
        # Get entity details
        entity = self.conn_mgr.get_entity_by_id(entity_id)
        entity_name = entity['name'] if entity else entity_id
        
        # Format as Schedule III (pass entity_id for custom mapping)
        formatter = Schedule3ProfitLoss(pl_df, entity_name, start_date, end_date, entity_id)
        return formatter.generate_profit_loss()
