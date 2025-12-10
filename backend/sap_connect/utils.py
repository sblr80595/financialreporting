"""
Utility Functions and Classes
==============================
Purpose: Common utilities for formatting, logging, and data processing
Author: SAP Connection Module
Date: November 2025
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from typing import Dict, Optional
import os


class Logger:
    """Simple logger for console output."""
    
    def __init__(self, name: str = "SAP"):
        self.name = name
        self.log_file = None
    
    def _log(self, level: str, message: str, symbol: str = "•"):
        """Internal logging method."""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_message = f"[{timestamp}] [{self.name}] {symbol} {message}"
        print(log_message)
        
        if self.log_file:
            with open(self.log_file, 'a') as f:
                f.write(log_message + '\n')
    
    def info(self, message: str):
        """Log info message."""
        self._log("INFO", message, "ℹ")
    
    def success(self, message: str):
        """Log success message."""
        self._log("SUCCESS", message, "✓")
    
    def warning(self, message: str):
        """Log warning message."""
        self._log("WARNING", message, "⚠")
    
    def error(self, message: str):
        """Log error message."""
        self._log("ERROR", message, "✗")
    
    def header(self, message: str):
        """Log header message."""
        separator = "=" * 80
        print(f"\n{separator}")
        print(message.center(80))
        print(separator)
    
    def set_log_file(self, log_file: str):
        """Set log file path."""
        self.log_file = log_file
        os.makedirs(os.path.dirname(log_file), exist_ok=True)


class ExcelFormatter:
    """Excel file formatting utilities."""
    
    def __init__(self):
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.header_font = Font(color='FFFFFF', bold=True)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def format_sheet(self, sheet):
        """Apply formatting to a worksheet."""
        # Format headers
        for cell in sheet[1]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze top row
        sheet.freeze_panes = 'A2'
    
    def save_multiple_sheets(self, filename: str, dataframes: Dict[str, pd.DataFrame]):
        """
        Save multiple DataFrames to Excel with formatting.
        
        Args:
            filename: Output Excel filename
            dataframes: Dictionary of sheet_name -> DataFrame
        """
        # Filter out empty dataframes
        valid_dfs = {name: df for name, df in dataframes.items() if not df.empty}
        
        if not valid_dfs:
            print("  Warning: No data to save to Excel")
            return
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            for sheet_name, df in valid_dfs.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Format all sheets
            workbook = writer.book
            for sheet_name in workbook.sheetnames:
                self.format_sheet(workbook[sheet_name])


class DataProcessor:
    """Data processing utilities."""
    
    @staticmethod
    def flatten_journal_entries(df: pd.DataFrame) -> pd.DataFrame:
        """
        Flatten journal entries DataFrame by expanding JournalEntryLines.
        
        Args:
            df: DataFrame with JournalEntryLines column
            
        Returns:
            Flattened DataFrame
        """
        flattened_rows = []
        
        for idx, row in df.iterrows():
            header_fields = {
                'JdtNum': row.get('JdtNum'),
                'Number': row.get('Number'),
                'Reference': row.get('Reference'),
                'Memo': row.get('Memo'),
                'TaxDate': row.get('TaxDate'),
                'DueDate': row.get('DueDate'),
                'TransactionCode': row.get('TransactionCode'),
                'ProjectCode': row.get('ProjectCode'),
            }
            
            lines_str = row.get('JournalEntryLines', '[]')
            
            try:
                import ast
                if isinstance(lines_str, str):
                    lines = ast.literal_eval(lines_str)
                else:
                    lines = lines_str if isinstance(lines_str, list) else []
            except (ValueError, SyntaxError):
                lines = []
            
            if not lines:
                flattened_rows.append(header_fields)
            else:
                for line in lines:
                    combined_row = {
                        **header_fields,
                        'Line_ID': line.get('Line_ID'),
                        'AccountCode': line.get('AccountCode'),
                        'AccountName': line.get('ShortName'),
                        'Debit': line.get('Debit', 0.0),
                        'Credit': line.get('Credit', 0.0),
                        'LineMemo': line.get('LineMemo'),
                        'ContraAccount': line.get('ContraAccount'),
                        'Reference1': line.get('Reference1'),
                        'Reference2': line.get('Reference2'),
                        'ProjectCode_Line': line.get('ProjectCode'),
                        'CostingCode': line.get('CostingCode'),
                        'BPLID': line.get('BPLID'),
                        'BPLName': line.get('BPLName'),
                        'DebitSys': line.get('DebitSys', 0.0),
                        'CreditSys': line.get('CreditSys', 0.0),
                    }
                    flattened_rows.append(combined_row)
        
        return pd.DataFrame(flattened_rows)
    
    @staticmethod
    def create_summary(df: pd.DataFrame) -> pd.DataFrame:
        """
        Create summary view of journal entries.
        
        Args:
            df: DataFrame with JournalEntryLines column
            
        Returns:
            Summary DataFrame
        """
        summary_rows = []
        
        for idx, row in df.iterrows():
            lines_str = row.get('JournalEntryLines', '[]')
            
            try:
                import ast
                if isinstance(lines_str, str):
                    lines = ast.literal_eval(lines_str)
                else:
                    lines = lines_str if isinstance(lines_str, list) else []
            except (ValueError, SyntaxError):
                lines = []
            
            total_debit = sum(line.get('Debit', 0.0) for line in lines)
            total_credit = sum(line.get('Credit', 0.0) for line in lines)
            num_lines = len(lines)
            account_codes = list(set(line.get('AccountCode', '') for line in lines if line.get('AccountCode')))
            
            summary_rows.append({
                'JdtNum': row.get('JdtNum'),
                'Number': row.get('Number'),
                'TaxDate': row.get('TaxDate'),
                'Reference': row.get('Reference'),
                'Memo': row.get('Memo'),
                'NumLines': num_lines,
                'TotalDebit': total_debit,
                'TotalCredit': total_credit,
                'AccountCodes': ', '.join(account_codes[:5]) + ('...' if len(account_codes) > 5 else ''),
                'ProjectCode': row.get('ProjectCode'),
                'TransactionCode': row.get('TransactionCode'),
            })
        
        return pd.DataFrame(summary_rows)
