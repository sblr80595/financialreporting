"""
Trial Balance 6-Rule Validation System
- Validates trial balance data against 6 fundamental accounting rules
- Generates compliance report in Excel format
- Checks debits/credits balance, account types, duplicates, and data integrity

Structured Trial Balance Validation Rules (India Accounting Standards):
1. Total Debits Equal Total Credits
   - Sum of all debit amounts must equal sum of all credit amounts
2. Balance Calculation Accuracy
   - Each row's Balance must equal (Debit - Credit)
3. No Duplicate Accounts
   - Each G/L Acct or BP Code must be unique
4. No Missing or Invalid Data
   - All rows must have valid codes, names, and numeric values
5. Logical Balance Signs by Account Type:
   - Assets (1xxxxx): Positive balance
   - Liabilities (2xxxxx): Negative balance
   - Revenue (3xxxxx): Negative balance
   - Expenses (4xxxxx): Positive balance
   - Equity (5xxxxx): Negative balance
6. Accounting Equation Validation
   - Assets = -(Liabilities + Equity + Revenue + Expenses)
   - Assets value must be positive

Integration: Step 5 in the Trial Balance Processing workflow
- Reads from: data/{entity}/output/adjusted-trialbalance/final_trialbalance.xlsx OR final_trial_balance.xlsx (output of Step 4)
- Outputs to: data/{entity}/output/adjusted-trialbalance/trialbalance_6rule_validation_report.xlsx
"""

import pandas as pd
import numpy as np
import os
import json
import sys
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Add project root to Python path for imports
project_root = Path(__file__).parent.parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

# Import PathService from services
from backend.services.path_service import PathService

# Default tolerance for floating-point comparisons (can be overridden by config)
DEFAULT_TOLERANCE = 0.01

# Currency symbol for reporting
CURRENCY_SYMBOL = "₹"


def load_validation_rules_config(entity: str) -> dict:
    """
    Load entity-specific validation rules configuration.

    Args:
        entity: Entity code (e.g., 'cpm', 'lifeline_diagnostics')

    Returns:
        Dictionary with validation rules configuration
    """
    try:
        # First, try to load validation_rules_config.json (new approach)
        rules_config_path = Path(__file__).parent.parent.parent / "data" / entity / "input" / "config" / "validation_rules_config.json"

        if rules_config_path.exists():
            with open(rules_config_path, 'r') as f:
                rules_config = json.load(f)

            print(f"  ✓ Loaded validation rules config for {entity}")
            print(f"    Entity: {rules_config.get('entity_name', entity)}")

            # Count enabled rules
            enabled_rules = sum(1 for rule in rules_config.get('validation_rules', {}).values() if rule.get('enabled', True))
            total_rules = len(rules_config.get('validation_rules', {}))
            print(f"    Enabled rules: {enabled_rules}/{total_rules}")

            return rules_config
        else:
            print(f"  ⚠️  No validation_rules_config.json found for {entity}, using defaults (all rules enabled)")
            # Return default config with all rules enabled
            return {
                "entity": entity,
                "entity_name": entity.replace('_', ' ').title(),
                "validation_rules": {
                    "rule_1": {
                        "enabled": True,
                        "rule_number": 1,
                        "rule_name": "Total Debits Equal Total Credits",
                        "description": "The sum of all debit amounts must equal the sum of all credit amounts",
                        "category": "Balance Validation",
                        "severity": "critical",
                        "notes": "This rule ensures the fundamental accounting equation is maintained"
                    },
                    "rule_2": {
                        "enabled": True,
                        "rule_number": 2,
                        "rule_name": "Balance Calculation Accuracy",
                        "description": "Each row's Balance must equal (Debit - Credit)",
                        "category": "Data Integrity",
                        "severity": "critical",
                        "notes": "Ensures individual GL account balances are correctly calculated"
                    },
                    "rule_3": {
                        "enabled": True,
                        "rule_number": 3,
                        "rule_name": "No Duplicate Accounts",
                        "description": "Each G/L Acct/BP Code must be unique",
                        "category": "Data Integrity",
                        "severity": "critical",
                        "notes": "Prevents duplicate GL codes which can cause reporting errors"
                    },
                    "rule_4": {
                        "enabled": True,
                        "rule_number": 4,
                        "rule_name": "No Missing or Invalid Data",
                        "description": "All rows must have valid codes, names, and numeric values",
                        "category": "Data Integrity",
                        "severity": "critical",
                        "notes": "Ensures completeness of trial balance data"
                    },
                    "rule_5": {
                        "enabled": True,
                        "rule_number": 5,
                        "rule_name": "Logical Balance Signs by Account Type",
                        "description": "Assets (1xxxxx): Positive | Liabilities (2xxxxx): Negative | Revenue (3xxxxx): Negative | Expenses (4xxxxx): Positive | Equity (5xxxxx): Negative",
                        "category": "Business Logic",
                        "severity": "warning",
                        "notes": "Validates account balances have expected signs based on account type"
                    },
                    "rule_6": {
                        "enabled": True,
                        "rule_number": 6,
                        "rule_name": "Accounting Equation Validation",
                        "description": "Assets = -(Liabilities + Equity + Revenue + Expenses)",
                        "category": "Balance Validation",
                        "severity": "critical",
                        "notes": "Ensures the fundamental accounting equation balances"
                    }
                },
                "tolerance_settings": {
                    "percentage_tolerance": 0.001,
                    "absolute_tolerance": 0.01,
                    "description": "Allowed variance as percentage (0.001%) and absolute rupee tolerance (0.01)"
                },
                "metadata": {
                    "last_updated": datetime.now().strftime("%Y-%m-%d"),
                    "updated_by": "System Default",
                    "version": "1.0"
                }
            }

    except Exception as e:
        print(f"  ⚠️  Error loading validation rules config: {e}")
        # Return default config
        return {
            "entity": entity,
            "entity_name": entity.replace('_', ' ').title(),
            "validation_rules": {
                "rule_1": {
                    "enabled": True,
                    "rule_number": 1,
                    "rule_name": "Total Debits Equal Total Credits",
                    "description": "The sum of all debit amounts must equal the sum of all credit amounts",
                    "category": "Balance Validation",
                    "severity": "critical",
                    "notes": "This rule ensures the fundamental accounting equation is maintained"
                },
                "rule_2": {
                    "enabled": True,
                    "rule_number": 2,
                    "rule_name": "Balance Calculation Accuracy",
                    "description": "Each row's Balance must equal (Debit - Credit)",
                    "category": "Data Integrity",
                    "severity": "critical",
                    "notes": "Ensures individual GL account balances are correctly calculated"
                },
                "rule_3": {
                    "enabled": True,
                    "rule_number": 3,
                    "rule_name": "No Duplicate Accounts",
                    "description": "Each G/L Acct/BP Code must be unique",
                    "category": "Data Integrity",
                    "severity": "critical",
                    "notes": "Prevents duplicate GL codes which can cause reporting errors"
                },
                "rule_4": {
                    "enabled": True,
                    "rule_number": 4,
                    "rule_name": "No Missing or Invalid Data",
                    "description": "All rows must have valid codes, names, and numeric values",
                    "category": "Data Integrity",
                    "severity": "critical",
                    "notes": "Ensures completeness of trial balance data"
                },
                "rule_5": {
                    "enabled": True,
                    "rule_number": 5,
                    "rule_name": "Logical Balance Signs by Account Type",
                    "description": "Assets (1xxxxx): Positive | Liabilities (2xxxxx): Negative | Revenue (3xxxxx): Negative | Expenses (4xxxxx): Positive | Equity (5xxxxx): Negative",
                    "category": "Business Logic",
                    "severity": "warning",
                    "notes": "Validates account balances have expected signs based on account type"
                },
                "rule_6": {
                    "enabled": True,
                    "rule_number": 6,
                    "rule_name": "Accounting Equation Validation",
                    "description": "Assets = -(Liabilities + Equity + Revenue + Expenses)",
                    "category": "Balance Validation",
                    "severity": "critical",
                    "notes": "Ensures the fundamental accounting equation balances"
                }
            },
            "tolerance_settings": {
                "percentage_tolerance": 0.001,
                "absolute_tolerance": 0.01,
                "description": "Allowed variance as percentage (0.001%) and absolute rupee tolerance (0.01)"
            },
            "metadata": {
                "last_updated": datetime.now().strftime("%Y-%m-%d"),
                "updated_by": "System Default",
                "version": "1.0"
            }
        }


def load_tolerance_from_config(entity: str) -> float:
    """
    Load tolerance variance from validation_rules_config.json or adjustment_config.json

    Args:
        entity: Entity code (e.g., 'cpm', 'hausen')

    Returns:
        Tolerance value as decimal (e.g., 0.1% = 0.001)
    """
    try:
        # First try validation_rules_config.json
        rules_config_path = Path(__file__).parent.parent.parent / "data" / entity / "input" / "config" / "validation_rules_config.json"
        if rules_config_path.exists():
            with open(rules_config_path, 'r') as f:
                config = json.load(f)

            tolerance_pct = config.get('tolerance_settings', {}).get('percentage_tolerance', 0.001)
            tolerance = tolerance_pct
            print(f"  Using tolerance from validation_rules_config: {tolerance_pct*100:.4f}% ({tolerance})")
            return tolerance

        # Fallback to adjustment_config.json
        config_path = Path(__file__).parent.parent.parent / "data" / entity / "input" / "config" / "adjustment_config.json"

        if config_path.exists():
            with open(config_path, 'r') as f:
                config = json.load(f)

            # Get tolerance percentage from config
            tolerance_pct = config.get('validation_settings', {}).get('tolerance_variance_percentage', 0.1)

            # Convert percentage to decimal (0.1% = 0.001)
            tolerance = tolerance_pct / 100.0

            print(f"  Using tolerance from adjustment_config: {tolerance_pct}% ({tolerance})")
            return tolerance
        else:
            print(f"  Config file not found, using default tolerance: {DEFAULT_TOLERANCE}")
            return DEFAULT_TOLERANCE

    except Exception as e:
        print(f"  Error loading config, using default tolerance: {e}")
        return DEFAULT_TOLERANCE


class TrialBalanceValidator:
    """Main class for tb 6-rule validation"""
    
    def __init__(self, entity: str = "cpm", input_file=None, output_file=None):
        """
        Initialize validator
        Args:
            entity: Entity code (e.g., 'cpm', 'hausen')
            input_file: Path to input file (Excel or CSV). Defaults to entity final TB
            output_file: Path to output file. Defaults to entity validation report
        """
        self.entity = entity
        self.path_service = PathService(entity)
        self.path_service.create_entity_structure(entity)

        # Load entity-specific validation rules configuration
        self.rules_config = load_validation_rules_config(entity)

        # Load percentage tolerance from config (e.g., 0.001 means 0.001%)
        # load_tolerance_from_config returns a decimal fraction (percentage/100),
        # e.g., config value 0.001 -> returns 0.00001.
        self.tolerance_pct = load_tolerance_from_config(entity)
        # Absolute rupee tolerance for row-level checks (fixed or from config)
        self.abs_tolerance = self.rules_config.get('tolerance_settings', {}).get('absolute_tolerance', DEFAULT_TOLERANCE)
        
        # Set default paths if not provided
        if input_file is None:
            # Try to find the final trial balance file with various naming patterns
            input_file = self._find_final_trial_balance()
        if output_file is None:
            output_file = str(self.path_service.get_validation_report_path(entity))
        
        self.input_file = input_file
        self.output_file = output_file
        self.df = None
        self.validation_results = {}
        self.violations = {}
    
    def _find_final_trial_balance(self):
        """
        Find the final trial balance file, checking for different naming patterns:
        - final_trialbalance.xlsx
        - final_trial_balance.xlsx
        """
        adjusted_tb_dir = self.path_service.get_adjusted_tb_dir(self.entity)
        
        # Check for different filename variations
        possible_filenames = [
            'final_trialbalance.xlsx',
            'final_trial_balance.xlsx',
        ]
        
        for filename in possible_filenames:
            file_path = adjusted_tb_dir / filename
            if file_path.exists():
                print(f"✓ Found final trial balance: {filename}")
                return str(file_path)
        
        # If no file found, return the default path (will fail later with proper error)
        default_path = str(self.path_service.get_final_tb_path(self.entity))
        print(f"⚠️ Warning: Final trial balance not found, using default: {default_path}")
        return default_path
        
    def load_trial_balance(self):
        """Load trial balance from Excel or CSV file"""
        print(f"\nLoading trial balance from: {self.input_file}")
        
        # Ensure output directory exists
        output_dir = self.path_service.get_adjusted_tb_dir(self.entity)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Determine file type and read accordingly
        file_ext = Path(self.input_file).suffix.lower()
        
        if file_ext in ['.xlsx', '.xls', '.xlsb']:
            # Read Excel file
            self.df = pd.read_excel(self.input_file)
            print(f"✓ Loaded Excel file with {len(self.df)} records")
        elif file_ext == '.csv':
            # Read CSV, skipping comment lines (lines starting with #)
            self.df = pd.read_csv(self.input_file, comment='#')
            print(f"✓ Loaded CSV file with {len(self.df)} records")
        else:
            raise ValueError(f"Unsupported file format: {file_ext}")
        
        # Clean column names (remove trailing spaces)
        self.df.columns = self.df.columns.str.strip()
        
        # Map column names from different formats
        column_mapping = {
            "GL Code": "GL_Code",
            "GL Code Description": "Description",
            "GL Description": "Description",  # Added mapping for actual column name
            "Adjusted Mar'25": "Balance",
            "Mar'25 Adjusted": "Balance",  # March adjusted
            "Jun'25 Adjusted": "Balance",  # June adjusted
            "Mar'25": "Original_Balance",
            "(Unaudited) Mar'25": "Original_Balance",  # Unaudited March
            "(Unaudited) Jun'25": "Original_Balance",  # Unaudited June
            "BSPL": "BSPL",
            "Ind AS Major": "Major_Category",
            "Ind AS Minor": "Minor_Category"
        }
        
        # Rename columns if they exist in the mapping
        rename_dict = {}
        for old_col, new_col in column_mapping.items():
            if old_col in self.df.columns:
                rename_dict[old_col] = new_col
        
        if rename_dict:
            self.df = self.df.rename(columns=rename_dict)
            print(f"  Renamed columns: {list(rename_dict.keys())}")
        
        # Ensure required columns exist
        required_cols = ["GL_Code", "Description", "Balance"]
        missing_cols = [col for col in required_cols if col not in self.df.columns]
        
        if missing_cols:
            raise ValueError(f"Missing required columns: {missing_cols}. Available columns: {list(self.df.columns)}")
        
        # Ensure Balance is numeric
        self.df["Balance"] = pd.to_numeric(self.df["Balance"], errors='coerce').fillna(0)
        
        # If Debit and Credit columns don't exist, calculate them from Balance
        if "Debit" not in self.df.columns or "Credit" not in self.df.columns:
            print("  Calculating Debit/Credit columns from Balance...")
            self.df["Debit"] = self.df["Balance"].apply(lambda x: x if x > 0 else 0)
            self.df["Credit"] = self.df["Balance"].apply(lambda x: abs(x) if x < 0 else 0)
        
        print(f"  Columns: {list(self.df.columns)}")
        
        return self.df
    
    def get_account_type(self, gl_code):
        """
        Determine account type based on GL code prefix (simplified)
        - 1xxxxx: Asset accounts
        - 2xxxxx: Liability accounts
        - 3xxxxx: Revenue accounts
        - 4xxxxx: Expense accounts
        - 5xxxxx: Equity accounts
        Returns: account type string
        """
        if pd.isna(gl_code):
            return "Unknown"
        
        gl_str = str(gl_code).strip()
        
        # Get first digit
        if len(gl_str) == 0:
            return "Unknown"
        
        first_digit = gl_str[0]
        
        if first_digit == '1':
            return "Asset"
        elif first_digit == '2':
            return "Liability"
        elif first_digit == '3':
            return "Revenue"
        elif first_digit == '4':
            return "Expense"
        elif first_digit == '5':
            return "Equity"
        else:
            return "Unknown"
    
    def is_share_capital_account(self, gl_code):
        """
        Detect Share Capital accounts (starting with 500000)
        Returns: True if account is a share capital account
        """
        if pd.isna(gl_code):
            return False
        
        gl_str = str(gl_code).strip()
        return gl_str.startswith('500000')
    
    def rule_1_total_debits_equal_credits(self):
        """
        Rule 1: Total Debits Equal Total Credits
        The sum of all debit amounts must equal the sum of all credit amounts.
        """
        print("\n" + "="*80)
        print("RULE 1: Total Debits Equal Total Credits")
        print("="*80)
        
        total_debits = self.df["Debit"].sum()
        total_credits = self.df["Credit"].sum()
        difference = total_debits - total_credits
        
        # Use percentage-based allowed variance based on total debits (if available)
        allowed_variance = abs(total_debits) * self.tolerance_pct if abs(total_debits) > 0 else self.abs_tolerance
        is_balanced = abs(difference) <= allowed_variance
        
        result = {
            "rule_number": 1,
            "rule_name": "Total Debits Equal Total Credits",
            "status": "✓ PASS" if is_balanced else "✗ FAIL",
            "total_debits": total_debits,
            "total_credits": total_credits,
            "difference": difference,
            "tolerance_pct": self.tolerance_pct,
            "allowed_variance": allowed_variance,
            "is_compliant": is_balanced,
            "details": f"Debits: {CURRENCY_SYMBOL}{total_debits:,.2f}, Credits: {CURRENCY_SYMBOL}{total_credits:,.2f}, Difference: {CURRENCY_SYMBOL}{difference:,.2f}"
        }
        
        print(f"Total Debits:     {CURRENCY_SYMBOL}{total_debits:>20,.2f}")
        print(f"Total Credits:    {CURRENCY_SYMBOL}{total_credits:>20,.2f}")
        print(f"Difference:       {CURRENCY_SYMBOL}{difference:>20,.2f}")
        print(f"Tolerance pct:    {self.tolerance_pct*100:.6f}% (amount allowed: {CURRENCY_SYMBOL}{allowed_variance:,.2f})")
        print(f"Status: {result['status']}")
        
        self.validation_results["rule_1"] = result
        return result
    
    def rule_2_balance_calculation_accuracy(self):
        """
        Rule 2: Balance Calculation Accuracy
        Each row's Balance must equal (Debit - Credit).
        """
        print("\n" + "="*80)
        print("RULE 2: Balance Calculation Accuracy")
        print("="*80)
        
        # Calculate expected balance: Balance = Debit - Credit
        self.df["Computed_Balance"] = self.df["Debit"] - self.df["Credit"]
        self.df["Balance_Diff"] = abs(self.df["Balance"] - self.df["Computed_Balance"])
        
        # Find violations using absolute rupee tolerance
        violations = self.df[self.df["Balance_Diff"] > self.abs_tolerance].copy()
        
        is_compliant = len(violations) == 0
        
        result = {
            "rule_number": 2,
            "rule_name": "Balance Calculation Accuracy",
            "status": "✓ PASS" if is_compliant else "✗ FAIL",
            "total_records": len(self.df),
            "violations_count": len(violations),
            "is_compliant": is_compliant,
            "details": f"Found {len(violations)} records with incorrect balance calculations"
        }
        
        print(f"Total Records:    {len(self.df):>10}")
        print(f"Violations:       {len(violations):>10}")
        print(f"Status: {result['status']}")
        
        if len(violations) > 0:
            print(f"\nSample violations (first 5):")
            for idx, row in violations.head(5).iterrows():
                print(f"  GL Code: {row['GL_Code']}, Balance: {row['Balance']:.2f}, " +
                      f"Computed: {row['Computed_Balance']:.2f}, Diff: {row['Balance_Diff']:.2f}")
            
            self.violations["rule_2"] = violations[[
                "GL_Code", "Description", "Debit", "Credit", "Balance", 
                "Computed_Balance", "Balance_Diff"
            ]].copy()
        
        self.validation_results["rule_2"] = result
        return result
    
    def rule_3_no_duplicate_accounts(self):
        """
        Rule 3: No Duplicate Accounts
        Each G/L Acct/BP Code must be unique.
        """
        print("\n" + "="*80)
        print("RULE 3: No Duplicate Accounts")
        print("="*80)
        
        # Find duplicates
        duplicates = self.df[self.df.duplicated(subset=["GL_Code"], keep=False)].copy()
        duplicates = duplicates.sort_values("GL_Code")
        
        duplicate_codes = duplicates["GL_Code"].unique()
        
        is_compliant = len(duplicate_codes) == 0
        
        result = {
            "rule_number": 3,
            "rule_name": "No Duplicate Accounts",
            "status": "✓ PASS" if is_compliant else "✗ FAIL",
            "total_records": len(self.df),
            "duplicate_gl_codes": len(duplicate_codes),
            "total_duplicate_records": len(duplicates),
            "is_compliant": is_compliant,
            "details": f"Found {len(duplicate_codes)} GL codes with duplicates ({len(duplicates)} total records)"
        }
        
        print(f"Total Records:         {len(self.df):>10}")
        print(f"Duplicate GL Codes:    {len(duplicate_codes):>10}")
        print(f"Total Duplicate Rows:  {len(duplicates):>10}")
        print(f"Status: {result['status']}")
        
        if len(duplicates) > 0:
            print(f"\nDuplicate GL Codes: {list(duplicate_codes[:10])}" + 
                  ("..." if len(duplicate_codes) > 10 else ""))
            
            self.violations["rule_3"] = duplicates[["GL_Code", "Description", "Balance"]].copy()
        
        self.validation_results["rule_3"] = result
        return result
    
    def rule_4_no_missing_invalid_data(self):
        """
        Rule 4: No Missing or Invalid Data
        All rows must have valid codes, names, and numeric values.
        """
        print("\n" + "="*80)
        print("RULE 4: No Missing or Invalid Data")
        print("="*80)
        
        # Find rows with missing or invalid data
        issues = self.df[
            self.df[['GL_Code', 'Description', 'Debit', 'Credit', 'Balance']].isnull().any(axis=1) |
            ~self.df[['Debit', 'Credit', 'Balance']].apply(lambda x: x.apply(np.isreal)).all(axis=1)
        ]
        
        is_compliant = len(issues) == 0
        
        result = {
            "rule_number": 4,
            "rule_name": "No Missing or Invalid Data",
            "status": "✓ PASS" if is_compliant else "✗ FAIL",
            "total_records": len(self.df),
            "issues_count": len(issues),
            "is_compliant": is_compliant,
            "details": f"Found {len(issues)} records with missing or invalid data"
        }
        
        print(f"Total Records: {len(self.df):>10}")
        print(f"Issues Found:  {len(issues):>10}")
        print(f"Status: {result['status']}")
        
        if len(issues) > 0:
            print(f"\nSample issues (first 5):")
            for idx, row in issues.head(5).iterrows():
                print(f"  GL Code: {row['GL_Code']}, Description: {row.get('Description', 'N/A')}")
            
            self.violations["rule_4"] = issues[["GL_Code", "Description", "Debit", "Credit", "Balance"]].copy()
        
        self.validation_results["rule_4"] = result
        return result
    
    def rule_5_logical_balance_signs(self):
        """
        Rule 5: Logical Balance Signs by Account Type (Simplified)
        - Assets (1xxxxxx): Positive balance
        - Liabilities (2xxxxxx): Negative balance
        - Revenue (3xxxxxx): Negative balance
        - Expenses (4xxxxxx): Positive balance
        - Equity (5xxxxxx): Negative balance
        
        Ignores zero balances.
        """
        print("\n" + "="*80)
        print("RULE 6: Logical Balance Signs by Account Type")
        print("="*80)
        
        # Add account type column
        self.df["Account_Type"] = self.df["GL_Code"].apply(self.get_account_type)
        
        violations = []
        
        for _, row in self.df.iterrows():
            code = str(row['GL_Code'])
            balance = row['Balance']
            
            # Skip zero or near-zero balances using absolute tolerance
            if abs(balance) <= self.abs_tolerance:
                continue
            
            # Check balance signs
            if code.startswith('1') and balance < 0:
                violations.append(row)
            elif code.startswith('2') and balance > 0:
                violations.append(row)
            elif code.startswith('5') and balance > 0:  # Equity
                violations.append(row)
            elif code.startswith('3') and balance > 0:  # Revenue
                violations.append(row)
            elif code.startswith('4') and balance < 0:  # Expense
                violations.append(row)
        
        violations_df = pd.DataFrame(violations)
        is_compliant = len(violations_df) == 0
        
        result = {
            "rule_number": 5,
            "rule_name": "Logical Balance Signs by Account Type",
            "status": "✓ PASS" if is_compliant else "✗ FAIL",
            "total_records": len(self.df),
            "violations_count": len(violations_df),
            "is_compliant": is_compliant,
            "details": f"Found {len(violations_df)} accounts with unexpected balance signs"
        }
        
        print(f"Total Records:    {len(self.df):>10}")
        print(f"Violations:       {len(violations_df):>10}")
        print(f"Status: {result['status']}")
        
        if len(violations_df) > 0:
            print(f"\nSample violations (first 10):")
            for idx, row in violations_df.head(10).iterrows():
                print(f"  {row['Account_Type']:10} GL Code: {row['GL_Code']}, Balance: {row['Balance']:>15,.2f}")
            
            self.violations["rule_5"] = violations_df[["GL_Code", "Description", "Account_Type", "Balance"]].copy()
        
        self.validation_results["rule_5"] = result
        return result
    
    def rule_6_accounting_equation(self):
        """
        Rule 6: Accounting Equation Validation
        Assets = -(Liabilities + Equity + Revenue + Expenses)
        Or: Assets + Liabilities + Equity + Revenue + Expenses = 0
        Additionally: Assets should be positive
        """
        print("\n" + "="*80)
        print("RULE 6: Accounting Equation Validation")
        print("="*80)
        
        # Ensure Account_Type column exists
        if "Account_Type" not in self.df.columns:
            self.df["Account_Type"] = self.df["GL_Code"].apply(self.get_account_type)
        
        # Calculate sums by account type
        assets = self.df[self.df['GL_Code'].astype(str).str.startswith('1')]['Balance'].sum()
        liabilities = self.df[self.df['GL_Code'].astype(str).str.startswith('2')]['Balance'].sum()
        equity = self.df[self.df['GL_Code'].astype(str).str.startswith('5')]['Balance'].sum()
        revenue = self.df[self.df['GL_Code'].astype(str).str.startswith('3')]['Balance'].sum()
        expenses = self.df[self.df['GL_Code'].astype(str).str.startswith('4')]['Balance'].sum()
        
        # Calculate: Assets = -(Liabilities + Equity + Revenue + Expenses)
        rhs = -(liabilities + equity + revenue + expenses)
        difference = abs(assets - rhs)
        # Use percentage-based tolerance relative to asset base
        equation_tolerance = abs(assets) * self.tolerance_pct if abs(assets) > 0 else self.abs_tolerance
        equation_balanced = difference <= equation_tolerance
        
        # Check if assets are positive (should be > 0)
        assets_positive = assets > 0
        
        # Rule passes if equation is balanced AND assets are positive
        is_compliant = equation_balanced and assets_positive
        
        result = {
            "rule_number": 6,
            "rule_name": "Accounting Equation Validation",
            "status": "✓ PASS" if is_compliant else "✗ FAIL",
            "assets": assets,
            "liabilities": liabilities,
            "equity": equity,
            "revenue": revenue,
            "expenses": expenses,
            "rhs": rhs,
            "difference": difference,
            "equation_tolerance": equation_tolerance,
            "equation_balanced": equation_balanced,
            "assets_positive": assets_positive,
            "is_compliant": is_compliant,
            "details": f"Assets: {CURRENCY_SYMBOL}{assets:,.2f}, RHS: {CURRENCY_SYMBOL}{rhs:,.2f}, Diff: {CURRENCY_SYMBOL}{difference:,.2f}, Assets positive: {assets_positive}"
        }
        
        print(f"Assets (1xxxxx):     {CURRENCY_SYMBOL}{assets:>15,.2f}")
        print(f"Liabilities (2xxxxx): {CURRENCY_SYMBOL}{liabilities:>15,.2f}")
        print(f"Equity (5xxxxx):      {CURRENCY_SYMBOL}{equity:>15,.2f}")
        print(f"Revenue (3xxxxx):     {CURRENCY_SYMBOL}{revenue:>15,.2f}")
        print(f"Expenses (4xxxxx):    {CURRENCY_SYMBOL}{expenses:>15,.2f}")
        print(f"" + "=" * 60)
        print(f"RHS [-(L+E+R+X)]:    {CURRENCY_SYMBOL}{rhs:>15,.2f}")
        print(f"Difference:          {CURRENCY_SYMBOL}{difference:>15,.2f}")
        print(f"Equation tolerance:  {CURRENCY_SYMBOL}{equation_tolerance:>15,.2f} ({self.tolerance_pct*100:.6f}%)")
        print(f"Equation balanced:   {'✓ YES' if equation_balanced else '✗ NO'}")
        print(f"Assets positive:     {'✓ YES' if assets_positive else '✗ NO'}")
        print(f"Status: {result['status']}")
        
        self.validation_results["rule_6"] = result
        return result
    
    def validate_all_rules(self):
        """Run all enabled validation rules (dynamic based on entity config)"""
        print("\n" + "="*80)
        print("TRIAL BALANCE DYNAMIC RULE VALIDATION")
        print("="*80)
        print(f"Entity: {self.rules_config.get('entity_name', self.entity)}")
        print(f"File: {self.input_file}")
        print(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Tolerance pct: {self.tolerance_pct*100:.6f}% | Absolute tol: {CURRENCY_SYMBOL}{self.abs_tolerance}")

        # Show enabled rules
        enabled_rules = [
            rule_key for rule_key, rule_config in self.rules_config.get('validation_rules', {}).items()
            if rule_config.get('enabled', True)
        ]
        print(f"Enabled rules: {len(enabled_rules)}/{len(self.rules_config.get('validation_rules', {}))}")
        print("="*80)

        # Load data
        self.load_trial_balance()

        # Run enabled rules dynamically
        rule_methods = {
            'rule_1': self.rule_1_total_debits_equal_credits,
            'rule_2': self.rule_2_balance_calculation_accuracy,
            'rule_3': self.rule_3_no_duplicate_accounts,
            'rule_4': self.rule_4_no_missing_invalid_data,
            'rule_5': self.rule_5_logical_balance_signs,
            'rule_6': self.rule_6_accounting_equation
        }

        # Execute only enabled rules
        for rule_key in enabled_rules:
            if rule_key in rule_methods:
                rule_config = self.rules_config['validation_rules'][rule_key]
                print(f"\n🔍 Running {rule_key.upper()}: {rule_config.get('rule_name', 'Unknown')}")
                if rule_config.get('notes'):
                    print(f"   Note: {rule_config['notes']}")
                rule_methods[rule_key]()
            else:
                print(f"\n⚠️  Warning: Rule method not found for {rule_key}")

        # Show skipped rules
        skipped_rules = [
            f"Rule {rule_config['rule_number']}: {rule_config.get('rule_name', 'Unknown')}"
            for rule_key, rule_config in self.rules_config.get('validation_rules', {}).items()
            if not rule_config.get('enabled', True)
        ]
        if skipped_rules:
            print("\n" + "="*80)
            print("SKIPPED RULES (DISABLED FOR THIS ENTITY)")
            print("="*80)
            for skipped in skipped_rules:
                print(f"  ⊗ {skipped}")

        # Summary
        print("\n" + "="*80)
        print("VALIDATION SUMMARY")
        print("="*80)

        total_rules_run = len(self.validation_results)
        passed_rules = sum(1 for r in self.validation_results.values() if r["is_compliant"])

        for key in sorted(self.validation_results.keys()):
            result = self.validation_results[key]
            print(f"Rule {result['rule_number']}: {result['status']:10} - {result['rule_name']}")

        print("="*80)
        print(f"OVERALL COMPLIANCE: {passed_rules}/{total_rules_run} rules passed (out of {total_rules_run} enabled rules)")
        print(f"Total rules configured: {len(self.rules_config.get('validation_rules', {}))}")
        print("="*80)

        return self.validation_results
    
    def generate_excel_report(self):
        """Generate comprehensive Excel report with only enabled rules"""
        print(f"\n{'='*80}")
        print("GENERATING EXCEL REPORT")
        print("="*80)

        # Use the output file path set in __init__
        output_file = self.output_file

        # Delete existing report file if it exists (ensure overwrite mode)
        if os.path.exists(output_file):
            os.remove(output_file)

        # Create comprehensive report data
        print("Creating comprehensive validation report...")
        report_data = []

        # Header section
        report_data.append({
            "Section": "HEADER",
            "Rule": "",
            "Metric": "Validation Date",
            "Value": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "Status": ""
        })
        report_data.append({
            "Section": "HEADER",
            "Rule": "",
            "Metric": "Entity",
            "Value": self.rules_config.get('entity_name', self.entity.upper()),
            "Status": ""
        })
        report_data.append({
            "Section": "HEADER",
            "Rule": "",
            "Metric": "Input File",
            "Value": self.input_file,
            "Status": ""
        })
        report_data.append({
            "Section": "HEADER",
            "Rule": "",
            "Metric": "Total Records",
            "Value": len(self.df),
            "Status": ""
        })
        report_data.append({
            "Section": "HEADER",
            "Rule": "",
            "Metric": "Tolerance (%)",
            "Value": f"{self.tolerance_pct*100:.6f}%",
            "Status": ""
        })
        report_data.append({
            "Section": "HEADER",
            "Rule": "",
            "Metric": "Enabled Rules",
            "Value": f"{len(self.validation_results)}/{len(self.rules_config.get('validation_rules', {}))}",
            "Status": ""
        })

        # Empty row for spacing
        report_data.append({"Section": "", "Rule": "", "Metric": "", "Value": "", "Status": ""})

        # Generate report sections ONLY for enabled rules that were actually run
        # Rule 1: Total Debits Equal Credits
        if "rule_1" in self.validation_results:
            rule1 = self.validation_results["rule_1"]
            report_data.append({
                "Section": "RULE 1",
                "Rule": "Total Debits Equal Credits",
                "Metric": "Status",
                "Value": "",
                "Status": rule1["status"]
            })
            report_data.append({
                "Section": "RULE 1",
                "Rule": "",
                "Metric": "Total Debits",
                "Value": f"{CURRENCY_SYMBOL}{rule1['total_debits']:,.2f}",
                "Status": ""
            })
            report_data.append({
                "Section": "RULE 1",
                "Rule": "",
                "Metric": "Total Credits",
                "Value": f"{CURRENCY_SYMBOL}{rule1['total_credits']:,.2f}",
                "Status": ""
            })
            report_data.append({
                "Section": "RULE 1",
                "Rule": "",
                "Metric": "Difference",
                "Value": f"{CURRENCY_SYMBOL}{rule1['difference']:,.2f}",
                "Status": ""
            })
            report_data.append({
                "Section": "RULE 1",
                "Rule": "",
                "Metric": "Allowed Variance",
                "Value": f"{CURRENCY_SYMBOL}{rule1['allowed_variance']:,.2f}",
                "Status": ""
            })
            report_data.append({"Section": "", "Rule": "", "Metric": "", "Value": "", "Status": ""})

        # Rule 2: Balance Calculation Accuracy
        if "rule_2" in self.validation_results:
            rule2 = self.validation_results["rule_2"]
            report_data.append({
                "Section": "RULE 2",
                "Rule": "Balance Calculation Accuracy",
                "Metric": "Status",
                "Value": "",
                "Status": rule2["status"]
            })
            report_data.append({
                "Section": "RULE 2",
                "Rule": "",
                "Metric": "Total Rows Checked",
                "Value": rule2["total_records"],
                "Status": ""
            })
            report_data.append({
                "Section": "RULE 2",
                "Rule": "",
                "Metric": "Errors Found",
                "Value": rule2["violations_count"],
                "Status": ""
            })
            report_data.append({"Section": "", "Rule": "", "Metric": "", "Value": "", "Status": ""})

        # Rule 3: No Duplicate Accounts
        if "rule_3" in self.validation_results:
            rule3 = self.validation_results["rule_3"]
            report_data.append({
                "Section": "RULE 3",
                "Rule": "No Duplicate Accounts",
                "Metric": "Status",
                "Value": "",
                "Status": rule3["status"]
            })
            report_data.append({
                "Section": "RULE 3",
                "Rule": "",
                "Metric": "Total Records",
                "Value": rule3["total_records"],
                "Status": ""
            })
            report_data.append({
                "Section": "RULE 3",
                "Rule": "",
                "Metric": "Duplicate GL Codes",
                "Value": rule3["duplicate_gl_codes"],
                "Status": ""
            })
            report_data.append({
                "Section": "RULE 3",
                "Rule": "",
                "Metric": "Total Duplicate Rows",
                "Value": rule3["total_duplicate_records"],
                "Status": ""
            })
            if "rule_3" in self.violations and len(self.violations["rule_3"]) > 0:
                report_data.append({
                    "Section": "RULE 3",
                    "Rule": "",
                    "Metric": "Violation Details",
                    "Value": "See 'Rule 3 Violations' sheet",
                    "Status": ""
                })
            report_data.append({"Section": "", "Rule": "", "Metric": "", "Value": "", "Status": ""})

        # Rule 4: No Missing or Invalid Data
        if "rule_4" in self.validation_results:
            rule4 = self.validation_results["rule_4"]
            report_data.append({
                "Section": "RULE 4",
                "Rule": "No Missing or Invalid Data",
                "Metric": "Status",
                "Value": "",
                "Status": rule4["status"]
            })
            report_data.append({
                "Section": "RULE 4",
                "Rule": "",
                "Metric": "Total Records",
                "Value": rule4["total_records"],
                "Status": ""
            })
            report_data.append({
                "Section": "RULE 4",
                "Rule": "",
                "Metric": "Issues Found",
                "Value": rule4["issues_count"],
                "Status": ""
            })
            if "rule_4" in self.violations and len(self.violations["rule_4"]) > 0:
                report_data.append({
                    "Section": "RULE 4",
                    "Rule": "",
                    "Metric": "Violation Details",
                    "Value": "See 'Rule 4 Violations' sheet",
                    "Status": ""
                })
            report_data.append({"Section": "", "Rule": "", "Metric": "", "Value": "", "Status": ""})

        # Rule 5: Logical Balance Signs by Account Type
        if "rule_5" in self.validation_results:
            rule5 = self.validation_results["rule_5"]
            report_data.append({
                "Section": "RULE 5",
                "Rule": "Logical Balance Signs by Account Type",
                "Metric": "Status",
                "Value": "",
                "Status": rule5["status"]
            })
            report_data.append({
                "Section": "RULE 5",
                "Rule": "",
                "Metric": "Total Rows Checked",
                "Value": rule5["total_records"],
                "Status": ""
            })
            report_data.append({
                "Section": "RULE 5",
                "Rule": "",
                "Metric": "Violations Found",
                "Value": rule5["violations_count"],
                "Status": ""
            })
            if "rule_5" in self.violations and len(self.violations["rule_5"]) > 0:
                report_data.append({
                    "Section": "RULE 5",
                    "Rule": "",
                    "Metric": "Violation Details",
                    "Value": "See 'Rule 5 Violations' sheet",
                    "Status": ""
                })
            report_data.append({"Section": "", "Rule": "", "Metric": "", "Value": "", "Status": ""})

        # Rule 6: Accounting Equation Validation
        if "rule_6" in self.validation_results:
            rule6 = self.validation_results["rule_6"]
            report_data.append({
                "Section": "RULE 6",
                "Rule": "Accounting Equation Validation",
                "Metric": "Status",
                "Value": "",
                "Status": rule6["status"]
            })
            report_data.append({
                "Section": "RULE 6",
                "Rule": "",
                "Metric": "Assets (1xxxxx)",
                "Value": f"{CURRENCY_SYMBOL}{rule6['assets']:,.2f}",
                "Status": ""
            })
            report_data.append({
                "Section": "RULE 6",
                "Rule": "",
                "Metric": "Liabilities (2xxxxx)",
                "Value": f"{CURRENCY_SYMBOL}{rule6['liabilities']:,.2f}",
                "Status": ""
            })
            report_data.append({
                "Section": "RULE 6",
                "Rule": "",
                "Metric": "Equity (5xxxxx)",
                "Value": f"{CURRENCY_SYMBOL}{rule6['equity']:,.2f}",
                "Status": ""
            })
            report_data.append({
                "Section": "RULE 6",
                "Rule": "",
                "Metric": "Revenue (3xxxxx)",
                "Value": f"{CURRENCY_SYMBOL}{rule6['revenue']:,.2f}",
                "Status": ""
            })
            report_data.append({
                "Section": "RULE 6",
                "Rule": "",
                "Metric": "Expenses (4xxxxx)",
                "Value": f"{CURRENCY_SYMBOL}{rule6['expenses']:,.2f}",
                "Status": ""
            })
            report_data.append({
                "Section": "RULE 6",
                "Rule": "",
                "Metric": "RHS [-(L+E+R+X)]",
                "Value": f"{CURRENCY_SYMBOL}{rule6['rhs']:,.2f}",
                "Status": ""
            })
            report_data.append({
                "Section": "RULE 6",
                "Rule": "",
                "Metric": "Difference (A - RHS)",
                "Value": f"{CURRENCY_SYMBOL}{rule6['difference']:,.2f}",
                "Status": ""
            })
            report_data.append({
                "Section": "RULE 6",
                "Rule": "",
                "Metric": "Equation Balanced",
                "Value": "✓ YES" if rule6["equation_balanced"] else "✗ NO",
                "Status": ""
            })
            report_data.append({
                "Section": "RULE 6",
                "Rule": "",
                "Metric": "Assets Positive",
                "Value": "✓ YES" if rule6["assets_positive"] else "✗ NO",
                "Status": ""
            })
            report_data.append({"Section": "", "Rule": "", "Metric": "", "Value": "", "Status": ""})
        
        # Summary section
        total_rules = len(self.validation_results)
        passed_rules = sum(1 for r in self.validation_results.values() if r["is_compliant"])
        failed_rules = total_rules - passed_rules
        
        report_data.append({
            "Section": "SUMMARY",
            "Rule": "Overall Validation Results",
            "Metric": "Total Rules Checked",
            "Value": total_rules,
            "Status": ""
        })
        report_data.append({
            "Section": "SUMMARY",
            "Rule": "",
            "Metric": "Rules Passed",
            "Value": passed_rules,
            "Status": "✓ PASS" if passed_rules > 0 else ""
        })
        report_data.append({
            "Section": "SUMMARY",
            "Rule": "",
            "Metric": "Rules Failed",
            "Value": failed_rules,
            "Status": "✗ FAIL" if failed_rules > 0 else ""
        })
        report_data.append({
            "Section": "SUMMARY",
            "Rule": "",
            "Metric": "Compliance Status",
            "Value": f"{passed_rules}/{total_rules}",
            "Status": "✓ COMPLIANT" if failed_rules == 0 else "✗ NON-COMPLIANT"
        })
        
        # Create DataFrame
        report_df = pd.DataFrame(report_data)
        
        # Create Excel writer and write main report
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Main comprehensive report
            report_df.to_excel(writer, sheet_name="Validation Report", index=False)
            
            # Add violation details sheets if violations exist
            if "rule_3" in self.violations and len(self.violations["rule_3"]) > 0:
                print("Adding Rule 3 violations sheet...")
                self.violations["rule_3"].to_excel(writer, sheet_name="Rule 3 Violations", index=False)
            
            if "rule_4" in self.violations and len(self.violations["rule_4"]) > 0:
                print("Adding Rule 4 violations sheet...")
                self.violations["rule_4"].to_excel(writer, sheet_name="Rule 4 Violations", index=False)
            
            if "rule_5" in self.violations and len(self.violations["rule_5"]) > 0:
                print("Adding Rule 5 violations sheet...")
                self.violations["rule_5"].to_excel(writer, sheet_name="Rule 5 Violations", index=False)
            
            if "rule_6" in self.violations and len(self.violations["rule_6"]) > 0:
                print("Adding Rule 6 violations sheet...")
                self.violations["rule_6"].to_excel(writer, sheet_name="Rule 6 Violations", index=False)
        
        # Format the workbook
        print("Formatting workbook...")
        wb = load_workbook(self.output_file)
        
        # Format main report sheet
        if "Validation Report" in wb.sheetnames:
            ws = wb["Validation Report"]
            self._format_validation_report_sheet(ws)
        
        # Format other sheets
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            self._auto_adjust_columns(ws)
        
        wb.save(self.output_file)
        print(f"✓ Report saved: {self.output_file}")
    
    def _format_validation_report_sheet(self, worksheet):
        """Apply formatting to the comprehensive validation report sheet"""
        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Section headers formatting (HEADER, RULE 1-6, SUMMARY)
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        section_font = Font(bold=True, size=11)
        
        # Status formatting
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        pass_font = Font(color="006100", bold=True)
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        fail_font = Font(color="9C0006", bold=True)
        
        for row in range(2, worksheet.max_row + 1):
            section_cell = worksheet[f"A{row}"]
            status_cell = worksheet[f"E{row}"]
            
            # Format section headers
            if section_cell.value and section_cell.value in ["HEADER", "RULE 1", "RULE 2", "RULE 3", "RULE 4", "RULE 5", "RULE 6", "SUMMARY"]:
                for col in ["A", "B", "C", "D", "E"]:
                    cell = worksheet[f"{col}{row}"]
                    cell.fill = section_fill
                    cell.font = section_font
            
            # Format status cells
            status_value = str(status_cell.value)
            if "PASS" in status_value or "COMPLIANT" in status_value:
                status_cell.fill = pass_fill
                status_cell.font = pass_font
            elif "FAIL" in status_value or "NON-COMPLIANT" in status_value:
                status_cell.fill = fail_fill
                status_cell.font = fail_font
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                       min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border
                # Align text
                if cell.row > 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    def _auto_adjust_columns(self, worksheet):
        """Auto-adjust column widths"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def run(self):
        """Main execution flow"""
        try:
            # Load trial balance
            self.load_trial_balance()
            
            # Validate all rules
            self.validate_all_rules()
            
            # Generate Excel report
            self.generate_excel_report()
            
            print("\n" + "="*80)
            print("VALIDATION COMPLETE")
            print("="*80)
            print(f"\nReport generated: {self.output_file}")
            
            return True
            
        except Exception as e:
            print(f"\n✗ Error during validation: {e}")
            import traceback
            traceback.print_exc()
            return False


def validate_final_trial_balance(entity: str = "cpm", input_file=None, output_file=None, rule_overrides=None):
    """
    Convenience function to validate final trial balance from Step 4.
    This is the main entry point for Step 5 integration.

    Args:
        entity: Entity code (e.g., 'cpm', 'hausen')
        input_file: Path to final trial balance (defaults to entity final TB)
        output_file: Path to validation report (defaults to entity validation report)
        rule_overrides: Optional dict with rule override settings from frontend toggles
                       e.g., {"rule_1": True, "rule_2": False, ...}

    Returns:
        tuple: (success: bool, message: str, output_file: str, summary: dict)
    """
    try:
        print("\n" + "="*80)
        print("STEP 5: TRIAL BALANCE 6-RULE VALIDATION")
        print(f"Entity: {entity.upper()}")
        if rule_overrides:
            print(f"Using custom rule overrides from UI")
        print("="*80)

        # Create validator and run (defaults handled in __init__)
        validator = TrialBalanceValidator(entity=entity, input_file=input_file, output_file=output_file)

        # Apply rule overrides if provided
        if rule_overrides:
            print(f"\n📋 Applying rule overrides from frontend toggles:")
            for rule_key, enabled in rule_overrides.items():
                if rule_key in validator.rules_config.get('validation_rules', {}):
                    old_value = validator.rules_config['validation_rules'][rule_key].get('enabled', True)
                    validator.rules_config['validation_rules'][rule_key]['enabled'] = enabled
                    if old_value != enabled:
                        status = "ENABLED" if enabled else "DISABLED"
                        print(f"   • {rule_key}: {status} (was: {'enabled' if old_value else 'disabled'})")
        
        # Check if input file exists
        if not os.path.exists(validator.input_file):
            error_msg = f"❌ Input file not found: {validator.input_file}"
            print(error_msg)
            return False, error_msg, None, None
        
        success = validator.run()
        
        if success:
            # Prepare summary for display
            failed_rules_list = [
                {
                    "rule_key": rule_key,
                    "rule_number": result.get("rule_number"),
                    "rule_name": result.get("rule_name"),
                    "status": result.get("status"),
                    "details": result.get("details", ""),
                    "has_violations": rule_key in validator.violations and len(validator.violations[rule_key]) > 0
                }
                for rule_key, result in validator.validation_results.items()
                if not result.get("is_compliant", True)
            ]
            
            summary = {
                "total_records": len(validator.df),
                "total_rules_checked": len(validator.validation_results),
                "rules_passed": sum(1 for r in validator.validation_results.values() if r.get("is_compliant", False)),
                "rules_failed": sum(1 for r in validator.validation_results.values() if not r.get("is_compliant", True)),
                "validation_results": validator.validation_results,
                "failed_rules": failed_rules_list,
                "violations_count": {
                    rule_key: len(violations_df) 
                    for rule_key, violations_df in validator.violations.items()
                }
            }
            
            success_msg = f"""
✅ Validation completed successfully!

📊 Summary:
   • Total GL Codes: {summary['total_records']}
   • Rules Checked: {summary['total_rules_checked']}
   • Rules Passed: {summary['rules_passed']}
   • Rules Failed: {summary['rules_failed']}
   
📁 Validation report saved to: {validator.output_file}

🎯 Next Step: Review validation report and proceed to financial statement generation
"""
            print(success_msg)
            return True, success_msg, validator.output_file, summary
        else:
            error_msg = "❌ Validation failed. Please check the error messages above."
            return False, error_msg, None, None
            
    except Exception as e:
        error_msg = f"❌ Error during validation: {str(e)}"
        print(error_msg)
        import traceback
        traceback.print_exc()
        return False, error_msg, None, None


# Legacy function removed - no longer needed with entity-based structure
# Use validate_final_trial_balance(entity="cpm") instead




if __name__ == "__main__":
    # Validate final trial balance for an entity
    import sys
    
    # Get entity from command line or use default
    entity = sys.argv[1] if len(sys.argv) > 1 else "cpm"
    
    print("\nTrial Balance 8-Rule Validation (Simplified)")
    print("="*50)
    print(f"Entity: {entity.upper()}")
    print("="*50)
    
    # Validate final trial balance from Step 4
    success, message, output_file, summary = validate_final_trial_balance(entity=entity)
    if not success:
        sys.exit(1)
