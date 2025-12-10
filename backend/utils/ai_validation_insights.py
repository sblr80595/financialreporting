"""
AI-Powered Validation Insights Generator
Generates comprehensive insights and recommendations for failed validation rules
using LLM analysis to help users understand failures and how to fix them.

This program:
1. Analyzes validation results from the 7-rule validation
2. Uses LLM to generate contextual insights for each failure
3. Provides actionable recommendations
4. Creates a detailed PDF/Excel report in the output folder
5. Integrates with Step 5 UI for download and viewing
"""

import os
import sys
import json
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Add project root to Python path for imports
project_root = Path(__file__).parent.parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

from backend.services.path_service import PathService


class AIValidationInsightsGenerator:
    """
    Generates AI-powered insights for validation failures
    """
    
    def __init__(self, entity: str = "cpm"):
        """
        Initialize the insights generator
        
        Args:
            entity: Entity code (e.g., 'cpm', 'hausen')
        """
        self.entity = entity
        self.path_service = PathService(entity)
        self.validation_report_path = None
        self.insights_report_path = None
        self.validation_data = None
        self.insights = {}
        
    def load_validation_report(self) -> bool:
        """
        Load the validation report from Step 5
        
        Returns:
            bool: True if report loaded successfully, False otherwise
        """
        try:
            # Get validation report path
            self.validation_report_path = self.path_service.get_validation_report_path(self.entity)
            
            if not os.path.exists(self.validation_report_path):
                error_msg = f"Validation report not found for entity '{self.entity.upper()}'. Please run Step 5 (6-Rules Validation) first."
                print(f"❌ {error_msg}")
                print(f"   Expected path: {self.validation_report_path}")
                raise FileNotFoundError(error_msg)
            
            # Load the Excel file (it has multiple sheets)
            excel_file = pd.ExcelFile(self.validation_report_path)
            
            self.validation_data = {
                'sheets': excel_file.sheet_names,
                'data': {}
            }
            
            # Load all sheets
            for sheet_name in excel_file.sheet_names:
                self.validation_data['data'][sheet_name] = pd.read_excel(
                    excel_file, 
                    sheet_name=sheet_name
                )
            
            print(f"✓ Loaded validation report with {len(self.validation_data['sheets'])} sheets")
            return True
            
        except FileNotFoundError:
            raise
        except Exception as e:
            print(f"❌ Error loading validation report: {e}")
            raise
    
    def analyze_rule_failures(self) -> Dict[str, Any]:
        """
        Analyze each validation rule and identify failures
        
        Returns:
            Dict containing analysis of each rule
        """
        analysis = {}
        
        # Check if we have the summary sheet
        if 'Summary' in self.validation_data['data']:
            summary_df = self.validation_data['data']['Summary']
            
            for idx, row in summary_df.iterrows():
                rule_name = row.get('Rule Name', 'Unknown')
                status = row.get('Status', 'Unknown')
                
                # Only analyze failed rules
                if '✗' in str(status) or 'FAIL' in str(status).upper():
                    analysis[rule_name] = {
                        'status': 'FAILED',
                        'details': row.to_dict(),
                        'violations': [],
                        'violation_count': 0
                    }
        
        # Load violation details from rule detail sheets
        # These sheets are named like "Rule 5 - Invalid Data", "Rule 6 - Balance Signs", etc.
        for sheet_name in self.validation_data['sheets']:
            if sheet_name.startswith('Rule') and sheet_name != 'Summary':
                violations_df = self.validation_data['data'][sheet_name]
                
                # Only include if there are actual violations (non-empty sheet)
                if len(violations_df) > 0:
                    # Match sheet name to rule name from summary
                    # Extract the descriptive part after the hyphen
                    matched_rule = None
                    for rule_name in analysis.keys():
                        # Check if the sheet name contains key parts of the rule name
                        if any(keyword in sheet_name for keyword in ['Invalid Data', 'Balance Signs', 'Empty Capital', 'Accounting Eq']):
                            if 'Invalid Data' in sheet_name and 'Invalid Data' in rule_name:
                                matched_rule = rule_name
                            elif 'Balance Signs' in sheet_name and 'Balance Signs' in rule_name:
                                matched_rule = rule_name
                            elif 'Empty Capital' in sheet_name and 'Capital' in rule_name:
                                matched_rule = rule_name
                            elif 'Accounting Eq' in sheet_name and 'Accounting' in rule_name:
                                matched_rule = rule_name
                    
                    # If we found a match, add the violations
                    if matched_rule:
                        analysis[matched_rule]['violations'] = violations_df.to_dict('records')
                        analysis[matched_rule]['violation_count'] = len(violations_df)
                    else:
                        # If no match found but sheet has data, create entry with sheet name
                        analysis[sheet_name] = {
                            'status': 'FAILED',
                            'details': {'Details': f'See {sheet_name} sheet for violations'},
                            'violations': violations_df.to_dict('records'),
                            'violation_count': len(violations_df)
                        }
        
        return analysis
    
    def generate_ai_insights(self, rule_name: str, failure_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Generate AI-powered insights for a specific rule failure
        
        Args:
            rule_name: Name of the failed validation rule
            failure_data: Data about the failure
            
        Returns:
            Dict containing AI-generated insights
        """
        # This is a template-based approach. In production, you would call
        # an LLM API (OpenAI, Anthropic, etc.) to generate dynamic insights
        
        insights = {
            'rule_name': rule_name,
            'severity': self._determine_severity(rule_name, failure_data),
            'summary': '',
            'root_causes': [],
            'recommendations': [],
            'examples': [],
            'impact': ''
        }
        
        # Generate insights based on rule type
        if 'Debits Equal Credits' in rule_name or 'Rule 1' in rule_name:
            insights.update(self._generate_debit_credit_insights(failure_data))
            
        elif 'Sum of Balances' in rule_name or 'Rule 2' in rule_name:
            insights.update(self._generate_balance_sum_insights(failure_data))
            
        elif 'Balance Calculation' in rule_name or 'Rule 3' in rule_name:
            insights.update(self._generate_calculation_insights(failure_data))
            
        elif 'Duplicate' in rule_name or 'Rule 4' in rule_name:
            insights.update(self._generate_duplicate_insights(failure_data))
            
        elif 'Missing or Invalid' in rule_name or 'Rule 5' in rule_name:
            insights.update(self._generate_data_quality_insights(failure_data))
            
        elif 'Balance Signs' in rule_name or 'Rule 6' in rule_name:
            insights.update(self._generate_sign_logic_insights(failure_data))
            
        elif 'Accounting Equation' in rule_name or 'Capital' in rule_name or 'Rule 7' in rule_name:
            insights.update(self._generate_equation_insights(failure_data))
        
        return insights
    
    def _determine_severity(self, rule_name: str, failure_data: Dict[str, Any]) -> str:
        """Determine severity level of the failure"""
        violation_count = failure_data.get('violation_count', 0)
        
        # Rules 1 and 2 are critical (fundamental accounting principles)
        if any(x in rule_name for x in ['Rule 1', 'Rule 2', 'Debits Equal Credits', 'Sum of Balances']):
            return 'CRITICAL'
        
        # High violation counts
        if violation_count > 50:
            return 'HIGH'
        elif violation_count > 10:
            return 'MEDIUM'
        else:
            return 'LOW'
    
    def _generate_debit_credit_insights(self, failure_data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate insights for Rule 1: Debits Equal Credits"""
        details = failure_data.get('details', {})
        total_debits = details.get('Total Debits', 0)
        total_credits = details.get('Total Credits', 0)
        difference = abs(total_debits - total_credits)
        
        return {
            'summary': f'The total debits (₹{total_debits:,.2f}) do not equal total credits (₹{total_credits:,.2f}). Difference: ₹{difference:,.2f}',
            'root_causes': [
                '🔍 **Data Entry Errors**: Manual entry mistakes in debit/credit amounts',
                '🔍 **Import Issues**: Incorrect mapping during data import from SAP/source system',
                '🔍 **Adjustment Errors**: Manual adjustments not properly balanced',
                '🔍 **System Rounding**: Accumulated rounding errors from multiple calculations',
                '🔍 **Missing Entries**: Some journal entries may be incomplete'
            ],
            'recommendations': [
                '✅ **Immediate Action**: Review recent manual adjustments for balancing errors',
                '✅ **Verify Import**: Re-check the original source data export and import mapping',
                '✅ **Reconcile**: Identify which GL codes contribute most to the imbalance',
                '✅ **Adjust**: Create a correcting journal entry to balance debits and credits',
                '✅ **Review Process**: Implement pre-validation checks during data import'
            ],
            'examples': [
                f'Example: If debits are ₹{difference:,.2f} higher, check for:\n  - Missing credit entries\n  - Doubled debit entries\n  - Sign reversal in credit columns',
                'Common pattern: Contra accounts (accumulated depreciation, allowances) often have reversed signs',
                'SAP export issue: Check if "Balance" column was used instead of separate Debit/Credit columns'
            ],
            'impact': '🚨 **CRITICAL IMPACT**: This is a fundamental accounting principle. Financial statements cannot be generated until this is resolved. All downstream reports will be incorrect.'
        }
    
    def _generate_balance_sum_insights(self, failure_data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate insights for Rule 2: Sum of Balances"""
        details = failure_data.get('details', {})
        total_balance = details.get('Total Balance', 0)
        
        return {
            'summary': f'All account balances should sum to zero, but current sum is ₹{total_balance:,.2f}',
            'root_causes': [
                '🔍 **Account Classification**: Some accounts may be classified incorrectly (Asset vs. Liability)',
                '🔍 **Sign Convention**: Inconsistent use of positive/negative signs across account types',
                '🔍 **Period Mismatch**: Opening balances from different periods mixed with current period',
                '🔍 **Incomplete Adjustments**: Year-end adjustments not fully applied',
                '🔍 **Intercompany Elimination**: Intercompany transactions not properly eliminated'
            ],
            'recommendations': [
                '✅ **Analyze by Type**: Break down the imbalance by account type (Asset, Liability, Equity, Revenue, Expense)',
                '✅ **Check Signs**: Verify that liabilities, equity, and revenue have negative balances',
                '✅ **Review Period**: Ensure all balances are from the same accounting period (Mar\'25)',
                '✅ **Audit Trail**: Trace back through adjustments to find where imbalance was introduced',
                '✅ **Cross-Check**: Compare with previous month\'s trial balance for consistency'
            ],
            'examples': [
                f'If total is ₹{abs(total_balance):,.2f}, investigate:\n  - Are all equity accounts negative?\n  - Are all expense accounts positive?\n  - Check for accounts with wrong GL code ranges',
                'Pattern: Check if retained earnings or profit/loss accounts are missing',
                'Common fix: Ensure P&L accounts close to retained earnings properly'
            ],
            'impact': '🚨 **CRITICAL IMPACT**: Indicates fundamental data integrity issues. The accounting equation Assets = Liabilities + Equity may not be satisfied.'
        }
    
    def _generate_calculation_insights(self, failure_data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate insights for Rule 3: Balance Calculation Accuracy"""
        violation_count = failure_data.get('violation_count', 0)
        violations = failure_data.get('violations', [])
        
        sample_violations = violations[:3] if violations else []
        
        return {
            'summary': f'Found {violation_count} accounts where Balance ≠ Opening Balance + Debits - Credits',
            'root_causes': [
                '🔍 **Formula Errors**: Incorrect calculation formulas in source system or Excel',
                '🔍 **Manual Overrides**: Balance column manually edited without updating components',
                '🔍 **Data Corruption**: Cell references broken during copy/paste operations',
                '🔍 **Import Mapping**: Wrong columns mapped during import process',
                '🔍 **Currency Conversion**: Exchange rate calculations not properly applied'
            ],
            'recommendations': [
                f'✅ **Fix Immediately**: Recalculate balances for {violation_count} affected accounts',
                '✅ **Use Formula**: Ensure Balance = Opening Balance + Debit - Credit for all accounts',
                '✅ **Lock Columns**: Protect calculated columns from manual editing',
                '✅ **Validation Check**: Add Excel formulas to highlight calculation mismatches',
                '✅ **Reimport**: If widespread, consider re-importing from source system'
            ],
            'examples': [
                'Sample violations:' if sample_violations else 'No violation samples available',
            ] + [
                f"  GL {v.get('GL_Code', 'N/A')}: Expected ₹{v.get('Expected_Balance', 0):,.2f}, Got ₹{v.get('Balance', 0):,.2f}"
                for v in sample_violations
            ],
            'impact': '⚠️ **HIGH IMPACT**: These calculation errors will propagate to financial statements and cause incorrect account balances in P&L and Balance Sheet.'
        }
    
    def _generate_duplicate_insights(self, failure_data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate insights for Rule 4: Duplicate Accounts"""
        violation_count = failure_data.get('violation_count', 0)
        violations = failure_data.get('violations', [])
        
        duplicate_codes = list(set([v.get('GL_Code', '') for v in violations]))[:5]
        
        return {
            'summary': f'Found {violation_count} duplicate GL code entries across the trial balance',
            'root_causes': [
                '🔍 **Multiple Imports**: Trial balance imported multiple times without clearing',
                '🔍 **Consolidation Issues**: Multiple entities/departments not properly merged',
                '🔍 **Manual Additions**: Same account added manually after automatic import',
                '🔍 **Template Errors**: Excel template had duplicate rows that weren\'t removed',
                '🔍 **Historical Data**: Old period data mixed with current period'
            ],
            'recommendations': [
                '✅ **Consolidate**: Combine balances for duplicate GL codes into single entries',
                '✅ **Remove Duplicates**: Use Excel "Remove Duplicates" feature carefully',
                '✅ **Verify Amounts**: Before merging, verify that amounts should be summed',
                '✅ **Source Review**: Check source system to confirm correct GL code structure',
                '✅ **Prevention**: Implement unique GL code constraint in import process'
            ],
            'examples': [
                f'Duplicate GL Codes found: {", ".join(duplicate_codes)}' if duplicate_codes else 'No specific duplicates listed',
                'Action required: For each duplicate, decide whether to:\n  - Sum the balances (if legitimate sub-accounts)\n  - Keep one, delete other (if true duplicate)\n  - Investigate further (if amounts conflict)',
                'Common scenario: Same GL code appears with different descriptions - review SAP master data'
            ],
            'impact': '⚠️ **MEDIUM IMPACT**: Duplicates inflate account balances and can double-count amounts. Must be resolved before generating statements.'
        }
    
    def _generate_data_quality_insights(self, failure_data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate insights for Rule 5: Data Quality"""
        details = failure_data.get('details', {})
        violations = failure_data.get('violations', [])
        violation_count = failure_data.get('violation_count', len(violations))
        
        # Analyze types of data quality issues
        null_gl = 0
        null_desc = 0
        null_balance = 0
        sample_issues = []
        
        for v in violations:
            gl_code = v.get('GL_Code')
            desc = v.get('Description')
            balance = v.get('Balance')
            
            # Check for missing or NaN values
            if pd.isna(gl_code) or not gl_code:
                null_gl += 1
                sample_issues.append(f"Missing GL Code - Balance: ₹{balance:,.2f}" if not pd.isna(balance) else "Missing GL Code")
            if pd.isna(desc) or not desc:
                null_desc += 1
                if gl_code and not pd.isna(gl_code):
                    sample_issues.append(f"GL {gl_code}: Missing description")
            if pd.isna(balance):
                null_balance += 1
        
        # Get first 5 sample issues
        sample_issues = sample_issues[:5]
        
        # Parse details from summary if available
        details_text = details.get('Details', '')
        if 'missing GL Code' in details_text or 'missing Description' in details_text:
            # Extract numbers from details text
            import re
            numbers = re.findall(r'(\d+)\s+records?\s+with\s+missing\s+(\w+)', details_text)
            for count, field in numbers:
                if field == 'GL':
                    null_gl = int(count)
                elif field == 'Description':
                    null_desc = int(count)
        
        return {
            'summary': f'Data quality issues detected: {null_gl} missing GL codes, {null_desc} missing descriptions, {null_balance} missing balances. Total violations: {violation_count}',
            'root_causes': [
                '🔍 **Incomplete Export**: SAP export stopped mid-process or had errors',
                '🔍 **Excel Corruption**: File damaged during transfer or editing',
                '🔍 **Manual Deletion**: Critical data accidentally deleted during cleanup',
                '🔍 **Import Errors**: Column mapping issues left some fields empty',
                '🔍 **Source Data**: Master data incomplete in source system (SAP/ERP)'
            ],
            'recommendations': [
                f'✅ **Fill Missing Data**: Review {violation_count} rows and populate missing GL codes/descriptions',
                '✅ **Cross-Reference**: Use SAP master data (chart of accounts) to fill in missing descriptions',
                '✅ **Remove Empty Rows**: Delete rows that are completely empty or have zero balances',
                '✅ **Validate Source**: Check if source system has these accounts properly defined',
                '✅ **Import Cleanup**: Add validation rules to import process to reject incomplete data',
                f'✅ **Priority Action**: Focus on {null_gl} missing GL codes first - these are critical'
            ],
            'examples': [
                f'❌ Critical issue: {null_gl} rows without GL codes cannot be classified or reported',
                f'⚠️ Warning: {null_desc} rows without descriptions make reports hard to read',
                f'❌ Data issue: {null_balance} rows with missing balance values',
                '\nSample violations:',
            ] + sample_issues + [
                '\n💡 Quick Fix: Use Excel\'s VLOOKUP with GL master data to populate descriptions:',
                '   =VLOOKUP(A2, MasterData!A:B, 2, FALSE)',
                '⚠️ Important: Never leave balances null - should be 0.00 if no activity'
            ],
            'impact': '🚨 **HIGH IMPACT**: Missing GL codes prevent proper classification into Balance Sheet and P&L. Missing descriptions make financial reports unreadable. Must be addressed before proceeding.',
            'violation_count': violation_count
        }
    
    def _generate_sign_logic_insights(self, failure_data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate insights for Rule 6: Balance Sign Logic"""
        violation_count = failure_data.get('violation_count', 0)
        violations = failure_data.get('violations', [])
        
        # Categorize violations by account type
        asset_violations = [v for v in violations if v.get('Account_Type') == 'Asset']
        liability_violations = [v for v in violations if v.get('Account_Type') == 'Liability']
        equity_violations = [v for v in violations if v.get('Account_Type') == 'Equity']
        revenue_violations = [v for v in violations if v.get('Account_Type') == 'Revenue']
        expense_violations = [v for v in violations if v.get('Account_Type') == 'Expense']
        
        # Get sample violations with details
        sample_violations = []
        for v in violations[:10]:  # First 10 violations
            gl_code = v.get('GL_Code', 'N/A')
            desc = v.get('Description', 'No description')
            acct_type = v.get('Account_Type', 'Unknown')
            balance = v.get('Balance', 0)
            is_contra = v.get('Is_Contra_Asset', False) or v.get('Is_Contra_Revenue', False) or v.get('Is_Contra_Equity', False)
            
            # Determine expected vs actual sign
            if acct_type in ['Asset', 'Expense']:
                expected = 'Positive'
                actual = 'Negative' if balance < 0 else 'Positive'
            else:  # Liability, Equity, Revenue
                expected = 'Negative'
                actual = 'Positive' if balance > 0 else 'Negative'
            
            if is_contra:
                expected = 'Negative (Contra)' if acct_type == 'Asset' else 'Positive (Contra)'
            
            sample_violations.append(
                f"  • {gl_code} ({acct_type}): ₹{balance:,.2f} [{actual}, expected {expected}]\n    {desc[:60]}"
            )
        
        return {
            'summary': f'Found {violation_count} accounts with incorrect debit/credit signs for their account type',
            'root_causes': [
                '🔍 **Sign Convention**: Misunderstanding of debit/credit sign conventions',
                '🔍 **Contra Accounts**: Contra accounts not properly identified (e.g., Accumulated Depreciation)',
                '🔍 **Period End**: Negative balances in revenue/expense accounts after period close',
                '🔍 **Corrections**: Adjusting entries reversed the normal account balance',
                '🔍 **GL Code Range**: Account classified in wrong GL code range (e.g., 1xxx vs 2xxx)'
            ],
            'recommendations': [
                '✅ **Review Convention**: Assets & Expenses = Positive (Debit), Liabilities & Revenue = Negative (Credit)',
                '✅ **Identify Contra Accounts**: Mark contra-asset, contra-revenue accounts appropriately',
                '✅ **Investigate Reversals**: Check if negative balances are legitimate (e.g., prepayments)',
                '✅ **Reclassify**: Move accounts to correct GL code range if misclassified',
                f'✅ **Fix Priority**: Focus on {len(asset_violations)} asset violations first',
                f'✅ **Review {len(liability_violations)} liability and {len(equity_violations)} equity violations'
            ],
            'examples': [
                f'📊 Breakdown by Account Type:',
                f'  • Assets with wrong sign: {len(asset_violations)} (should be positive unless contra-asset)',
                f'  • Liabilities with wrong sign: {len(liability_violations)} (should be negative)',
                f'  • Equity with wrong sign: {len(equity_violations)} (should be negative)',
                f'  • Revenue with wrong sign: {len(revenue_violations)} (should be negative)',
                f'  • Expenses with wrong sign: {len(expense_violations)} (should be positive)',
                '',
                '🔍 Sample Violations (first 10):',
            ] + sample_violations + [
                '',
                '💡 Common Patterns to Check:',
                '  - Accumulated Depreciation (12xxxxx): Should be NEGATIVE (contra-asset)',
                '  - Accounts Payable (21xxxxx): Should be NEGATIVE (liability)',
                '  - Sales Returns/Allowances: Should be POSITIVE (contra-revenue)',
                '  - Purchase Discounts: Should be NEGATIVE (contra-expense)',
                '',
                '⚠️ Keywords indicating contra accounts: "Accumulated", "Allowance", "Provision", "Reserve"'
            ],
            'impact': '⚠️ **MEDIUM-HIGH IMPACT**: Incorrect signs will misrepresent account positions in financial statements. Assets may show as liabilities and vice versa. Should be fixed for accurate reporting.',
            'violation_count': violation_count
        }
    
    def _generate_equation_insights(self, failure_data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate insights for Rule 7: Accounting Equation"""
        details = failure_data.get('details', {})
        violations = failure_data.get('violations', [])
        
        # Parse equation sum from Details field
        details_text = details.get('Details', '')
        equation_sum = 0
        capital_accounts = 0
        empty_capital = 0
        
        import re
        # Extract equation sum
        sum_match = re.search(r'Equation Sum:\s*₹?([-\d,\.]+)', details_text)
        if sum_match:
            equation_sum = float(sum_match.group(1).replace(',', ''))
        
        # Extract capital account info
        capital_match = re.search(r'Capital Accounts:\s*(\d+)\s*\((\d+)\s*empty\)', details_text)
        if capital_match:
            capital_accounts = int(capital_match.group(1))
            empty_capital = int(capital_match.group(2))
        
        # Get sample empty capital accounts
        empty_capital_accounts = []
        for v in violations[:5]:
            gl_code = v.get('GL_Code', 'N/A')
            desc = v.get('Description', 'No description')
            balance = v.get('Balance', 0)
            empty_capital_accounts.append(f"  • {gl_code}: {desc} - Balance: ₹{balance:,.2f}")
        
        return {
            'summary': f'Accounting equation validation failed. Found {empty_capital} empty capital account(s) out of {capital_accounts} total. Equation sum: ₹{equation_sum:,.2f} (should be ₹0.00)',
            'root_causes': [
                '🔍 **Incomplete P&L Close**: Profit/Loss not transferred to retained earnings',
                '🔍 **Missing Capital Accounts**: Equity accounts (32xxxxx) not properly set up or populated',
                '🔍 **Share-Based Payments**: Share-based payment reserves may be empty if no such transactions',
                '🔍 **Period Mismatch**: Balance sheet items from different accounting periods',
                '🔍 **New Company**: If new company, some equity accounts may legitimately be zero',
                '🔍 **Classification Errors**: Accounts in wrong category affecting equation balance'
            ],
            'recommendations': [
                '✅ **Check Empty Accounts**: Review the empty capital account(s) - are they expected to be zero?',
                f'✅ **Review {empty_capital} Empty Capital Account(s)**: Determine if these should have balances',
                '✅ **Verify P&L Close**: Ensure current year profit/loss is in retained earnings (32xxxxx)',
                '✅ **Check Share Capital**: Verify share capital and reserves are properly recorded',
                '✅ **Reconcile Equation**: Assets = Liabilities + Equity (including current year profit)',
                '✅ **Professional Review**: Complex equity issues may require CFO/Controller input'
            ],
            'examples': (
                [
                    f'📊 Capital Account Analysis:',
                    f'  • Total Capital Accounts: {capital_accounts}',
                    f'  • Empty Capital Accounts: {empty_capital}',
                    f'  • Equation Imbalance: ₹{equation_sum:,.2f}',
                    '',
                    '🔍 Empty Capital Accounts:',
                ] + (empty_capital_accounts if empty_capital_accounts else ['  None listed']) + [
                    '',
                    '💡 Common Scenarios:',
                    '  1. Share-based payment reserve (32100000): Often zero if no ESOP/stock options',
                    '  2. Check if profit for the year has been moved to retained earnings',
                    '  3. For new companies, some equity accounts may be legitimately zero',
                    '',
                    '⚠️ Expected Capital Account Structure (32xxxxx):',
                    '  - Share Capital (32010000): Should have balance (founders\' investment)',
                    '  - Retained Earnings (32020000): Should include prior years\' P&L',
                    '  - Current Year Profit/Loss: Should balance the equation',
                    '',
                    f'🎯 Action Required: Investigate {empty_capital} empty account(s) and verify if they need balances'
                ]
            ),
            'impact': '🚨 **CRITICAL IMPACT**: The accounting equation must balance. Empty capital accounts or equation imbalance indicates fundamental issues with equity structure or P&L closing process.',
            'violation_count': len(violations)
        }
    
    def generate_insights_report(self) -> bool:
        """
        Generate comprehensive insights report in Excel format
        
        Returns:
            bool: True if report generated successfully
        """
        try:
            # Analyze failures
            print("\n" + "="*80)
            print("AI VALIDATION INSIGHTS GENERATION")
            print("="*80)
            
            failures = self.analyze_rule_failures()
            
            if not failures:
                print("✓ No validation failures found - all rules passed!")
                # Set a placeholder path even when there are no failures
                output_dir = self.path_service.get_adjusted_tb_dir(self.entity)
                output_dir.mkdir(parents=True, exist_ok=True)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                self.insights_report_path = output_dir / f"validation_insights_report_{timestamp}.xlsx"
                return True
            
            print(f"\nAnalyzing {len(failures)} failed validation rules...")
            
            # Generate insights for each failure
            for rule_name, failure_data in failures.items():
                print(f"\n  Generating insights for: {rule_name}")
                self.insights[rule_name] = self.generate_ai_insights(rule_name, failure_data)
            
            # Create output path
            output_dir = self.path_service.get_adjusted_tb_dir(self.entity)
            output_dir.mkdir(parents=True, exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.insights_report_path = output_dir / f"validation_insights_report_{timestamp}.xlsx"
            
            # Create Excel report
            self._create_excel_report()
            
            print(f"\n✓ Insights report generated: {self.insights_report_path}")
            return True
            
        except Exception as e:
            print(f"❌ Error generating insights report: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _create_excel_report(self):
        """Create formatted Excel report with insights"""
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # 1. Executive Summary Sheet
        self._create_executive_summary_sheet(wb)
        
        # 2. Detailed Insights Sheet (one for each failed rule)
        for rule_name, insights in self.insights.items():
            self._create_rule_insights_sheet(wb, rule_name, insights)
        
        # 3. Action Plan Sheet
        self._create_action_plan_sheet(wb)
        
        # Save workbook
        wb.save(self.insights_report_path)
    
    def _create_executive_summary_sheet(self, wb: Workbook):
        """Create executive summary sheet"""
        ws = wb.create_sheet("Executive Summary", 0)
        
        # Title
        ws['A1'] = "VALIDATION INSIGHTS - EXECUTIVE SUMMARY"
        ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="8B0010", end_color="8B0010", fill_type="solid")
        ws.merge_cells('A1:E1')
        
        # Metadata
        ws['A3'] = "Entity:"
        ws['B3'] = self.entity.upper()
        ws['A4'] = "Report Generated:"
        ws['B4'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws['A5'] = "Total Failed Rules:"
        ws['B5'] = len(self.insights)
        
        # Summary table
        ws['A7'] = "Failed Rule"
        ws['B7'] = "Severity"
        ws['C7'] = "Violation Count"
        ws['D7'] = "Summary"
        ws['E7'] = "Impact"
        
        # Header formatting
        for col in ['A7', 'B7', 'C7', 'D7', 'E7']:
            ws[col].font = Font(bold=True, color="FFFFFF")
            ws[col].fill = PatternFill(start_color="8B0010", end_color="8B0010", fill_type="solid")
        
        # Add data
        row = 8
        for rule_name, insights in self.insights.items():
            ws[f'A{row}'] = rule_name
            ws[f'B{row}'] = insights['severity']
            ws[f'C{row}'] = insights.get('violation_count', 'N/A')
            ws[f'D{row}'] = insights['summary'][:100] + "..." if len(insights['summary']) > 100 else insights['summary']
            ws[f'E{row}'] = insights['impact'][:50] + "..." if len(insights['impact']) > 50 else insights['impact']
            
            # Color code by severity
            severity_color = {
                'CRITICAL': 'FFC7CE',  # Red
                'HIGH': 'FFEB9C',      # Yellow
                'MEDIUM': 'C6EFCE',    # Light Green
                'LOW': 'DDEBF7'        # Blue
            }
            color = severity_color.get(insights['severity'], 'FFFFFF')
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 40
    
    def _create_rule_insights_sheet(self, wb: Workbook, rule_name: str, insights: Dict[str, Any]):
        """Create detailed insights sheet for a specific rule"""
        # Sanitize sheet name (Excel has 31 char limit)
        sheet_name = rule_name[:31].replace(':', '-').replace('/', '-')
        ws = wb.create_sheet(sheet_name)
        
        # Title
        ws['A1'] = rule_name
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="8B0010", end_color="8B0010", fill_type="solid")
        ws.merge_cells('A1:D1')
        
        # Severity
        ws['A3'] = "Severity:"
        ws['B3'] = insights['severity']
        ws['B3'].font = Font(bold=True)
        
        # Summary
        ws['A5'] = "Summary:"
        ws['A5'].font = Font(bold=True, size=12)
        ws['A6'] = insights['summary']
        ws.merge_cells('A6:D6')
        ws['A6'].alignment = Alignment(wrap_text=True)
        
        # Root Causes
        ws['A8'] = "Root Causes:"
        ws['A8'].font = Font(bold=True, size=12)
        row = 9
        for cause in insights['root_causes']:
            ws[f'A{row}'] = cause
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            row += 1
        
        # Recommendations
        row += 1
        ws[f'A{row}'] = "Recommendations:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        for rec in insights['recommendations']:
            ws[f'A{row}'] = rec
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            ws[f'A{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            row += 1
        
        # Examples
        row += 1
        ws[f'A{row}'] = "Examples & Patterns:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        for example in insights['examples']:
            ws[f'A{row}'] = example
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            row += 1
        
        # Impact
        row += 1
        ws[f'A{row}'] = "Business Impact:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        ws[f'A{row}'] = insights['impact']
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws[f'A{row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 80
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
    
    def _create_action_plan_sheet(self, wb: Workbook):
        """Create prioritized action plan sheet"""
        ws = wb.create_sheet("Action Plan")
        
        # Title
        ws['A1'] = "PRIORITIZED ACTION PLAN"
        ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="8B0010", end_color="8B0010", fill_type="solid")
        ws.merge_cells('A1:E1')
        
        # Instructions
        ws['A3'] = "Follow this prioritized action plan to resolve validation failures:"
        ws.merge_cells('A3:E3')
        ws['A3'].font = Font(italic=True)
        
        # Headers
        ws['A5'] = "Priority"
        ws['B5'] = "Rule"
        ws['C5'] = "Action Required"
        ws['D5'] = "Estimated Effort"
        ws['E5'] = "Owner/Responsible"
        
        for col in ['A5', 'B5', 'C5', 'D5', 'E5']:
            ws[col].font = Font(bold=True, color="FFFFFF")
            ws[col].fill = PatternFill(start_color="8B0010", end_color="8B0010", fill_type="solid")
        
        # Sort insights by severity
        severity_order = {'CRITICAL': 1, 'HIGH': 2, 'MEDIUM': 3, 'LOW': 4}
        sorted_insights = sorted(
            self.insights.items(),
            key=lambda x: severity_order.get(x[1]['severity'], 5)
        )
        
        row = 6
        priority = 1
        for rule_name, insights in sorted_insights:
            ws[f'A{row}'] = priority
            ws[f'B{row}'] = rule_name
            
            # First recommendation as action
            action = insights['recommendations'][0] if insights['recommendations'] else "Review and resolve"
            ws[f'C{row}'] = action
            ws[f'C{row}'].alignment = Alignment(wrap_text=True)
            
            # Estimate effort based on severity
            effort_map = {
                'CRITICAL': 'High (2-4 hours)',
                'HIGH': 'Medium (1-2 hours)',
                'MEDIUM': 'Low (30-60 min)',
                'LOW': 'Minimal (< 30 min)'
            }
            ws[f'D{row}'] = effort_map.get(insights['severity'], 'TBD')
            
            ws[f'E{row}'] = ""  # To be filled by user
            
            # Color code by severity
            severity_color = {
                'CRITICAL': 'FFC7CE',
                'HIGH': 'FFEB9C',
                'MEDIUM': 'C6EFCE',
                'LOW': 'DDEBF7'
            }
            color = severity_color.get(insights['severity'], 'FFFFFF')
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            
            row += 1
            priority += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 25
    
    def run(self) -> Tuple[bool, str, str]:
        """
        Main execution flow
        
        Returns:
            Tuple of (success: bool, message: str, report_path: str)
        """
        try:
            # Load validation report
            try:
                if not self.load_validation_report():
                    return False, "Failed to load validation report", None
            except FileNotFoundError as e:
                return False, str(e), None
            
            # Generate insights
            if not self.generate_insights_report():
                return False, "Failed to generate insights report", None
            
            # Check if there were any failures
            if not self.insights:
                success_msg = f"""
✅ All Validation Rules Passed!

📊 Analysis Complete:
   • Failed Rules: 0
   • All 7 validation rules have passed successfully
   
🎯 Next Steps:
   1. Proceed to the next stage of processing
   2. No corrections needed
"""
            else:
                success_msg = f"""
✅ AI Validation Insights Generated Successfully!

📊 Analysis Complete:
   • Failed Rules Analyzed: {len(self.insights)}
   • Insights Report: {self.insights_report_path}
   
🎯 Next Steps:
   1. Download and review the insights report
   2. Follow the prioritized action plan
   3. Fix critical issues first
   4. Re-run validation after corrections
"""
            print(success_msg)
            return True, success_msg, str(self.insights_report_path) if self.insights_report_path else None
            
        except Exception as e:
            error_msg = f"Error generating insights: {str(e)}"
            print(f"❌ {error_msg}")
            import traceback
            traceback.print_exc()
            return False, error_msg, None


def generate_validation_insights(entity: str = "cpm") -> Tuple[bool, str, str]:
    """
    Convenience function to generate validation insights
    
    Args:
        entity: Entity code
        
    Returns:
        Tuple of (success: bool, message: str, report_path: str)
    """
    generator = AIValidationInsightsGenerator(entity=entity)
    return generator.run()


if __name__ == "__main__":
    import sys
    
    # Get entity from command line or use default
    entity = sys.argv[1] if len(sys.argv) > 1 else "cpm"
    
    print("\n" + "="*80)
    print("AI VALIDATION INSIGHTS GENERATOR")
    print("="*80)
    print(f"Entity: {entity.upper()}")
    print("="*80)
    
    success, message, report_path = generate_validation_insights(entity=entity)
    
    if success:
        print("\n✓ Success!")
        print(f"  Report: {report_path}")
        sys.exit(0)
    else:
        print(f"\n✗ Failed: {message}")
        sys.exit(1)
